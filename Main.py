# [INFO] '24.02.26 Update

# usb upload interface module
from multiprocessing.sharedctypes import Value
from stat import ST_CTIME
from textwrap import fill
from cv2 import line
from Module.database import Database
from Module.SCP import Transfer
from Module.jsonParser import JASONparser
from tqdm import tqdm
import os
os.environ["GIT_PYTHON_REFRESH"] = "quiet"
import sys

import configparser
from os.path import isfile, join
from collections import Counter
from tkinter import END, messagebox
from PIL import Image, ImageSequence, ImageTk
import tkinter.font as tkFont
from tkcalendar import Calendar
from collections import deque
from os.path import getsize
# import keras.backend as K
from queue import Queue
from os import listdir
from _thread import *
import tkinter as tk
from tkinter import ttk
import numpy as np
import threading
from threading import Thread
import datetime

# import binascii
import imutils
import pymysql
import pickle
import serial
import socket
import copy
import time
import cv2
import json

# import re
# import gc
import traceback
import logging
import git
import shutil

#modbus
from pymodbus.client.sync import ModbusTcpClient

from operator import index
from tabnanny import check
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font

# import math as mt
# import pygame

# tf.disable_v2_behavior()

logger = logging.getLogger()
logger.setLevel(logging.INFO)
formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")

if not os.path.exists("log"):
    os.makedirs("log")

nowtime = datetime.datetime.now().strftime("%Y_%m_%d")

file_handler = logging.FileHandler(f"log/log_{nowtime}.log")
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

# config = tf.ConfigProto()
# config.gpu_options.allow_growth = True
# sess = tf.Session(config=config)

# gpus = tf.config.experimental.list_physical_devices('GPU')
# tf.config.experimental.set_memory_growth(gpus[0], True)
# K.clear_session()

root = tk.Tk()

# Hyper parameter
imwrite_switch = True
txtwrite_switch = False

# for test
path = os.path.dirname(os.path.abspath(__file__)).split("\\") 
my_folder = path[-1] # e.g. CONEMAIN_SS11
print('[INFO] NOW PC NAME:', my_folder)

if "CONEMAIN" in my_folder:
    ConnectClientCheck = 6
    MAINHOST = "192.168.0.110"
    ModebusHostIp = '192.168.0.2'
    ModebusHostPort = 502
    LINE = my_folder.split('_')[1] # e.g. S11
    CodeSetup = "CONE"
    BgSetup = "bg_cone"
else:
    ConnectClientCheck = 5
    MAINHOST = "192.168.0.100"
    ModebusHostIp = '192.168.0.2'
    ModebusHostPort = 502
    LINE = my_folder.split('_')[1] # e.g. S11
    CodeSetup = "CUP"
    BgSetup = "bg_cup"

print(f"[INFO] {LINE}, {CodeSetup}")
try:
    os.system("sudo ifmetric enp37s0 30000")
except:
    pass

try:
    os.system("sudo ifmetric enp34s0 100")
except:
    pass


class MainFrame(tk.Frame):
    # 변수 선언
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.main_frame_bad = [None] * 37
        self.main_frame_img = ImageTk.PhotoImage(file=f"bg/{BgSetup}/bg.png")
        self.main_frame_img2 = ImageTk.PhotoImage(file=f"bg/{BgSetup}/bg2.png")
        self.main_frame_ok = ImageTk.PhotoImage(file=f"bg/{BgSetup}/OK.png")
        self.main_frame_ng = ImageTk.PhotoImage(file=f"bg/{BgSetup}/NG.png")
        self.main_frame_connecticon = ImageTk.PhotoImage(file=f"bg/{BgSetup}/connecticon.png")
        self.main_frame_back = ImageTk.PhotoImage(file=f"bg/{BgSetup}/back.png")
        self.main_frame_save = ImageTk.PhotoImage(file=f"bg/{BgSetup}/SAVE.png")
        self.main_frame_ngsetting_btn = ImageTk.PhotoImage(file=f"bg/{BgSetup}/NGSETTING_btn.png")
        self.main_frame_ngsetting_bg = ImageTk.PhotoImage(file=f"bg/{BgSetup}/NGSETTING_bg.png")
        self.main_frame_CameraBlur_img = ImageTk.PhotoImage(file=f"bg/{BgSetup}/CameraBlur.png")
        self.main_frame_adminPassword_img = ImageTk.PhotoImage(file=f"bg/{BgSetup}/admin_Password.png")

        self.disk_error_pop = ImageTk.PhotoImage(file=f"bg/ING/disk_error.png") # OH

        # for i in range(37):
        #     print(i)
        #     self.main_frame_bad[i] = ImageTk.PhotoImage(file = ff'bg/{BgSetup}/bad{i}.png')
        self.main_frame_bad = ImageTk.PhotoImage(file=f"bg/{BgSetup}/badSection.png")
        self.main_frame_listview = ImageTk.PhotoImage(file=f"bg/{BgSetup}/NgCheckListView_inner.png" if CodeSetup == "CONE" else f"bg/{BgSetup}/NgCheckListView_outer.png")
        self.main_frame_modelloading = ImageTk.PhotoImage(file=f"bg/{BgSetup}/Loading.png")

        # 초기 상태 표시 INI 가져오기
        self.init_state_config = configparser.ConfigParser()
        self.initstateINI = 'CountSave/State.INI'

        if os.path.isfile(self.initstateINI) == False: 
            self.init_state_config['STATE'] = {'index' : '0'}
            self.init_state_config['BYPASS'] = {'bypass' : 'off'}
            
            # INI 파일 생성 
            with open(self.initstateINI, 'w') as f:
                self.init_state_config.write(f)

        self.init_state_config.read('CountSave/State.INI', encoding='euc-kr')
        self.state = self.init_state_config['STATE']['index']
        print(f'▶ 현재 상태 : {self.state}')
        
        # 변수 선언
        self.main_cam = [None] * 6
        self.main_ok = [None] * 6
        self.main_ng = [None] * 6
        self.connectionCheck = [None] * 6
        self.updateImage = [None] * 6
        self.bad_image1 = [None] * 10
        self.bad_image2 = [None] * 10
        self.bad_text1 = [None] * 10
        self.bad_text1_inputData = [None] * 10
        self.bad_text2 = [None] * 10
        self.bad_text2_inputData = [None] * 10
        self.CameraBlur = [None] * 6

        self.total_count = 0
        self.ok_count = 0
        self.ng_count = 0
        self.capture_mode = self.setting_mode = self.view_mode = self.management_mode =  False
        self.bypass_mode = False
        self.ngCheckMode = False
        self.ngSettingMode = False
        self.NgCaptureCheck = False

        #entry mode
        self.PasswordMode = False  # 패스워드 화면이 켜져있는지 아닌지 확인하는 변수
        self.PasswordCheck = False
        self.Password = 'art105'
        ####################################
        self.test_mode = None
        ####################################

        self.mainbadCheckList = [[0] * 25, [0] * 25, [0] * 25, [0] * 25, [0] * 25, [0] * 25] if CodeSetup == "CONE" else [[0] * 25, [0] * 25, [0] * 25, [0] * 25, [0] * 25]
        self.mainbadCheckListboxData = [[], [], [], [], [], []] if CodeSetup == "CONE" else [[], [], [], [], []]
        self.mainbadCheckListTotalcount = [0, 0, 0, 0, 0, 0] if CodeSetup == "CONE" else [0, 0, 0, 0, 0]

        # window box
        self.labelHeightNum = [None] * 6
        self.scaleframe = [None] * 6
        self.scaleframeShow = [None] * 6
        self.framecanvas = [None] * 6
        self.framecanvasScale = [None] * 6
        self.filelistboxscrollbarH = [None] * 6
        self.classLabel = [None] * 6
        self.scaleBar = [None] * 6
        self.scaleBar_count1 = [None] * 6
        self.scaleBar_count2 = [None] * 6
        self.framecanvasScaleShow = [None] * 6
        
        # ng 연속검출 세팅 인터페이스 필요 변수
        self.ngSettingBg = [None] * 6
        self.entryBox = [None] * 12
        self.entryBoxView = [None] * 12
        self.ngSettingList = [[20, 10, False],[20, 10, False],[20, 10, False],[20, 10, False],[20, 10, False],[20, 10, False]]
        self.ngSettingList_backup = [[20, 10, False],[20, 10, False],[20, 10, False],[20, 10, False],[20, 10, False],[20, 10, False]]
        self.ngSettingText = [None] * 6

        # loading gif moduls
        self.loadingFrame = []
        im = Image.open(f"bg/{BgSetup}/loading_main.gif")
        index = 0
        for frame in ImageSequence.Iterator(im):
            # np_frame = np.asarray(frame)
            frame = frame.convert('RGBA')
            frame = frame.resize((150, 150), Image.ANTIALIAS)
            frame_tk = ImageTk.PhotoImage(frame)
            
            self.loadingFrame.append(frame_tk)
            index += 1
        self.loadingProcessTrigger = None

        # 카메라 연결중 팝업 이미지
        self.popup_image = [None] * 2
        self.popup_image[0] = ImageTk.PhotoImage(file=f"bg/point_{CodeSetup}_1.png")
        self.popup_image[1] = ImageTk.PhotoImage(file=f"bg/point_{CodeSetup}_2.png")

        self.ng_check_click = False

        # 화면 세팅
        self.master.attributes("-fullscreen", True)
        self.master.bind(
            "<F11>", lambda event: self.master.attributes("-fullscreen", not self.master.attributes("-fullscreen")),
        )
        self.master.bind("<Escape>", lambda event: self.master.attributes("-fullscreen", False))
        self.create_widgets()

    def create_widgets(self):
        self.grid(row=0, column=0)
        self.main_canvas = tk.Canvas(self, width=1920, height=1080)
        # image
        # [topMain, topclient, sideclient1, sideclient2, sideclient3, sideclient4]
        self.background_image = self.main_canvas.create_image(0, 0, image=self.main_frame_img, anchor="nw")
        self.stateText = self.main_canvas.create_text(1122, 1008, text="상태 미설정", font=("gothic", 20, "bold"), fill="white", anchor="center",)
        self.main_cam[0] = self.main_canvas.create_image(36, 548, image="", anchor="nw", state="normal")
        self.main_cam[1] = self.main_canvas.create_image(36, 131, image="", anchor="nw", state="normal")
        self.main_cam[2] = self.main_canvas.create_image(484, 131, image="", anchor="nw", state="normal")
        if CodeSetup == "CONE":
            self.main_cam[3] = self.main_canvas.create_image(932, 131, image="", anchor="nw", state="normal")
            self.main_cam[4] = self.main_canvas.create_image(484, 548, image="", anchor="nw", state="normal")
            self.main_cam[5] = self.main_canvas.create_image(932, 548, image="", anchor="nw", state="normal")
        else:
            self.main_cam[3] = self.main_canvas.create_image(484, 548, image="", anchor="nw", state="normal")
            self.main_cam[4] = self.main_canvas.create_image(932, 548, image="", anchor="nw", state="normal")
        self.main_ok[0] = self.main_canvas.create_image(31, 856, image=self.main_frame_ok, anchor="nw", state="hidden")
        self.main_ok[1] = self.main_canvas.create_image(31, 438, image=self.main_frame_ok, anchor="nw", state="hidden")
        self.main_ok[2] = self.main_canvas.create_image(479, 438, image=self.main_frame_ok, anchor="nw", state="hidden")
        if CodeSetup == "CONE":
            self.main_ok[3] = self.main_canvas.create_image(927, 438, image=self.main_frame_ok, anchor="nw", state="hidden")
            self.main_ok[4] = self.main_canvas.create_image(479, 856, image=self.main_frame_ok, anchor="nw", state="hidden")
            self.main_ok[5] = self.main_canvas.create_image(927, 856, image=self.main_frame_ok, anchor="nw", state="hidden")
        else:
            self.main_ok[3] = self.main_canvas.create_image(479, 856, image=self.main_frame_ok, anchor="nw", state="hidden")
            self.main_ok[4] = self.main_canvas.create_image(927, 856, image=self.main_frame_ok, anchor="nw", state="hidden")
        self.main_ng[0] = self.main_canvas.create_image(251, 856, image=self.main_frame_ng, anchor="nw", state="hidden")
        self.main_ng[1] = self.main_canvas.create_image(251, 438, image=self.main_frame_ng, anchor="nw", state="hidden")
        self.main_ng[2] = self.main_canvas.create_image(699, 438, image=self.main_frame_ng, anchor="nw", state="hidden")
        if CodeSetup == "CONE":
            self.main_ng[3] = self.main_canvas.create_image(1147, 438, image=self.main_frame_ng, anchor="nw", state="hidden")
            self.main_ng[4] = self.main_canvas.create_image(699, 856, image=self.main_frame_ng, anchor="nw", state="hidden")
            self.main_ng[5] = self.main_canvas.create_image(1147, 856, image=self.main_frame_ng, anchor="nw", state="hidden")
        else:
            self.main_ng[3] = self.main_canvas.create_image(699, 856, image=self.main_frame_ng, anchor="nw", state="hidden")
            self.main_ng[4] = self.main_canvas.create_image(1147, 856, image=self.main_frame_ng, anchor="nw", state="hidden")
            pass

        self.CameraBlur[0] = self.main_canvas.create_image(31, 856, image=self.main_frame_CameraBlur_img, anchor="nw", state="hidden")
        self.CameraBlur[1] = self.main_canvas.create_image(31, 438, image=self.main_frame_CameraBlur_img, anchor="nw", state="hidden")
        self.CameraBlur[2] = self.main_canvas.create_image(479, 438, image=self.main_frame_CameraBlur_img, anchor="nw", state="hidden")
        if CodeSetup == "CONE":
            self.CameraBlur[3] = self.main_canvas.create_image(927, 438, image=self.main_frame_CameraBlur_img, anchor="nw", state="hidden")
            self.CameraBlur[4] = self.main_canvas.create_image(479, 856, image=self.main_frame_CameraBlur_img, anchor="nw", state="hidden")
            self.CameraBlur[5] = self.main_canvas.create_image(927, 856, image=self.main_frame_CameraBlur_img, anchor="nw", state="hidden")
        else:
            self.CameraBlur[3] = self.main_canvas.create_image(479, 856, image=self.main_frame_CameraBlur_img, anchor="nw", state="hidden")
            self.CameraBlur[4] = self.main_canvas.create_image(927, 856, image=self.main_frame_CameraBlur_img, anchor="nw", state="hidden")
        self.connectionCheck[0] = self.main_canvas.create_image(127, 74, image=self.main_frame_connecticon, anchor="nw", state="hidden")
        self.connectionCheck[1] = self.main_canvas.create_image(127, 30, image=self.main_frame_connecticon, anchor="nw", state="hidden")
        self.connectionCheck[2] = self.main_canvas.create_image(307, 30, image=self.main_frame_connecticon, anchor="nw", state="hidden")
        if CodeSetup == "CONE":
            self.connectionCheck[3] = self.main_canvas.create_image(487, 30, image=self.main_frame_connecticon, anchor="nw", state="hidden")
            self.connectionCheck[4] = self.main_canvas.create_image(307, 74, image=self.main_frame_connecticon, anchor="nw", state="hidden")
            self.connectionCheck[5] = self.main_canvas.create_image(487, 74, image=self.main_frame_connecticon, anchor="nw", state="hidden")
        else:
            self.connectionCheck[3] = self.main_canvas.create_image(307, 74, image=self.main_frame_connecticon, anchor="nw", state="hidden")
            self.connectionCheck[4] = self.main_canvas.create_image(487, 74, image=self.main_frame_connecticon, anchor="nw", state="hidden")
        self.back_image = self.main_canvas.create_image(1146, 973, image=self.main_frame_back, anchor="nw", state="hidden")
        self.save_image = self.main_canvas.create_image(927, 973, image=self.main_frame_save, anchor="nw", state="hidden")
        check = 0
        for i in range(9, -1, -1):
            self.bad_image1[i] = self.main_canvas.create_image(1376, 995 - check, image="", anchor="nw", state="normal")
            self.bad_text1[i] = self.main_canvas.create_text(1498, 1020 - check, text="", font=("gothic", 18, "bold"), fill="white", anchor="center",)
            check += 57
        check = 0
        for i in range(9, -1, -1):
            self.bad_image2[i] = self.main_canvas.create_image(1636, 995 - check, image="", anchor="nw", state="normal")
            self.bad_text2[i] = self.main_canvas.create_text(1755, 1021 - check, text="", font=("gothic", 18, "bold"), fill="white", anchor="center",)
            check += 57

        self.ngSettingBtn = self.main_canvas.create_image(479, 972, image=self.main_frame_ngsetting_btn, anchor="nw", state="hidden")
        # self.ngSettingBg = self.main_canvas.create_image(1376, 424, image=self.main_frame_ngsetting_bg, anchor="nw", state="hidden")
        self.ngSettingBg[0] = self.main_canvas.create_image(31, 543, image=self.main_frame_ngsetting_bg, anchor = "nw", state = "hidden")
        self.ngSettingBg[1] = self.main_canvas.create_image(31, 125, image=self.main_frame_ngsetting_bg, anchor = "nw", state = "hidden")
        self.ngSettingBg[2] = self.main_canvas.create_image(479, 125, image=self.main_frame_ngsetting_bg, anchor = "nw", state = "hidden")
        self.ngSettingBg[3] = self.main_canvas.create_image(927, 125, image=self.main_frame_ngsetting_bg, anchor = "nw", state = "hidden")
        self.ngSettingBg[4] = self.main_canvas.create_image(479, 543, image=self.main_frame_ngsetting_bg, anchor = "nw", state = "hidden")
        if CodeSetup == "CONE":
            self.ngSettingBg[5] = self.main_canvas.create_image(927, 543, image=self.main_frame_ngsetting_bg, anchor = "nw", state = "hidden")

        # 팝업창 OH
        self.show_pop = self.main_canvas.create_image(678, 972, image='', anchor="nw", state="hidden")

        # 팝업창
        self.popup_place = self.main_canvas.create_image(23, 23, anchor="nw", state="hidden")

        def _from_rgb(rgb):
            """translates an rgb tuple of int to a tkinter friendly color code
            """
            r, g, b = rgb
            return f'#{r:02x}{g:02x}{b:02x}'

        self.entryBox[0] = tk.Entry(self, bg = _from_rgb((47,78,108)), font=(None, 25, 'bold'), fg='white',width=15, justify='center')
        self.entryBoxView[0] = self.main_canvas.create_window(95, 647, anchor='nw', window='')
        self.entryBox[1] = tk.Entry(self, bg = _from_rgb((47,78,108)), font=(None, 25, 'bold'), fg='white', width=15, justify='center')
        self.entryBoxView[1] = self.main_canvas.create_window(95, 763, anchor='nw', window='')
        self.entryBox[2] = tk.Entry(self, bg = _from_rgb((47,78,108)), font=(None, 25, 'bold'), fg='white', width=15, justify='center')
        self.entryBoxView[2] = self.main_canvas.create_window(95, 229, anchor='nw', window='')
        self.entryBox[3] = tk.Entry(self, bg = _from_rgb((47,78,108)), font=(None, 25, 'bold'), fg='white', width=15, justify='center')
        self.entryBoxView[3] = self.main_canvas.create_window(95, 345, anchor='nw', window='')
        self.entryBox[4] = tk.Entry(self, bg = _from_rgb((47,78,108)), font=(None, 25, 'bold'), fg='white', width=15, justify='center')
        self.entryBoxView[4] = self.main_canvas.create_window(543, 229, anchor='nw', window='')
        self.entryBox[5] = tk.Entry(self, bg = _from_rgb((47,78,108)), font=(None, 25, 'bold'), fg='white', width=15, justify='center')
        self.entryBoxView[5] = self.main_canvas.create_window(543, 345, anchor='nw', window='')
        self.entryBox[6] = tk.Entry(self, bg = _from_rgb((47,78,108)), font=(None, 25, 'bold'), fg='white', width=15, justify='center')
        self.entryBoxView[6] = self.main_canvas.create_window(991, 229, anchor='nw', window='')
        self.entryBox[7] = tk.Entry(self, bg = _from_rgb((47,78,108)), font=(None, 25, 'bold'), fg='white', width=15, justify='center')
        self.entryBoxView[7] = self.main_canvas.create_window(991, 345, anchor='nw', window='')
        self.entryBox[8] = tk.Entry(self, bg = _from_rgb((47,78,108)), font=(None, 25, 'bold'), fg='white', width=15, justify='center')
        self.entryBoxView[8] = self.main_canvas.create_window(543, 647, anchor='nw', window='')
        self.entryBox[9] = tk.Entry(self, bg = _from_rgb((47,78,108)), font=(None, 25, 'bold'), fg='white', width=15, justify='center')
        self.entryBoxView[9] = self.main_canvas.create_window(543, 763, anchor='nw', window='')
        if CodeSetup == "CONE":
            self.entryBox[10] = tk.Entry(self, bg = _from_rgb((47,78,108)), font=(None, 25, 'bold'), fg='white', width=15, justify='center')
            self.entryBoxView[10] = self.main_canvas.create_window(991, 647, anchor='nw', window='')
            self.entryBox[11] = tk.Entry(self, bg = _from_rgb((47,78,108)), font=(None, 25, 'bold'), fg='white', width=15, justify='center')
            self.entryBoxView[11] = self.main_canvas.create_window(991, 763, anchor='nw', window='')

        self.ngCheckListView = self.main_canvas.create_image(30, 124, image=self.main_frame_listview, anchor="nw", state="hidden")

        # text
        self.totalText = self.main_canvas.create_text(1628, 272, text=self.total_count, font=("gothic", 20, "bold"), fill="white", anchor="center",)
        self.okText = self.main_canvas.create_text(1498, 379, text=self.ok_count, font=("gothic", 20, "bold"), fill="white", anchor="center",)
        self.ngText = self.main_canvas.create_text(1757, 379, text=self.ng_count, font=("gothic", 20, "bold"), fill="white", anchor="center",)
        self.productName = self.main_canvas.create_text(1624, 154, text="None Select Product", font=("gothic", 20, "bold"), fill="white", anchor="center",)

        self.ngSettingText[0] = self.main_canvas.create_text(233, 897, text="비활성화", font=("gothic", 25, "bold"), fill="white", anchor="center", state = 'hidden')
        self.ngSettingText[1] = self.main_canvas.create_text(233, 479, text="비활성화", font=("gothic", 25, "bold"), fill="white", anchor="center", state = 'hidden')
        self.ngSettingText[2] = self.main_canvas.create_text(685, 479, text="비활성화", font=("gothic", 25, "bold"), fill="white", anchor="center", state = 'hidden')
        self.ngSettingText[3] = self.main_canvas.create_text(1130, 479, text="비활성화", font=("gothic", 25, "bold"), fill="white", anchor="center", state = 'hidden')
        self.ngSettingText[4] = self.main_canvas.create_text(685, 897, text="비활성화", font=("gothic", 25, "bold"), fill="white", anchor="center", state = 'hidden')
        self.ngSettingText[5] = self.main_canvas.create_text(1130, 897, text="비활성화", font=("gothic", 25, "bold"), fill="white", anchor="center", state = 'hidden')

        # GPU 텍스트 2022-12-09 OH
        self.gputext1 = self.main_canvas.create_text(363, 152, text = "GPU check", font=("gothic", 15, "bold"), fill='red', anchor='center', state='hidden')
        self.gputext2 = self.main_canvas.create_text(811, 152, text = "GPU check", font=("gothic", 15, "bold"), fill='red', anchor='center', state='hidden')
        self.gputext3 = self.main_canvas.create_text(1259, 152, text = "GPU check", font=("gothic", 15, "bold"), fill='red', anchor='center', state='hidden')
        self.gputext4 = self.main_canvas.create_text(363, 567, text = "GPU check", font=("gothic", 15, "bold"), fill='red', anchor='center', state='hidden')
        self.gputext5 = self.main_canvas.create_text(811, 567, text = "GPU check", font=("gothic", 15, "bold"), fill='red', anchor='center', state='hidden')
        self.gputext6 = self.main_canvas.create_text(1259, 567, text = "GPU check", font=("gothic", 15, "bold"), fill='red', anchor='center', state='hidden')

        # listbox
        color = "#%02x%02X%02x" % (19, 47, 60)
        self.mainframe_log = tk.Listbox(self.main_canvas, fg="white", bg=color, font=("gothic", 12, "bold"))
        self.mainframe_log.place(x=2000, y=2000, width=428, height=539)
        self.mainframe_scrollbar = tk.Scrollbar(self.main_canvas, bg=color, orient="vertical")
        self.mainframe_scrollbar.config(command=self.mainframe_log.yview)
        self.mainframe_scrollbar.place(x=2000, y=2000, width=20, height=539)
        self.mainframe_log.config(yscrollcommand=self.mainframe_scrollbar.set)

        self.resultNgViewListbox_main = tk.Listbox(self.main_canvas, fg="white", bg=color, font=("gothic", 12, "bold"))
        self.resultNgViewListbox_main.place(x=2000, y=2000, width=366, height=343)
        self.resultNgViewListbox_main_scrollbar = tk.Scrollbar(self.main_canvas, bg=color, orient="vertical")
        self.resultNgViewListbox_main_scrollbar.config(command=self.resultNgViewListbox_main.yview)
        self.resultNgViewListbox_main_scrollbar.place(x=2000, y=2000, width=20, height=343)

        self.resultNgViewListbox_topclient = tk.Listbox(self.main_canvas, fg="white", bg=color, font=("gothic", 12, "bold"))
        self.resultNgViewListbox_topclient.place(x=2000, y=2000, width=366, height=343)
        self.resultNgViewListbox_topclient_scrollbar = tk.Scrollbar(self.main_canvas, bg=color, orient="vertical")
        self.resultNgViewListbox_topclient_scrollbar.config(command=self.resultNgViewListbox_topclient.yview)
        self.resultNgViewListbox_topclient_scrollbar.place(x=2000, y=2000, width=20, height=343)

        self.resultNgViewListbox_sideclient1 = tk.Listbox(self.main_canvas, fg="white", bg=color, font=("gothic", 12, "bold"))
        self.resultNgViewListbox_sideclient1.place(x=2000, y=2000, width=366, height=343)
        self.resultNgViewListbox_sideclient1_scrollbar = tk.Scrollbar(self.main_canvas, bg=color, orient="vertical")
        self.resultNgViewListbox_sideclient1_scrollbar.config(command=self.resultNgViewListbox_sideclient1.yview)
        self.resultNgViewListbox_sideclient1_scrollbar.place(x=2000, y=2000, width=20, height=343)

        self.resultNgViewListbox_sideclient2 = tk.Listbox(self.main_canvas, fg="white", bg=color, font=("gothic", 12, "bold"))
        self.resultNgViewListbox_sideclient2.place(x=2000, y=2000, width=366, height=343)
        self.resultNgViewListbox_sideclient2_scrollbar = tk.Scrollbar(self.main_canvas, bg=color, orient="vertical")
        self.resultNgViewListbox_sideclient2_scrollbar.config(command=self.resultNgViewListbox_sideclient2.yview)
        self.resultNgViewListbox_sideclient2_scrollbar.place(x=2000, y=2000, width=20, height=343)

        self.resultNgViewListbox_sideclient3 = tk.Listbox(self.main_canvas, fg="white", bg=color, font=("gothic", 12, "bold"))
        self.resultNgViewListbox_sideclient3.place(x=2000, y=2000, width=366, height=343)
        self.resultNgViewListbox_sideclient3_scrollbar = tk.Scrollbar(self.main_canvas, bg=color, orient="vertical")
        self.resultNgViewListbox_sideclient3_scrollbar.config(command=self.resultNgViewListbox_sideclient3.yview)
        self.resultNgViewListbox_sideclient3_scrollbar.place(x=2000, y=2000, width=20, height=343)

        if CodeSetup == "CONE":
            self.resultNgViewListbox_sideclient4 = tk.Listbox(self.main_canvas, fg="white", bg=color, font=("gothic", 12, "bold"))
            self.resultNgViewListbox_sideclient4.place(x=2000, y=2000, width=366, height=343)
            self.resultNgViewListbox_sideclient4_scrollbar = tk.Scrollbar(self.main_canvas, bg=color, orient="vertical")
            self.resultNgViewListbox_sideclient4_scrollbar.config(command=self.resultNgViewListbox_sideclient4.yview)
            self.resultNgViewListbox_sideclient4_scrollbar.place(x=2000, y=2000, width=20, height=343)
        else:
            pass

        def _from_rgb(rgb):
            """translates an rgb tuple of int to a tkinter friendly color code
            """
            r, g, b = rgb
            return f'#{r:02x}{g:02x}{b:02x}'

        self.AdminPasswordBg = self.main_canvas.create_image(960, 540, image = self.main_frame_adminPassword_img, anchor = 'center', state = 'hidden')
        self.entry = tk.Entry(self, bg = _from_rgb((44,116,132)), font=(None, 25, 'bold'), width=19, justify='center', show = "*")
        self.showEntry = self.main_canvas.create_window(786, 509, anchor='nw', window='')

        self.modelloadingView = self.main_canvas.create_image(960,600, image = self.main_frame_modelloading, anchor = 'center', state = 'hidden')
        self.loadingGifImage = self.main_canvas.create_image(960,700, image = '', anchor = 'center', state = 'hidden')
        
        # 상태 표시 초기 설정
        self.updatestate(self.state)

        # event
        self.entry.bind("<Return>", self.main_password_check)
        self.main_canvas.bind("<Button-1>", self.main_btn)
        self.main_canvas.pack()

    def loadingProcessGIF(self):
        while True:
            for i in range(len(self.loadingFrame)):
                # print('test')
                self.main_canvas.itemconfig(self.loadingGifImage, image=self.loadingFrame[i], state = 'normal')
                time.sleep(0.1)
                if self.loadingProcessTrigger == False:
                    print('complete')
                    self.loadingProcessTrigger = True
                    self.main_canvas.itemconfig(self.loadingGifImage, image='', state = 'hidden')
                    return

    def main_password_check(self, event):
        readPassword = tk.Entry.get(self.entry)
        if readPassword == self.Password:
            self.main_canvas.itemconfig(self.stateText, state = 'hidden')
            self.main_canvas.itemconfig(self.AdminPasswordBg, state = 'hidden')
            self.main_canvas.itemconfig(self.showEntry, window = '')
            self.entry.delete(0,999)
            self.PasswordCheck = True
            self.PasswordMode = False
            print('password Check')

            if self.ngSettingMode == None:
                print("NG SETTINGS Btn Click")
                self.ngSettingLoad("setting")
                self.main_canvas.itemconfig(self.ngSettingBtn, state="normal")
                for i in range(ConnectClientCheck):
                    self.main_canvas.itemconfig(self.ngSettingBg[i], state="normal")
                for i in range(ConnectClientCheck*2):
                    self.main_canvas.itemconfig(self.entryBoxView[i], window = self.entryBox[i])
                self.main_canvas.itemconfig(self.back_image, state="normal")
                self.main_canvas.itemconfig(self.save_image, state="normal")
                self.ngSettingMode = True
            if self.setting_mode == None:
                print("SETTING MODE")
                # if self.setting_mode == False:
                self.setting_mode = True

                if self.setting_mode == True:
                    # DTT.pickleListData[0] = copy.deepcopy(EFFI.ModelSetupList)
                    # self.makeLabelWindow(36, 548, DTT.pickleListData[0], 0)
                    TSD.sendAllClient("SETTING")
                    self.main_canvas.itemconfig(self.back_image, state="normal")
                    self.main_canvas.itemconfig(self.save_image, state="normal")
            if self.management_mode == None:
                self.management_mode = True
                management_frame.tkraise()
        else:
            messagebox.showerror("에러발생", "정확한 패스워드를 기입하여 주세요.")
            self.entry.delete(0,999)

    def ngSettingSave(self):
        self.ngSettingEntryLoad()
        WriteData = self.ngSettingList_backup.copy()
        self.ngSettingList = self.ngSettingList_backup.copy()
        STH.ngContinuousCount = 0
        STH.ngContinuousList = [0,0,0,0,0,0]
        try:
            try:
                if not (os.path.isdir("NGsetting")):
                    os.makedirs(os.path.join("NGsetting"))
            except:
                pass

            with open("NGsetting/NGsettingCount.pickle", "wb") as file:
                pickle.dump(WriteData, file)

            for i in range(6):
                if i == 5 and CodeSetup == 'CUP':
                    break
                self.main_canvas.itemconfig(self.ngSettingText[i], state="hidden")
            main_frame.main_canvas.itemconfig(main_frame.ngSettingBtn, state="hidden")
            for i in range(ConnectClientCheck):
                self.main_canvas.itemconfig(self.ngSettingBg[i], state="hidden")
            for i in range(ConnectClientCheck*2):
                self.main_canvas.itemconfig(self.entryBoxView[i], window = '')
            self.main_canvas.itemconfig(self.back_image, state="hidden")
            self.main_canvas.itemconfig(self.save_image, state="hidden")

            print(f'Saveing Data - {WriteData}')
        except Exception as ex:
            logger.info(f"Warning : NGsetting 피클파일 저장 에러 - {ex} \n {traceback.format_exc()}")

    def ngSettingLoad(self, Mode):
        try:
            # self.vision_value = []
            try:
                with open("NGsetting/NGsettingCount.pickle", "rb") as file:
                    data = pickle.load(file)

                self.ngSettingList_backup = data.copy()
                self.ngSettingList = data.copy()
            except:
                pass
            if Mode == "load":
                pass
            else:
                print(f'Read Data - {self.ngSettingList}')
                for i in range(6):
                    if i == 5 and CodeSetup == 'CUP':
                        break
                    main_frame.main_canvas.itemconfig(main_frame.ngSettingText[i], text="활성화" if self.ngSettingList_backup[i][2] else "비활성화", state="normal")
                    self.ngSettingInterfaceUpdate(i)
                self.main_canvas.itemconfig(self.back_image, state="normal")
                self.main_canvas.itemconfig(self.save_image, state="normal")

            print("피클 파일 로딩 완료")
        except:
            logger.info(f"NGsetting 피클파일이 존재하지 않습니다. {traceback.format_exc()}")
            print("피클파일이 존재하지 않습니다.")
            self.ngSettingSave()
            logger.info(f"NGsetting 피클파일을 Default로 생성합니다. {traceback.format_exc()}")
            print("피클파일을 Default로 생성합니다.")

    def ngSettingEntryLoad(self):
        for i in range(6 if CodeSetup == 'CONE' else 5):
            IdleData = tk.Entry.get(self.entryBox[(i*2)])
            self.ngSettingList_backup[i][0] = IdleData
            IdleData = tk.Entry.get(self.entryBox[(i*2)+1])
            self.ngSettingList_backup[i][1] = IdleData
            pass

    def ngSettingInterfaceUpdate(self, Index):
        self.entryBox[(Index*2)].delete(0,999)
        self.entryBox[(Index*2)+1].delete(0,999)
        self.entryBox[(Index*2)].insert(0, self.ngSettingList[Index][0])
        self.entryBox[(Index*2)+1].insert(0, self.ngSettingList[Index][1])

    def visionNgViewUpdate(self):
        self.mainbadCheckListboxData = [[], [], [], [], [], []] if CodeSetup == "CONE" else [[], [], [], [], []]
        self.mainbadCheckListTotalcount = [0, 0, 0, 0, 0, 0] if CodeSetup == "CONE" else [0, 0, 0, 0, 0]

        self.resultNgViewListbox_main.delete(0, "end")
        self.resultNgViewListbox_topclient.delete(0, "end")
        self.resultNgViewListbox_sideclient1.delete(0, "end")
        self.resultNgViewListbox_sideclient2.delete(0, "end")
        self.resultNgViewListbox_sideclient3.delete(0, "end")
        if CodeSetup == "CONE":
            self.resultNgViewListbox_sideclient4.delete(0, "end")
        else:
            pass
        # self.mainbadCheckList = [[72, 0, 0, 4, 0, 62, 0, 0, 31, 0, 6, 0, 0, 0, 88, 0, 0, 15, 0, 14, 0, 0, 0, 19, 0], [32, 13, 0, 0, 0, 177, 0, 52, 112, 0, 73, 0, 0, 0, 11, 0, 0, 23, 42, 15, 0, 68, 16, 0, 0], [823, 0, 13, 15, 0, 1, 62, 0, 73, 0, 13, 0, 0, 83, 11, 152, 0, 0, 72, 42, 11, 0, 0, 42, 0], [72, 19, 18, 84, 11, 16, 62, 2, 88, 3, 2, 1, 88, 2, 21, 15, 9, 3, 1, 1, 3, 6, 36, 13, 20], [0, 0, 2, 0, 0, 9, 0, 0, 7, 0, 10, 0, 0, 0, 2, 0, 0, 8, 0, 4, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]]

        for i in range(6 if CodeSetup == "CONE" else 5):
            for x in range(len(DTT.badType.keys())):
                self.mainbadCheckListTotalcount[i] += self.mainbadCheckList[i][x]
        print(f'[Check] mainbadCheckListTotalcount - {self.mainbadCheckListTotalcount}')
        # main data trim
        try:
            for i in range(6 if CodeSetup == "CONE" else 5):
                updateCount = 0
                for x in range(len(DTT.badType.keys())):
                    try:
                        if self.mainbadCheckList[i][x] > 0:
                            print(f'{DTT.badType[x]} : {self.mainbadCheckList[i][x]}개 불량확인')
                            updateCount += 1
                            updateCountText = str(updateCount) + ". "
                            updateText = f"{updateCountText}{DTT.badType[x]} : {self.mainbadCheckList[i][x]}개 불량확인"
                            self.mainbadCheckListboxData[i].append(updateText)
                        else:
                            pass
                    except:
                        print(f'{x+1}. DUMMY : {self.mainbadCheckList[i][x]}개 불량확인')
                        pass
        except:
            print("trim Error : ", traceback.format_exc())
        print(f'[Check] mainbadCheckListboxData - {self.mainbadCheckListboxData}')
        try:
            for i in range(6):
                for x in reversed(range(len(self.mainbadCheckListboxData[i]))):
                    if i == 0:
                        # main update
                        self.resultNgViewListbox_main.insert(0, (self.mainbadCheckListboxData[i][x]))
                    elif i == 1:
                        # top client update
                        self.resultNgViewListbox_topclient.insert(0, (self.mainbadCheckListboxData[i][x]))
                    elif i == 2:
                        # side client1 update
                        self.resultNgViewListbox_sideclient1.insert(0, (self.mainbadCheckListboxData[i][x]))
                    elif i == 3:
                        # side client2 update
                        self.resultNgViewListbox_sideclient2.insert(0, (self.mainbadCheckListboxData[i][x]))
                    elif i == 4:
                        # side client3 update
                        self.resultNgViewListbox_sideclient3.insert(0, (self.mainbadCheckListboxData[i][x]))
                    elif i == 5:
                        # side client3 update
                        self.resultNgViewListbox_sideclient4.insert(0, (self.mainbadCheckListboxData[i][x]))
                if i == 0:
                    # main update
                    self.resultNgViewListbox_main.insert(0, ("total Count : " + str(self.mainbadCheckListTotalcount[i])))
                elif i == 1:
                    # top client update
                    self.resultNgViewListbox_topclient.insert(0, ("total Count : " + str(self.mainbadCheckListTotalcount[i])))
                elif i == 2:
                    # side client1 update
                    self.resultNgViewListbox_sideclient1.insert(0, ("total Count : " + str(self.mainbadCheckListTotalcount[i])))
                elif i == 3:
                    # side client2 update
                    self.resultNgViewListbox_sideclient2.insert(0, ("total Count : " + str(self.mainbadCheckListTotalcount[i])))
                elif i == 4:
                    # side client3 update
                    self.resultNgViewListbox_sideclient3.insert(0, ("total Count : " + str(self.mainbadCheckListTotalcount[i])))
                elif i == 5:
                    # side client3 update
                    self.resultNgViewListbox_sideclient4.insert(0, ("total Count : " + str(self.mainbadCheckListTotalcount[i])))
        except:
            print("update Error : ", traceback.format_exc())

    def show_img(self, image, number):
        try:
            image = cv2.resize(image, dsize=(391, 290), interpolation=cv2.INTER_AREA)
        except:
            img = cv2.imread(f"bg/{BgSetup}/NoneData.png")
            image = cv2.resize(img, dsize=(391, 290), interpolation=cv2.INTER_AREA)

        self.cvtimg = cv2.cvtColor(image, cv2.COLOR_BGR2RGBA)
        self.cvtshow = Image.fromarray(self.cvtimg)
        self.updateImage[number] = ImageTk.PhotoImage(image=self.cvtshow)
        self.main_canvas.itemconfig(self.main_cam[number], image=self.updateImage[number])

    def update_signal(self, message):
        self.now = datetime.datetime.now()
        signalTime = self.now.strftime("[%H:%M:%S] ")
        message = signalTime + str(message)
        self.mainframe_log.insert(0, message)

    def makeLabelWindow(self, x, y, dictValue, number):
        print("setting start : ", number)

        if dictValue is not None:
            self.labelHeightNum[number] = int(len(dictValue)) * 250

            self.scaleframe[number] = tk.Frame(self.main_canvas)
            self.scaleframeShow[number] = self.main_canvas.create_window(x, y, anchor="nw", window=self.scaleframe[number], width=389, height=289, state="normal",)

            self.framecanvas[number] = tk.Canvas(self.scaleframe[number], width=439, height=self.labelHeightNum[number])

            self.framecanvasScale[number] = tk.Frame(self.framecanvas[number], width=439, height=self.labelHeightNum[number])
            self.filelistboxscrollbarH[number] = tk.Scrollbar(self.scaleframe[number], orient="vertical", command=self.framecanvas[number].yview)
            self.filelistboxscrollbarH[number].pack(side="right", fill="y")

            self.classLabel[number] = list(range(len(dictValue)))
            self.scaleBar[number] = list(range(len(dictValue)))
            self.scaleBar_count1[number] = list(range(len(dictValue)))
            self.scaleBar_count2[number] = list(range(len(dictValue)))

            fontStyle = tkFont.Font(family="gothic", size=10)

            check = len(dictValue)
            for num, (LableName, LabelValue) in enumerate(dictValue):
                if num == check - 1:
                    print(LableName, "///", LabelValue[0])
                    LableName = str(num + 1) + ". " + LableName
                    LabelName2 = "[COUNT(총)]"

                    self.classLabel[number][num] = tk.Label(self.framecanvasScale[number], text=LableName, font=fontStyle)
                    self.classLabel[number][num].pack()
                    self.classLabel[number][num] = tk.Label(self.framecanvasScale[number], text=LabelName2, font=fontStyle)
                    self.classLabel[number][num].pack()

                    # label size x, y
                    self.scaleBar_count1[number][num] = tk.Scale(self.framecanvasScale[number], from_=0, to=20, orient=tk.HORIZONTAL, length=339, resolution=1, command=lambda v, n=num, n2=number, n3=0: self.editParam(v, n, n2, n3),)
                    self.scaleBar_count1[number][num].set(int(LabelValue[0]))
                    self.scaleBar_count1[number][num].pack(padx=5)

                else:
                    print(LableName, "///", LabelValue[0], "///", LabelValue[1], "///", LabelValue[2])
                    LableName = str(num + 1) + ". " + LableName
                    LabelName2 = "[SCORE(검증값) / COUNT1(연속) / COUNT2(총)]"

                    self.classLabel[number][num] = tk.Label(self.framecanvasScale[number], text=LableName, font=fontStyle)
                    self.classLabel[number][num].pack()
                    self.classLabel[number][num] = tk.Label(self.framecanvasScale[number], text=LabelName2, font=fontStyle)
                    self.classLabel[number][num].pack()
                    self.scaleBar[number][num] = tk.Scale(self.framecanvasScale[number], from_=0, to=100, orient=tk.HORIZONTAL, length=339, resolution=1, command=lambda v, n=num, n2=number, n3=0: self.editParam(v, n, n2, n3),)
                    self.scaleBar[number][num].set(int(LabelValue[0]))
                    self.scaleBar[number][num].pack(padx=5)

                    # label size x, y
                    self.scaleBar_count1[number][num] = tk.Scale(self.framecanvasScale[number], from_=0, to=20, orient=tk.HORIZONTAL, length=339, resolution=1, command=lambda v, n=num, n2=number, n3=1: self.editParam(v, n, n2, n3),)
                    self.scaleBar_count1[number][num].set(int(LabelValue[1]))
                    self.scaleBar_count1[number][num].pack(padx=5)
                    self.scaleBar_count2[number][num] = tk.Scale(self.framecanvasScale[number], from_=0, to=20, orient=tk.HORIZONTAL, length=339, resolution=1, command=lambda v, n=num, n2=number, n3=2: self.editParam(v, n, n2, n3),)
                    self.scaleBar_count2[number][num].set(int(LabelValue[2]))
                    self.scaleBar_count2[number][num].pack(padx=5)

            # self.framecanvas.configure(xscrollcommand=self.filelistboxscrollbarW.set)
            self.framecanvas[number].configure(background="white", yscrollcommand=self.filelistboxscrollbarH[number].set)
            self.framecanvas[number].config(scrollregion=(0, 0, 0, self.labelHeightNum[number]))
            self.framecanvasScaleShow[number] = self.framecanvas[number].create_window(0, 0, anchor="nw", window=self.framecanvasScale[number], width=389, height=self.labelHeightNum[number],)

            self.framecanvas[number].pack()
        else:
            messagebox.showerror("에러발생", "정상적인 접근방법이 아닙니다.")

    def updatestate(self, state):
        if state == '0':
            self.main_canvas.itemconfig(self.stateText, text='상태 미설정')
        elif state == '1' :
            self.main_canvas.itemconfig(self.stateText, text='비전 검사 적용 완료')
        elif state == '2' :
            self.main_canvas.itemconfig(self.stateText, text='비전 검사 미적용')
        elif state == '3' :
            self.main_canvas.itemconfig(self.stateText, text='유효성 평가 대상')

    def editParam(self, value, number, index, check):
        DTT.pickleListData[index][number][1][check] = int(value)

    def saveLabelParam(self):
        filepath = f"models/{STH.nowModel}"

        try:
            if not (os.path.isdir(filepath)):
                os.makedirs(os.path.join(filepath))
        except:
            pass

        with open("{}/InsValue.pickle".format(filepath), "wb") as file:
            pickle.dump(DTT.pickleListData[0], file)

        print("피클파일 저장 완료 : ", filepath)

        # self.settingSavePickle()

        # self.scaleframe[].destroy()

    def OneCycleTest(self):
        try:
            #Start Signal Check
            print(f'[Signal] Modbus Inspection Start Signal Recv_Menual')
            logger.info(f'[Signal] Modbus Inspection Start Signal Recv_Menual')

            STH.StartMode = "11"
            print(f"Start Mode - {STH.StartMode}")
            logger.info(f"[SIGNAL Check] StartMode - {STH.StartMode} _ Menual")

            DTT.AllDataRecv = 0
            DTT.AllImageRecv = 0

            STH.startTime = time.time()

            print("[INFO] START 수신 : ", time.time() - STH.startTime)

            TSD.sendAllClient("START")

            DTT.inspectionSession = True
            print("[INFO] Start 시그널 처리 완료 : ", time.time() - STH.startTime)
            # STH.SendDataTrim("W", "D3003", "0")
            STH.ModbusSignalSend(Mode = "READY", Value = 0)
            STH.ModbusSignalSend(Mode = "BUSY", Value = 1)
            STH.ModbusSignalSend(Mode = "STATE", Value = 1)

            time.sleep(1)

            #ResultRequest Signal Check
            print('[Signal] Modbus ResultRequest ON Signal recv_Menual')
            logger.info('[Signal] Modbus ResultRequest ON Signal recv_Menual')
            main_frame.update_signal("Result Requast recv")
            TSD.sendAllClient("RESULTREQUEST")
            STH.ModbusSignalSend(Mode = "RESULTSESSION", Value = 1)
            print("[INFO] RESULTREQUEST 송신 : ", time.time() - STH.startTime)

            time.sleep(6)

            #Ins Result PLC Check Complete Signal Check
            print('[Signal] Modbus PLC Complete ON Signal recv_Menual')
            logger.info('[Signal] Modbus PLC Complete ON Signal recv_Menual')
            STH.ModbusSignalSend(Mode = "READY", Value = 1)
            STH.ModbusSignalSend(Mode = "BUSY", Value = 0)
            STH.ModbusSignalSend(Mode = "RESULTSESSION", Value = 0)
            STH.ModbusSignalSend(Mode = "RESULT1", Value = 0)
            STH.ModbusSignalSend(Mode = "RESULT2", Value = 0)
            STH.ModbusSignalSend(Mode = "STATE", Value = 0)
            # STH.SendDataTrim("W", "D3006", "0")
            # STH.SendDataTrim("W", "D3007", "0")
            # STH.SendDataTrim("W", "D3003", "1")
        except:
            print(traceback.format_exc())

    def main_btn(self, event):
        # sessionCheck = False
        x = event.x
        y = event.y
        print(x, y)

        if self.PasswordMode == False:
            # if x < 100 and y > 900:
            #     STH.SendDataTrim("W", "D3003", "0")

            # NG SETTINGS 클릭
            if 479 < x < 479 + 180 and 972 < y < 972 + 73 and self.ngSettingMode == False:
                self.ngSettingMode = None
                self.main_canvas.itemconfig(self.AdminPasswordBg, state = 'normal')
                self.main_canvas.itemconfig(self.showEntry, window = self.entry)
                self.PasswordMode = True
                
            if self.ngSettingMode == True:
                #기능 사용 유무
                if 89 < x < 374 and 461 < y < 508:
                    # self.ngSettingList_backup = [[20, 10, False],[20, 10, False],[20, 10, False],[20, 10, False],[20, 10, False],[20, 10, False]]
                    self.ngSettingList_backup[1][2] = not self.ngSettingList_backup[1][2]
                    self.main_canvas.itemconfig(self.ngSettingText[1], text = '활성화' if self.ngSettingList_backup[1][2] == True else '비활성화')

                if 537 < x < 823 and 461 < y < 509:
                    self.ngSettingList_backup[2][2] = not self.ngSettingList_backup[2][2]
                    self.main_canvas.itemconfig(self.ngSettingText[2], text = '활성화' if self.ngSettingList_backup[2][2] == True else '비활성화')

                if 985 < x < 1271 and 461 < y < 509:
                    self.ngSettingList_backup[3][2] = not self.ngSettingList_backup[3][2]
                    self.main_canvas.itemconfig(self.ngSettingText[3], text = '활성화' if self.ngSettingList_backup[3][2] == True else '비활성화')

                if 90 < x < 375 and 880 < y < 927:
                    self.ngSettingList_backup[0][2] = not self.ngSettingList_backup[0][2]
                    self.main_canvas.itemconfig(self.ngSettingText[0], text = '활성화' if self.ngSettingList_backup[0][2] == True else '비활성화')

                if 537 < x < 822 and 880 < y < 927:
                    self.ngSettingList_backup[4][2] = not self.ngSettingList_backup[4][2]
                    self.main_canvas.itemconfig(self.ngSettingText[4], text = '활성화' if self.ngSettingList_backup[4][2] == True else '비활성화')

                if CodeSetup == 'CONE':
                    if 985 < x < 1271 and 880 < y < 927:
                        self.ngSettingList_backup[5][2] = not self.ngSettingList_backup[5][2]
                        self.main_canvas.itemconfig(self.ngSettingText[5], text = '활성화' if self.ngSettingList_backup[5][2] == True else '비활성화')

                # SAVE 클릭
                if 928 < x < 1108 and 973 < y < 1046:
                    print("ngsetting 최종 저장")
                    self.ngSettingSave()
                    self.main_canvas.itemconfig(self.stateText, state = 'normal')
                    DTT.NgContinuityCheck = [[] for _ in range(main_frame.ngSettingText[0])]
                    self.ngSettingMode = False

                # BACK 클릭
                if 1147 < x < 1328 and 973 < y < 1046:
                    print("ngsetting 저장하지 않고 나가기")
                    self.main_canvas.itemconfig(self.stateText, state = 'normal')
                    self.ngSettingMode = False
                    self.ngSettingList_backup = [[20, 10, False],[20, 10, False],[20, 10, False],[20, 10, False],[20, 10, False],[20, 10, False]]
                    for i in range(6):
                        if i == 5 and CodeSetup == 'CUP':
                            break
                        self.main_canvas.itemconfig(self.ngSettingText[i], state="hidden")

                    main_frame.main_canvas.itemconfig(main_frame.ngSettingBtn, state="hidden")
                    for i in range(ConnectClientCheck):
                        self.main_canvas.itemconfig(self.ngSettingBg[i], state="hidden")
                    for i in range(ConnectClientCheck*2):
                        self.main_canvas.itemconfig(self.entryBoxView[i], window = '')
                    self.main_canvas.itemconfig(self.back_image, state="hidden")
                    self.main_canvas.itemconfig(self.save_image, state="hidden")
            
            # NG 클릭 시 NG 카운트 정보 출력되는 이벤트
            if 1636 < x < 1636 + 245 and 308 < y < 308 + 100 and self.ngCheckMode == False:
                self.main_canvas.itemconfig(self.ngCheckListView, state="normal")
                self.visionNgViewUpdate()
                # main
                self.resultNgViewListbox_main.place(x=48, y=587, width=366, height=343)
                self.resultNgViewListbox_main_scrollbar.place(x=48 + 346, y=587, width=20, height=343)
                # topclient
                self.resultNgViewListbox_topclient.place(x=48, y=168, width=366, height=343)
                self.resultNgViewListbox_topclient_scrollbar.place(x=48 + 346, y=168, width=20, height=343)
                # sideclient1
                self.resultNgViewListbox_sideclient1.place(x=494, y=168, width=366, height=343)
                self.resultNgViewListbox_sideclient1_scrollbar.place(x=494 + 346, y=168, width=20, height=343)
                if CodeSetup == "CONE":
                    # sideclient2
                    self.resultNgViewListbox_sideclient2.place(x=943, y=168, width=366, height=343)
                    self.resultNgViewListbox_sideclient2_scrollbar.place(x=943 + 346, y=168, width=20, height=343)
                    # sideclient3
                    self.resultNgViewListbox_sideclient3.place(x=494, y=587, width=366, height=343)
                    self.resultNgViewListbox_sideclient3_scrollbar.place(x=494 + 346, y=587, width=20, height=343)
                    # sideclient4
                    self.resultNgViewListbox_sideclient4.place(x=943, y=587, width=366, height=343)
                    self.resultNgViewListbox_sideclient4_scrollbar.place(x=943 + 346, y=587, width=20, height=343)
                else:
                    # sideclient2
                    self.resultNgViewListbox_sideclient2.place(x=494, y=587, width=366, height=343)
                    self.resultNgViewListbox_sideclient2_scrollbar.place(x=494 + 346, y=168, width=20, height=343)
                    # sideclient3
                    self.resultNgViewListbox_sideclient3.place(x=943, y=587, width=366, height=343)
                    self.resultNgViewListbox_sideclient3_scrollbar.place(x=943 + 346, y=587, width=20, height=343)

                self.ngCheckMode = True

            # NG 클릭 시 NG 카운트 정보 출력되는 이벤트 종료 버튼
            if 1143 < x < 1321 and 966 < y < 1037 and self.ngCheckMode == True:
                self.main_canvas.itemconfig(self.ngCheckListView, state="hidden")
                # main
                self.resultNgViewListbox_main.place(x=2000, y=2000, width=366, height=343)
                self.resultNgViewListbox_main_scrollbar.place(x=2000, y=2000, width=20, height=343)
                # topclient
                self.resultNgViewListbox_topclient.place(x=2000, y=2000, width=366, height=343)
                self.resultNgViewListbox_topclient_scrollbar.place(x=2000, y=2000, width=20, height=343)
                # sideclient1
                self.resultNgViewListbox_sideclient1.place(x=2000, y=2000, width=366, height=343)
                self.resultNgViewListbox_sideclient1_scrollbar.place(x=2000, y=2000, width=20, height=343)
                if CodeSetup == "CONE":
                    # sideclient2
                    self.resultNgViewListbox_sideclient2.place(x=2000, y=2000, width=366, height=343)
                    self.resultNgViewListbox_sideclient2_scrollbar.place(x=2000, y=2000, width=20, height=343)
                    # sideclient3
                    self.resultNgViewListbox_sideclient3.place(x=2000, y=2000, width=366, height=343)
                    self.resultNgViewListbox_sideclient3_scrollbar.place(x=2000, y=2000, width=20, height=343)
                    # sideclient4
                    self.resultNgViewListbox_sideclient4.place(x=2000, y=2000, width=366, height=343)
                    self.resultNgViewListbox_sideclient4_scrollbar.place(x=2000, y=2000, width=20, height=343)
                else:
                    # sideclient2
                    self.resultNgViewListbox_sideclient2.place(x=2000, y=2000, width=366, height=343)
                    self.resultNgViewListbox_sideclient2_scrollbar.place(x=2000, y=2000, width=20, height=343)
                    # sideclient3
                    self.resultNgViewListbox_sideclient3.place(x=2000, y=2000, width=366, height=343)
                    self.resultNgViewListbox_sideclient3_scrollbar.place(x=2000, y=2000, width=20, height=343)
                self.ngCheckMode = False

            # ILJIN 클릭 시 LOG 출력
            if 1446 < x < 1832 and 43 < y < 92:
                self.view_mode = False if self.view_mode == True else True
                if self.view_mode == False:
                    self.main_canvas.itemconfig(self.stateText, state = 'normal')
                    self.mainframe_log.place(x=2000, y=2000, width=428, height=349)
                    self.mainframe_scrollbar.place(x=2000, y=2000, width=20, height=349)
                    self.main_canvas.itemconfig(self.background_image, image=self.main_frame_img)
                    for i in range(10):
                        self.main_canvas.itemconfig(self.bad_image1[i], state="normal")
                        self.main_canvas.itemconfig(self.bad_image2[i], state="normal")
                        self.main_canvas.itemconfig(self.bad_text1[i], state="normal")
                        self.main_canvas.itemconfig(self.bad_text2[i], state="normal")
                else:
                    self.main_canvas.itemconfig(self.stateText, state = 'hidden')
                    self.mainframe_log.place(x=1414, y=492, width=428, height=539)
                    self.mainframe_scrollbar.place(x=1411 + 408, y=492, width=20, height=539)
                    self.main_canvas.itemconfig(self.background_image, image=self.main_frame_img2)
                    for i in range(10):
                        self.main_canvas.itemconfig(self.bad_image1[i], state="hidden")
                        self.main_canvas.itemconfig(self.bad_image2[i], state="hidden")
                        self.main_canvas.itemconfig(self.bad_text1[i], state="hidden")
                        self.main_canvas.itemconfig(self.bad_text2[i], state="hidden")

            # 관리자 모드 클릭 (우측 상단 톱니바퀴)
            if 1846 < x < 1910 and 9 < y < 84:
                self.main_canvas.itemconfig(self.AdminPasswordBg, state = 'normal')
                self.main_canvas.itemconfig(self.showEntry, window = self.entry)
                self.PasswordMode = True
                self.management_mode = None
                # management_frame.tkraise()
                pass

            check = 0
            for i in range(9, -1, -1):
                if 1376 < x < 1619 and 994 - check < y < 1044 - check and self.ngSettingMode == False:
                    print(i)
                    ngcage_frame.ngcageStage = 1
                    ngcage_frame.ngcageNumber = i
                    if ngcage_frame.ngCageUpdate1[ngcage_frame.ngcageNumber][0] == True:
                        ngcage_frame.ngcageInit(ngcage_frame.ngcageStage, ngcage_frame.ngcageNumber)
                        ngcage_frame.tkraise()
                    else:
                        print("ngCageUpdate1 not data")
                if 1619 < x < 1879 and 994 - check < y < 1044 - check and self.ngSettingMode == False:
                    print(i)
                    ngcage_frame.ngcageStage = 2
                    ngcage_frame.ngcageNumber = i
                    if ngcage_frame.ngCageUpdate2[ngcage_frame.ngcageNumber][0] == True:
                        ngcage_frame.ngcageInit(ngcage_frame.ngcageStage, ngcage_frame.ngcageNumber)
                        ngcage_frame.tkraise()
                    else:
                        print("ngCageUpdate2 not data")
                # self.bad_image1[i] = self.main_canvas.create_image(1636, 995-check, image = self.main_frame_bad, anchor = 'nw', state = 'normal')
                # self.bad_text1[i] = self.main_canvas.create_text(1755, 1021-check, text='', font=('gothic', 20, 'bold'), fill='white', anchor='center')
                check += 57

            if 31 < x < 31 + 180 and 972 < y < 972 + 73:
                print("HISTORY MODE")
                history_frame.tkraise()

            # SETTINGS 클릭 이벤트
            if 251 < x < 251 + 180 and 972 < y < 972 + 73:
                self.setting_mode = None
                self.main_canvas.itemconfig(self.AdminPasswordBg, state = 'normal')
                self.main_canvas.itemconfig(self.showEntry, window = self.entry)
                self.PasswordMode = True

            # SETTINGS 화면 SAVE 버튼
            if 927 < x < 1105 and 972 < y < 1044:
                if self.setting_mode == True:
                    for i in range(5 if CodeSetup == 'CUP' else 6):
                        try:
                            self.scaleframe[i].destroy()
                            print("distory comp")
                            if STH.ManualMode == True:
                                print('수동 상태 수치 조절 저장')
                                TSD.sendAllClient(f"MPICKLE{i}") # 'M'anual mode
                            else :
                                print('자동 상태 수치 조절 저장')
                                TSD.sendAllClient(f"APICKLE{i}") # 'A'uto mode
                            time.sleep(0.1)
                        except Exception as ex:
                            print(ex)
                            continue

                    self.main_canvas.itemconfig(self.stateText, state = 'normal')
                    self.main_canvas.itemconfig(self.back_image, state="hidden")
                    self.main_canvas.itemconfig(self.save_image, state="hidden")
                    self.setting_mode = False

            # SETTINGS 화면 BACK 버튼
            if 1146 < x < 1326 and 972 < y < 1044:
                if self.setting_mode == True:
                    self.setting_mode = False

                    for i in range(5 if CodeSetup == 'CUP' else 6):
                        try:
                            self.scaleframe[i].destroy()
                        except:
                            pass
                    self.main_canvas.itemconfig(self.stateText, state = 'normal')
                    self.main_canvas.itemconfig(self.back_image, state="hidden")
                    self.main_canvas.itemconfig(self.save_image, state="hidden")

            if self.management_mode == True:
                self.management_mode = False
                main_frame.tkraise()
        else:
            # 패스워드 확인 클릭 시 (이때 SETTINGS, NG SETTINGS 화면이 전환됨)
            if 801 < x < 941 and 579 < y < 634:
                readPassword = tk.Entry.get(self.entry)
                if readPassword == self.Password:
                    self.main_canvas.itemconfig(self.stateText, state = 'hidden')
                    self.main_canvas.itemconfig(self.AdminPasswordBg, state = 'hidden')
                    self.main_canvas.itemconfig(self.showEntry, window = '')
                    self.entry.delete(0,999)
                    self.PasswordCheck = True
                    self.PasswordMode = False
                    print('password Check')

                    # NG SETTINGS 화면
                    if self.ngSettingMode == None:
                        print("NG SETTINGS Btn Click")
                        self.ngSettingLoad("setting")

                        for i in range(6):
                            if i == 5 and CodeSetup == 'CUP':
                                break
                            self.main_canvas.itemconfig(self.ngSettingText[i], state="normal")
                        main_frame.main_canvas.itemconfig(main_frame.ngSettingBtn, state="normal")
                        for i in range(ConnectClientCheck):
                            self.main_canvas.itemconfig(self.ngSettingBg[i], state="normal")
                        for i in range(ConnectClientCheck*2):
                            self.main_canvas.itemconfig(self.entryBoxView[i], window = '')
                        self.main_canvas.itemconfig(self.back_image, state="normal")
                        self.main_canvas.itemconfig(self.save_image, state="normal")
                        self.ngSettingMode = True

                    # SETTINGS 화면
                    if self.setting_mode == None:
                        print("SETTING MODE")
                        self.setting_mode = True

                        if self.setting_mode == True:
                            TSD.sendAllClient("SETTING")
                            self.main_canvas.itemconfig(self.back_image, state="normal")
                            self.main_canvas.itemconfig(self.save_image, state="normal")
                    if self.management_mode == None:
                        self.management_mode = True
                        management_frame.tkraise()
                else:
                    messagebox.showerror("에러발생", "정확한 패스워드를 기입하여 주세요.")
                    self.entry.delete(0,999)

            if 977 < x < 1116 and 579 < y < 634:
                print('password Pass, exit')
                self.main_canvas.itemconfig(self.AdminPasswordBg, state = 'hidden')
                self.main_canvas.itemconfig(self.showEntry, window = '')
                self.entry.delete(0,999)
                self.PasswordMode = False
                self.ngSettingMode = False                
                self.setting_mode = False
                self.management_mode = False

        # 불량 확인 # # # # # # # # # # # # # # # # # # # # # # # # # # 
        if 699 < x < 699 + 187 and 969 < y < 969 + 79:
            result = messagebox.askyesno("확인", "불량을 확인하시겠습니까?")
            if result == True:
                print('불량 확인')
                self.ng_check_click = True
                # 초기화
                NF.ng_canvas.itemconfig(NF.part_name, text='')
                NF.ng_canvas.itemconfig(NF.ng_name, text='')
                NF.image_data = {
                    'OK': {'filepaths': [], 'current_index': 0},
                    'NG': {'filepaths': [], 'current_index': 0}
                }
                NF.ok_image = ''
                NF.ng_image = ''
            else:
                self.ng_check_click = False
            
        # 111
        if 25 < x < 25 + 415 and 119 < y < 119 + 316 :
            if self.ng_check_click == True :
                result = messagebox.askyesno("확인", "1차 상부를 확인하시겠습니까?")
                if result == True:
                    print('1차 상부 선택')
                    self.ng_check_click = False
                    if CodeSetup == "CONE":
                        NF.cam_num = '111'
                    else :
                        NF.cam_num = '101'
                    NF.tkraise()
        # 112
        if 471 < x < 471 + 415 and 119 < y < 119 + 316 :
            if self.ng_check_click == True :
                result = messagebox.askyesno("확인", "1차 대각을 확인하시겠습니까?")
                if result == True:
                    print('1차 대각 선택')
                    self.ng_check_click = False
                    if CodeSetup == "CONE":
                        NF.cam_num = '112'
                    else :
                        NF.cam_num = '102'
                    NF.tkraise()
        # 113
        if 920 < x < 920 + 415 and 119 < y < 119 + 316 and CodeSetup == "CONE":
            if self.ng_check_click == True :
                result = messagebox.askyesno("확인", "1차 측면을 확인하시겠습니까?")
                if result == True:
                    print('1차 측면 선택')
                    self.ng_check_click = False
                    NF.cam_num = '113'
                    NF.tkraise()
        # 114
        if 25 < x < 25 + 415 and 537 < y < 537 + 316 :
            if self.ng_check_click == True :
                result = messagebox.askyesno("확인", "2차 상부를 확인하시겠습니까?")
                if result == True:
                    print('2차 상부 선택')
                    self.ng_check_click = False
                    if CodeSetup == "CONE":
                        NF.cam_num = '114'
                    else :
                        NF.cam_num = '104'
                    NF.tkraise()
        # 115
        if 471 < x < 471 + 415 and 537 < y < 537 + 316 :
            if self.ng_check_click == True :
                result = messagebox.askyesno("확인", "2차 대각을 확인하시겠습니까?")
                if result == True:
                    print('2차 대각 클릭')
                    self.ng_check_click = False
                    if CodeSetup == "CONE":
                        NF.cam_num = '115'
                    else :
                        NF.cam_num = '103'
                    NF.tkraise()
        # 116
        if 920 < x < 920 + 415 and 537 < y < 537 + 316 :
            if self.ng_check_click == True :
                result = messagebox.askyesno("확인", "2차 측면을 확인하시겠습니까?")
                if result == True:
                    print('2차 측면 클릭')
                    self.ng_check_click = False
                    if CodeSetup == "CONE":
                        NF.cam_num = '116'
                    else :
                        NF.cam_num = '105'
                    NF.tkraise()
        # BACK 버튼
        if 1683 < x < 1683 + 197 and 954 < y < 954 + 87 :
            if self.ng_check_click == True :
                self.ng_check_click = False

    # 팝업창 gif 효과 
    def twinkle(self, imagelist, cnt, done) :
        if done == True :
            self.main_canvas.itemconfig(self.popup_place, state="hidden")
        elif cnt % 2 == 0 :
            self.main_canvas.itemconfig(self.popup_place, image=imagelist[0], state="normal")
        elif cnt % 2 == 1 :
            self.main_canvas.itemconfig(self.popup_place, image=imagelist[1], state="normal")

    # 팝업창 스레드
    def popup_thread(self):
        cnt = 0
        while 1:
            try :
                time.sleep(0.5)
                cnt += 1
                if self.ng_check_click == False :
                    self.twinkle(self.popup_image, cnt, True)
                elif self.ng_check_click == True :
                    self.twinkle(self.popup_image, cnt, False)

                # 초기화
                if cnt == 20 :
                    cnt = 0
            except :
                pass


class ngCageShowFrame(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.ngcageBgImage = ImageTk.PhotoImage(file=f"bg/{BgSetup}/ngcage.png")
        self.ngcage_frame_connecticon = ImageTk.PhotoImage(file=f"bg/{BgSetup}/connecticon.png")

        self.ngcageStage = 0
        self.ngcageNumber = 0
        self.ngcage_cam = [None] * (6 if CodeSetup == "CONE" else 5)
        self.updateImage = [None] * (6 if CodeSetup == "CONE" else 5)
        self.connectionCheck = [None] * (6 if CodeSetup == "CONE" else 5)
        self.ngcage_ok = [None] * (6 if CodeSetup == "CONE" else 5)
        self.ngcage_ng = [None] * (6 if CodeSetup == "CONE" else 5)
        self.nowSection = 0

        self.ngCageUpdate1 = [
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
        ]
        self.ngCageUpdate2 = [
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
            [False, [None, None, None, None, None, None], [None, None, None, None, None, None], [None],],
        ]

        self.bad_image_ngcage = [None] * 10
        self.bad_text_ngcage = [None] * 10
        self.create_widgets()

    def create_widgets(self):
        self.grid(row=0, column=0)
        self.ngcage_canvas = tk.Canvas(self, width=1920, height=1080)
        self.ngcageBg = self.ngcage_canvas.create_image(0, 0, image=self.ngcageBgImage, anchor="nw")
        self.ngcage_cam[0] = self.ngcage_canvas.create_image(75, 596, image="", anchor="nw", state="normal")
        self.ngcage_cam[1] = self.ngcage_canvas.create_image(75, 178, image="", anchor="nw", state="normal")
        self.ngcage_cam[2] = self.ngcage_canvas.create_image(523, 179, image="", anchor="nw", state="normal")
        if CodeSetup == "CONE":
            self.ngcage_cam[3] = self.ngcage_canvas.create_image(971, 179, image="", anchor="nw", state="normal")
            self.ngcage_cam[4] = self.ngcage_canvas.create_image(523, 597, image="", anchor="nw", state="normal")
            self.ngcage_cam[5] = self.ngcage_canvas.create_image(971, 597, image="", anchor="nw", state="normal")
        else:
            self.ngcage_cam[3] = self.ngcage_canvas.create_image(523, 597, image="", anchor="nw", state="normal")
            self.ngcage_cam[4] = self.ngcage_canvas.create_image(971, 597, image="", anchor="nw", state="normal")
        self.ngcage_ok[0] = self.ngcage_canvas.create_image(71, 905, image=main_frame.main_frame_ok, anchor="nw", state="hidden")
        self.ngcage_ok[1] = self.ngcage_canvas.create_image(71, 488, image=main_frame.main_frame_ok, anchor="nw", state="hidden")
        self.ngcage_ok[2] = self.ngcage_canvas.create_image(517, 488, image=main_frame.main_frame_ok, anchor="nw", state="hidden")
        if CodeSetup == "CONE":
            self.ngcage_ok[3] = self.ngcage_canvas.create_image(965, 488, image=main_frame.main_frame_ok, anchor="nw", state="hidden")
            self.ngcage_ok[4] = self.ngcage_canvas.create_image(517, 905, image=main_frame.main_frame_ok, anchor="nw", state="hidden")
            self.ngcage_ok[5] = self.ngcage_canvas.create_image(965, 905, image=main_frame.main_frame_ok, anchor="nw", state="hidden")
        else:
            self.ngcage_ok[3] = self.ngcage_canvas.create_image(517, 905, image=main_frame.main_frame_ok, anchor="nw", state="hidden")
            self.ngcage_ok[4] = self.ngcage_canvas.create_image(965, 905, image=main_frame.main_frame_ok, anchor="nw", state="hidden")
        self.ngcage_ng[0] = self.ngcage_canvas.create_image(291, 905, image=main_frame.main_frame_ng, anchor="nw", state="hidden")
        self.ngcage_ng[1] = self.ngcage_canvas.create_image(291, 488, image=main_frame.main_frame_ng, anchor="nw", state="hidden")
        self.ngcage_ng[2] = self.ngcage_canvas.create_image(738, 488, image=main_frame.main_frame_ng, anchor="nw", state="hidden")
        if CodeSetup == "CONE":
            self.ngcage_ng[3] = self.ngcage_canvas.create_image(1186, 488, image=main_frame.main_frame_ng, anchor="nw", state="hidden")
            self.ngcage_ng[4] = self.ngcage_canvas.create_image(738, 905, image=main_frame.main_frame_ng, anchor="nw", state="hidden")
            self.ngcage_ng[5] = self.ngcage_canvas.create_image(1186, 905, image=main_frame.main_frame_ng, anchor="nw", state="hidden")
        else:
            self.ngcage_ng[3] = self.ngcage_canvas.create_image(738, 905, image=main_frame.main_frame_ng, anchor="nw", state="hidden")
            self.ngcage_ng[4] = self.ngcage_canvas.create_image(1186, 905, image=main_frame.main_frame_ng, anchor="nw", state="hidden")

        self.connectionCheck[0] = self.ngcage_canvas.create_image(127, 74, image=self.ngcage_frame_connecticon, anchor="nw", state="hidden")
        self.connectionCheck[1] = self.ngcage_canvas.create_image(127, 30, image=self.ngcage_frame_connecticon, anchor="nw", state="hidden")
        self.connectionCheck[2] = self.ngcage_canvas.create_image(307, 30, image=self.ngcage_frame_connecticon, anchor="nw", state="hidden")
        if CodeSetup == "CONE":
            self.connectionCheck[3] = self.ngcage_canvas.create_image(487, 30, image=self.ngcage_frame_connecticon, anchor="nw", state="hidden")
            self.connectionCheck[4] = self.ngcage_canvas.create_image(307, 74, image=self.ngcage_frame_connecticon, anchor="nw", state="hidden")
            self.connectionCheck[5] = self.ngcage_canvas.create_image(487, 74, image=self.ngcage_frame_connecticon, anchor="nw", state="hidden")
        else:
            self.connectionCheck[3] = self.ngcage_canvas.create_image(307, 74, image=self.ngcage_frame_connecticon, anchor="nw", state="hidden")
            self.connectionCheck[4] = self.ngcage_canvas.create_image(487, 74, image=self.ngcage_frame_connecticon, anchor="nw", state="hidden")

        check = 0
        for i in range(10):
            self.bad_image_ngcage[i] = self.ngcage_canvas.create_image(1527, 266 + check, image=main_frame.main_frame_bad, anchor="nw", state="hidden")
            self.bad_text_ngcage[i] = self.ngcage_canvas.create_text(1653, 292 + check, text="", font=("gothic", 18, "bold"), fill="white", anchor="center", state="normal",)
            check += 57
        self.nowSectionText = self.ngcage_canvas.create_text(1645, 204, text=self.nowSection, font=("gothic", 18, "bold"), fill="white", anchor="center", state="normal",)

        self.ngcage_canvas.bind("<Button-1>", self.ngcage_btn)
        self.ngcage_canvas.pack()

    def ngcageInit(self, cagenum, number):
        if cagenum == 1:
            # print(self.ngCageUpdate1[number])
            for x, i in enumerate(self.ngCageUpdate1[number][3]):
                self.ngcage_canvas.itemconfig(self.bad_text_ngcage[x], text=DTT.badType[i])
                self.ngcage_canvas.itemconfig(self.bad_image_ngcage[x], state="normal")

            for i in range(6 if CodeSetup == "CONE" else 5):
                try:
                    self.show_img(self.ngCageUpdate1[number][2][i][0], i)
                except:
                    NoneImg = cv2.imread(f"bg/{BgSetup}/NoneData.png")
                    self.show_img(NoneImg, i)

            self.imageUpdate()

        elif cagenum == 2:
            # print(self.ngCageUpdate2[number])
            for x, i in enumerate(self.ngCageUpdate2[number][3]):
                self.ngcage_canvas.itemconfig(self.bad_text_ngcage[x], text=DTT.badType[int(i)])
                self.ngcage_canvas.itemconfig(self.bad_image_ngcage[x], state="normal")

            for i in range(6 if CodeSetup == "CONE" else 5):
                # print(type(self.ngCageUpdate2[number][2][i][0]))
                try:
                    self.show_img(self.ngCageUpdate2[number][2][i][0], i)
                except:
                    NoneImg = cv2.imread(f"bg/{BgSetup}/NoneData.png")
                    self.show_img(NoneImg, i)

            self.imageUpdate()

    def show_img(self, image, number):
        # try:
        #     image = cv2.resize(image, dsize=(391, 290), interpolation=cv2.INTER_AREA)
        # except:
        #     img = cv2.imread(f'bg/{BgSetup}/NoneData.png')
        #     image = cv2.resize(img, dsize=(391, 290), interpolation=cv2.INTER_AREA)

        self.cvtimg = cv2.cvtColor(image, cv2.COLOR_BGR2RGBA)
        self.cvtshow = Image.fromarray(self.cvtimg)
        # self.cvtshow = Image.fromarray(image)
        self.updateImage[number] = ImageTk.PhotoImage(image=self.cvtshow)
        self.ngcage_canvas.itemconfig(self.ngcage_cam[number], image=self.updateImage[number])

    def imageUpdate(self):
        if self.ngcageStage == 1:
            for i in range(6 if CodeSetup == "CONE" else 5):
                # print(type(self.ngCageUpdate2[self.ngcageNumber][2][i][self.nowSection]))
                try:
                    self.show_img(self.ngCageUpdate1[self.ngcageNumber][2][i][self.nowSection], i)
                    print(self.ngCageUpdate1[self.ngcageNumber][1][i])
                    if self.ngCageUpdate1[self.ngcageNumber][1][i] == DTT.NoneData:
                        self.ngcage_canvas.itemconfig(self.ngcage_ng[i], state="hidden")
                        self.ngcage_canvas.itemconfig(self.ngcage_ok[i], state="hidden")
                    elif "1" in self.ngCageUpdate1[self.ngcageNumber][1][i]:
                        self.ngcage_canvas.itemconfig(self.ngcage_ng[i], state="normal")
                        self.ngcage_canvas.itemconfig(self.ngcage_ok[i], state="hidden")
                    else:
                        self.ngcage_canvas.itemconfig(self.ngcage_ng[i], state="hidden")
                        self.ngcage_canvas.itemconfig(self.ngcage_ok[i], state="normal")
                except:
                    NoneDataImg = cv2.imread(f"bg/{BgSetup}/NoneData.png")
                    self.show_img(NoneDataImg, i)
                    self.ngcage_canvas.itemconfig(self.ngcage_ng[i], state="hidden")
                    self.ngcage_canvas.itemconfig(self.ngcage_ok[i], state="hidden")

        elif self.ngcageStage == 2:
            for i in range(6 if CodeSetup == "CONE" else 5):
                try:
                    # print(type(self.ngCageUpdate2[self.ngcageNumber][2][i][self.nowSection]))
                    self.show_img(self.ngCageUpdate2[self.ngcageNumber][2][i][self.nowSection], i)
                    print(self.ngCageUpdate2[self.ngcageNumber][1][i])

                    if self.ngCageUpdate2[self.ngcageNumber][1][i] == DTT.NoneData:
                        self.ngcage_canvas.itemconfig(self.ngcage_ng[i], state="hidden")
                        self.ngcage_canvas.itemconfig(self.ngcage_ok[i], state="hidden")
                    elif "1" in self.ngCageUpdate2[self.ngcageNumber][1][i]:
                        self.ngcage_canvas.itemconfig(self.ngcage_ng[i], state="normal")
                        self.ngcage_canvas.itemconfig(self.ngcage_ok[i], state="hidden")
                    else:
                        self.ngcage_canvas.itemconfig(self.ngcage_ng[i], state="hidden")
                        self.ngcage_canvas.itemconfig(self.ngcage_ok[i], state="normal")
                except:
                    NoneDataImg = cv2.imread(f"bg/{BgSetup}/NoneData.png")
                    self.show_img(NoneDataImg, i)
                    self.ngcage_canvas.itemconfig(self.ngcage_ng[i], state="hidden")
                    self.ngcage_canvas.itemconfig(self.ngcage_ok[i], state="hidden")

    def ngcage_btn(self, event):
        x = event.x
        y = event.y

        print(x, y)

        if 1448 < x < 1847 and 875 < y < 1023:
            for i in range(10):
                self.ngcage_canvas.itemconfig(self.bad_text_ngcage[i], text="")
                self.ngcage_canvas.itemconfig(self.bad_image_ngcage[i], state="hidden")
            self.nowSection = 0
            self.ngcage_canvas.itemconfig(self.nowSectionText, text=self.nowSection)
            main_frame.tkraise()

        if 1491 < x < 1619 and 160 < y < 249:
            print("left image move")
            self.nowSection -= 1
            if self.nowSection == -1:
                self.nowSection = 0
            self.ngcage_canvas.itemconfig(self.nowSectionText, text=self.nowSection)

            self.imageUpdate()

        if 1674 < x < 1797 and 160 < y < 252:
            print("right image move")
            self.nowSection += 1
            if self.nowSection == 11:
                self.nowSection = 10
            self.ngcage_canvas.itemconfig(self.nowSectionText, text=self.nowSection)

            self.imageUpdate()


class HistoryFrame(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.historyBgImage = ImageTk.PhotoImage(file=f"bg/{BgSetup}/history.png")
        self.NoneDataImage = ImageTk.PhotoImage(file=f"bg/{BgSetup}/NoneData.png")
        self.historyModelCheck = ['ALL']
        with open('CheckValue/product_info.json') as json_file:
            json_data = json.load(json_file)
        historyModelCheck_append = json_data[LINE][CodeSetup][1]
        self.historyModelCheck.extend(historyModelCheck_append)
        print(self.historyModelCheck)
        # self.historyModelCheck = (
        #     ["ALL","ST-558522APC","32007XJgH6X","ST-HE417119mH","ST-E336221mH","ST-356217gH","ST-386923","ST-387523gH","ST-417119SH","ST-417319SH",] if CodeSetup == "CONE" else ["ALL","ST-558522ASH","32007XJgH6X","ST-HE417119mH","ST-E336221mH","ST-356217gH","ST-386923","ST-387523gH","ST-417119SH","ST-417319SH"]
        # )
        self.histroyResultCheck = ["ALL", "OK", "NG", "NO"]
        self.historyInputText = [None] * 6
        self.historyInputData = ["", "", "", "", "", ""]
        self.historyImageData = [[], [], [], [], [], []]
        self.historyResultData = [None, None, None, None, None, None]
        self.history_ok = [None] * (6 if CodeSetup == "CONE" else 5)
        self.history_ng = [None] * (6 if CodeSetup == "CONE" else 5)
        self.history_cam = [None] * (6 if CodeSetup == "CONE" else 5)
        self.indexCount1 = 0
        self.indexCount2 = 0
        self.showingIndex = 0
        self.create_widgets()

    def create_widgets(self):
        self.grid(row=0, column=0)
        self.history_canvas = tk.Canvas(self, width=1920, height=1080)
        self.historyBg = self.history_canvas.create_image(0, 0, image=self.historyBgImage, anchor="nw")

        self.history_cam[0] = self.history_canvas.create_image(36, 548, image="", anchor="nw", state="normal")
        self.history_cam[1] = self.history_canvas.create_image(36, 131, image="", anchor="nw", state="normal")
        self.history_cam[2] = self.history_canvas.create_image(484, 131, image="", anchor="nw", state="normal")
        if CodeSetup == "CONE":
            self.history_cam[3] = self.history_canvas.create_image(932, 131, image="", anchor="nw", state="normal")
            self.history_cam[4] = self.history_canvas.create_image(484, 548, image="", anchor="nw", state="normal")
            self.history_cam[5] = self.history_canvas.create_image(932, 548, image="", anchor="nw", state="normal")
        else:
            self.history_cam[3] = self.history_canvas.create_image(484, 548, image="", anchor="nw", state="normal")
            self.history_cam[4] = self.history_canvas.create_image(932, 548, image="", anchor="nw", state="normal")

        self.history_ok[0] = self.history_canvas.create_image(31, 856, image=main_frame.main_frame_ok, anchor="nw", state="hidden")
        self.history_ok[1] = self.history_canvas.create_image(31, 438, image=main_frame.main_frame_ok, anchor="nw", state="hidden")
        self.history_ok[2] = self.history_canvas.create_image(479, 438, image=main_frame.main_frame_ok, anchor="nw", state="hidden")
        if CodeSetup == "CONE":
            self.history_ok[3] = self.history_canvas.create_image(927, 438, image=main_frame.main_frame_ok, anchor="nw", state="hidden")
            self.history_ok[4] = self.history_canvas.create_image(479, 856, image=main_frame.main_frame_ok, anchor="nw", state="hidden")
            self.history_ok[5] = self.history_canvas.create_image(927, 856, image=main_frame.main_frame_ok, anchor="nw", state="hidden")
        else:
            self.history_ok[3] = self.history_canvas.create_image(479, 856, image=main_frame.main_frame_ok, anchor="nw", state="hidden")
            self.history_ok[4] = self.history_canvas.create_image(927, 856, image=main_frame.main_frame_ok, anchor="nw", state="hidden")

        self.history_ng[0] = self.history_canvas.create_image(251, 856, image=main_frame.main_frame_ng, anchor="nw", state="hidden")
        self.history_ng[1] = self.history_canvas.create_image(251, 438, image=main_frame.main_frame_ng, anchor="nw", state="hidden")
        self.history_ng[2] = self.history_canvas.create_image(699, 438, image=main_frame.main_frame_ng, anchor="nw", state="hidden")
        if CodeSetup == "CONE":
            self.history_ng[3] = self.history_canvas.create_image(1147, 438, image=main_frame.main_frame_ng, anchor="nw", state="hidden")
            self.history_ng[4] = self.history_canvas.create_image(699, 856, image=main_frame.main_frame_ng, anchor="nw", state="hidden")
            self.history_ng[5] = self.history_canvas.create_image(1147, 856, image=main_frame.main_frame_ng, anchor="nw", state="hidden")
        else:
            self.history_ng[3] = self.history_canvas.create_image(699, 856, image=main_frame.main_frame_ng, anchor="nw", state="hidden")
            self.history_ng[4] = self.history_canvas.create_image(1147, 856, image=main_frame.main_frame_ng, anchor="nw", state="hidden")

        # TEXT
        self.historyInputText[0] = self.history_canvas.create_text(140, 1008, text="ALL", font=("gothic", 15, "bold"), fill="white", anchor="center")
        self.historyInputText[1] = self.history_canvas.create_text(416, 1008, text="ALL", font=("gothic", 15, "bold"), fill="white", anchor="center")
        self.historyInputText[2] = self.history_canvas.create_text(654, 1008, text="None", font=("gothic", 15, "bold"), fill="white", anchor="center")
        self.historyInputText[3] = self.history_canvas.create_text(894, 1008, text="None", font=("gothic", 15, "bold"), fill="white", anchor="center")
        self.historyIndexView = self.history_canvas.create_text(1632, 155, text="", font=("gothic", 20, "bold"), fill="white", anchor="center")

        # listbox
        color = "#%02x%02X%02x" % (19, 47, 60)
        self.history_log = tk.Listbox(self.history_canvas, fg="white", bg=color, font=("gothic", 12, "bold"))
        self.history_log.place(x=1392, y=239, width=476, height=699)
        self.history_scrollbar = tk.Scrollbar(self.history_canvas, bg=color, orient="vertical")
        self.history_scrollbar.config(command=self.history_log.yview)
        self.history_scrollbar.place(x=1392 + 436, y=239, width=40, height=699)
        self.history_log.config(yscrollcommand=self.history_scrollbar.set)
        self.history_log.bind("<<ListboxSelect>>", self.on_select)

        #model ComboBox
        self.select_model = ttk.Combobox(self, width = 14, justify = 'center', values= self.historyModelCheck, font= ("gothic", 15, "bold"), state = "readonly")
        self.select_model.current(0)
        self.show_select_Model = self.history_canvas.create_window(137, 1013, anchor='center', window=self.select_model, state = 'normal')

        #startTime ComboBox
        values_time = ['Default', '00', '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14','15', '16','17', '18', '19', '20', '21', '22', '23', '23']
        self.select_time_start = ttk.Combobox(self, width = 7, justify = 'center', values= values_time, font= ("gothic", 15, "bold"), state = "readonly")
        self.select_time_start.current(0)
        self.show_select_Stime = self.history_canvas.create_window(1115, 1013, anchor='center', window=self.select_time_start, state = 'normal')

        #endTime ComboBox
        values_time = ['Default', '00', '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14','15', '16','17', '18', '19', '20', '21', '22', '23', '23']
        self.select_time_end = ttk.Combobox(self, width = 7, justify = 'center', values= values_time, font= ("gothic", 15, "bold"), state = "readonly")
        self.select_time_end.current(0)
        self.show_select_Etime = self.history_canvas.create_window(1269, 1013, anchor='center', window=self.select_time_end, state = 'normal')

        self.history_canvas.bind("<Button-1>", self.history_btn)
        self.history_canvas.pack()

    def DbRead(self):
        self.history_log.delete(0, "end")

        # 입력 데이터 확인
        # if self.historyInputData[0] == "ALL" or self.historyInputData[0] == "":
        #     self.historyInputData[0] = None
        if self.historyInputData[1] == "ALL" or self.historyInputData[1] == "":
            self.historyInputData[1] = None
        if self.historyInputData[2] == "":
            self.historyInputData[2] = None
        if self.historyInputData[3] == "":
            self.historyInputData[3] = None

        if self.select_model.get() == 'ALL':
            self.historyInputData[0] = None
        else:
            self.historyInputData[0] = self.select_model.get()

        if self.select_time_start.get() == 'Default':
            self.historyInputData[4] = None
        else:
            self.historyInputData[4] = self.select_time_start.get()

        if self.select_time_end.get() == 'Default':
            self.historyInputData[5] = None
        else:
            self.historyInputData[5] = self.select_time_end.get()

        if self.historyInputData[4] != None:
            if self.historyInputData[2] == None:
                TodayDate = datetime.datetime.now().strftime("%Y-%m-%d")
                self.historyInputData[2] = TodayDate
                self.historyInputData[2] = self.historyInputData[2]+' '+self.historyInputData[4]
            else:
                if len(self.historyInputData[2]) > 10:
                    self.historyInputData[2] = self.historyInputData[2][0:10]
                self.historyInputData[2] = self.historyInputData[2]+' '+self.historyInputData[4]
        
        if self.historyInputData[5] != None:
            if self.historyInputData[3] == None:
                TodayDate = datetime.datetime.now().strftime("%Y-%m-%d")
                self.historyInputData[3] = TodayDate
                self.historyInputData[3] = self.historyInputData[3]+' '+self.historyInputData[5]
            else:
                if len(self.historyInputData[3]) > 10:
                    self.historyInputData[3] = self.historyInputData[3][0:10]
                self.historyInputData[3] = self.historyInputData[3]+' '+self.historyInputData[5]

        print(self.historyInputData)

        if self.historyInputData[1] == "OK":
            self.historyInputData[1] = DTT.OKdata
        elif self.historyInputData[1] == "NO":
            self.historyInputData[1] = DTT.NoneData

        if CodeSetup == "CONE":
            self.dbsqlRow = bearingartDB.readSql_Cone(ModelName=self.historyInputData[0], Result1=self.historyInputData[1], Result2=self.historyInputData[1], Result3=self.historyInputData[1], Result4=self.historyInputData[1], Result5=self.historyInputData[1], Result6=self.historyInputData[1], StartDate=self.historyInputData[2], EndDate=self.historyInputData[3],)  # self.history_canvas.itemconfig(self.productName, text = self.historyInputData[0])
        else:
            self.dbsqlRow = bearingartDB.readSql_Cup(ModelName=self.historyInputData[0], Result1=self.historyInputData[1], Result2=self.historyInputData[1], Result3=self.historyInputData[1], Result4=self.historyInputData[1], Result5=self.historyInputData[1], StartDate=self.historyInputData[2], EndDate=self.historyInputData[3],)  # self.history_canvas.itemconfig(self.productName, text = self.historyInputData[0])
        # print(self.dbsqlRow)
        # print(self.dbsqlRow[0][2])
        # print(self.dbsqlRow[0][3])
        # print(self.dbsqlRow[0][4])
        # print(self.dbsqlRow[0][5])
        # print(self.dbsqlRow[0][6])
        # print(self.dbsqlRow[0][7])
        print('SQL Low Data Reading Complete')
        dbUpdateList = [None] * len(self.dbsqlRow)
        if CodeSetup == "CONE":
            for i in range(len(self.dbsqlRow)):
                dbUpdateList[i] = self.dbsqlRow[i][1]
                try:
                    if "1" in self.dbsqlRow[i][2][0] or "1" in self.dbsqlRow[i][3][0] or "1" in self.dbsqlRow[i][4][0] or "1" in self.dbsqlRow[i][5][0] or "1" in self.dbsqlRow[i][6][0] or "1" in self.dbsqlRow[i][7][0]:
                        dbUpdateList[i] += " | NO | "
                    elif "1" == self.dbsqlRow[i][2][1] or "1" == self.dbsqlRow[i][2][2] or "1" == self.dbsqlRow[i][2][3] or "1" == self.dbsqlRow[i][2][4] or "1" == self.dbsqlRow[i][2][5] or "1" == self.dbsqlRow[i][3][1] or "1" == self.dbsqlRow[i][3][2] or "1" == self.dbsqlRow[i][3][3] or "1" == self.dbsqlRow[i][3][4] or "1" == self.dbsqlRow[i][3][5] or "1" == self.dbsqlRow[i][4][1] or "1" == self.dbsqlRow[i][4][2] or "1" == self.dbsqlRow[i][4][3] or "1" == self.dbsqlRow[i][4][4] or "1" == self.dbsqlRow[i][4][5] or "1" == self.dbsqlRow[i][5][1] or "1" == self.dbsqlRow[i][5][2] or "1" == self.dbsqlRow[i][5][3] or "1" == self.dbsqlRow[i][5][4] or "1" == self.dbsqlRow[i][5][5] or "1" == self.dbsqlRow[i][6][1] or "1" == self.dbsqlRow[i][6][2] or "1" == self.dbsqlRow[i][6][3] or "1" == self.dbsqlRow[i][6][4] or "1" == self.dbsqlRow[i][6][5] or "1" == self.dbsqlRow[i][7][1] or "1" == self.dbsqlRow[i][7][2] or "1" == self.dbsqlRow[i][7][3] or "1" == self.dbsqlRow[i][7][4] or "1" == self.dbsqlRow[i][7][5]:
                        dbUpdateList[i] += " | FU | "
                    elif "1" in self.dbsqlRow[i][2] or "1" in self.dbsqlRow[i][3] or "1" in self.dbsqlRow[i][4] or "1" in self.dbsqlRow[i][5] or "1" in self.dbsqlRow[i][6] or "1" in self.dbsqlRow[i][7]:
                        dbUpdateList[i] += " | NG | "
                    else:
                        dbUpdateList[i] += " | OK | "
                except:
                    if "1" in self.dbsqlRow[i][2][0] or "1" in self.dbsqlRow[i][3][0] or "1" in self.dbsqlRow[i][4][0] or "1" in self.dbsqlRow[i][5][0] or "1" in self.dbsqlRow[i][6][0] or "1" in self.dbsqlRow[i][7][0]:
                        dbUpdateList[i] += " | NO | "
                    elif "1" in self.dbsqlRow[i][2] or "1" in self.dbsqlRow[i][3] or "1" in self.dbsqlRow[i][4] or "1" in self.dbsqlRow[i][5] or "1" in self.dbsqlRow[i][6] or "1" in self.dbsqlRow[i][7]:
                        dbUpdateList[i] += " | NG | "
                    else:
                        dbUpdateList[i] += " | OK | "
                time = self.dbsqlRow[i][14]
                time = str(time)
                dbUpdateList[i] += time + " | "
                dbUpdateList[i] += self.dbsqlRow[i][8] + " | "
                dbUpdateList[i] += self.dbsqlRow[i][9] + " | "
                dbUpdateList[i] += self.dbsqlRow[i][10] + " | "
                dbUpdateList[i] += self.dbsqlRow[i][11] + " | "
                dbUpdateList[i] += self.dbsqlRow[i][12] + " | "
                dbUpdateList[i] += self.dbsqlRow[i][13] + " | "
                dbUpdateList[i] += self.dbsqlRow[i][2] + " | "
                dbUpdateList[i] += self.dbsqlRow[i][3] + " | "
                dbUpdateList[i] += self.dbsqlRow[i][4] + " | "
                dbUpdateList[i] += self.dbsqlRow[i][5] + " | "
                dbUpdateList[i] += self.dbsqlRow[i][6] + " | "
                dbUpdateList[i] += self.dbsqlRow[i][7]
        else:
            for i in range(len(self.dbsqlRow)):
                dbUpdateList[i] = self.dbsqlRow[i][1]
                try:
                    if "1" in self.dbsqlRow[i][2][0] or "1" in self.dbsqlRow[i][3][0] or "1" in self.dbsqlRow[i][4][0] or "1" in self.dbsqlRow[i][5][0] or "1" in self.dbsqlRow[i][6][0]:
                        dbUpdateList[i] += " | NO | "
                    elif "1" == self.dbsqlRow[i][2][1] or "1" == self.dbsqlRow[i][2][2] or "1" == self.dbsqlRow[i][2][3] or "1" == self.dbsqlRow[i][2][4] or "1" == self.dbsqlRow[i][2][5] or "1" == self.dbsqlRow[i][3][1] or "1" == self.dbsqlRow[i][3][2] or "1" == self.dbsqlRow[i][3][3] or "1" == self.dbsqlRow[i][3][4] or "1" == self.dbsqlRow[i][3][5] or "1" == self.dbsqlRow[i][4][1] or "1" == self.dbsqlRow[i][4][2] or "1" == self.dbsqlRow[i][4][3] or "1" == self.dbsqlRow[i][4][4] or "1" == self.dbsqlRow[i][4][5] or "1" == self.dbsqlRow[i][5][1] or "1" == self.dbsqlRow[i][5][2] or "1" == self.dbsqlRow[i][5][3] or "1" == self.dbsqlRow[i][5][4] or "1" == self.dbsqlRow[i][5][5] or "1" == self.dbsqlRow[i][6][1] or "1" == self.dbsqlRow[i][6][2] or "1" == self.dbsqlRow[i][6][3] or "1" == self.dbsqlRow[i][6][4] or "1" == self.dbsqlRow[i][6][5]:
                        dbUpdateList[i] += " | FU | "
                    elif "1" in self.dbsqlRow[i][2] or "1" in self.dbsqlRow[i][3] or "1" in self.dbsqlRow[i][4] or "1" in self.dbsqlRow[i][5] or "1" in self.dbsqlRow[i][6]:
                        dbUpdateList[i] += " | NG | "
                    else:
                        dbUpdateList[i] += " | OK | "
                except:
                    if "1" in self.dbsqlRow[i][2][0] or "1" in self.dbsqlRow[i][3][0] or "1" in self.dbsqlRow[i][4][0] or "1" in self.dbsqlRow[i][5][0] or "1" in self.dbsqlRow[i][6][0]:
                        dbUpdateList[i] += " | NO | "
                    elif "1" in self.dbsqlRow[i][2] or "1" in self.dbsqlRow[i][3] or "1" in self.dbsqlRow[i][4] or "1" in self.dbsqlRow[i][5] or "1" in self.dbsqlRow[i][6]:
                        dbUpdateList[i] += " | NG | "
                    else:
                        dbUpdateList[i] += " | OK | "
                time = self.dbsqlRow[i][12]
                time = str(time)
                dbUpdateList[i] += time + " | "
                dbUpdateList[i] += self.dbsqlRow[i][7] + " | "
                dbUpdateList[i] += self.dbsqlRow[i][8] + " | "
                dbUpdateList[i] += self.dbsqlRow[i][9] + " | "
                dbUpdateList[i] += self.dbsqlRow[i][10] + " | "
                dbUpdateList[i] += self.dbsqlRow[i][11] + " | "
                # dbUpdateList[i] += self.dbsqlRow[i][13] + ' | '
                dbUpdateList[i] += self.dbsqlRow[i][2] + " | "
                dbUpdateList[i] += self.dbsqlRow[i][3] + " | "
                dbUpdateList[i] += self.dbsqlRow[i][4] + " | "
                dbUpdateList[i] += self.dbsqlRow[i][5] + " | "
                dbUpdateList[i] += self.dbsqlRow[i][6]  # + ' | '
                # dbUpdateList[i] += self.dbsqlRow[i][7]
        print('SQL Low Data Trim Complete')
        dbUpdateList.reverse()
        print('SQL Low Trim Data Reverse Complete')

        for i in range(len(self.dbsqlRow)):
            self.history_log.insert(0, f'{i+1}. '+dbUpdateList[i])

        print('SQL Low Trim Reverse Data List Input Complete')

        x = 0
        for i in reversed(range(len(dbUpdateList))):
            try:
                # print(i)
                if "NG" in dbUpdateList[i]:
                    self.history_log.itemconfig(x, bg="red")
                elif "NO" in dbUpdateList[i]:
                    self.history_log.itemconfig(x, bg="black")
                elif "FU" in dbUpdateList[i]:
                    self.history_log.itemconfig(x, bg="magenta")
                else:
                    self.history_log.itemconfig(x, bg="green")
                x += 1
            except:
                x += 1

        print('SQL Low Trim Reverse Data List Color Setup Complete')

        return

        # self.badCheckList = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        # for i in range(len(self.dbsqlRow)):
        #     for x in range(5 if CodeSetup == 'CUP' else 6):
        #         # print(self.dbsqlRow[i][x + 2])
        #         for z in range(14):
        #             try:
        #                 if self.dbsqlRow[i][x + 2][z] == "1":
        #                     self.badCheckList[z] += 1
        #             except:
        #                 print("except break")
        #                 break

        # print('SQL Low Trim Data BadProduct Check')

        # for i in range(14):
        #     try:
        #         print(f"{DTT.badType[i]} : {self.badCheckList[i]}개 불량확인")
        #     except:
        #         print(f"DUMMY : {self.badCheckList[i]}개 불량확인")

    def on_select(self, event):
        w = event.widget
        idx = int(w.curselection()[0])
        value = w.get(idx)

        indexlist = value.split(" | ")
        print(indexlist, len(indexlist))
        self.historyImageData = [[], [], [], [], [], []]
        self.showingIndex = 0
        self.selectProduct = indexlist[0]
        self.history_canvas.itemconfig(self.historyIndexView, text=f"{indexlist[0]} - {self.showingIndex}")
        """
        0 : 기종
        1 : 결과
        2 : 시간
        3 ~ 8 : 파일 경로
        9 ~ 14 : 개별 결과
        """
        for i in range(3, 9 if CodeSetup == "CONE" else 8):
            pathlist = []
            pathlist = indexlist[i].split("||")
            for x in range(len(pathlist)):
                try:
                    image = ImageTk.PhotoImage(file=pathlist[x])
                except:
                    print(traceback.format_exc())
                    image = ImageTk.PhotoImage(file=f"bg/{BgSetup}/OKDATA.png")
                self.historyImageData[i - 3].append(image)
            self.history_canvas.itemconfig(self.history_cam[i - 3], image=self.historyImageData[i - 3][0])

        for i in range(9 if CodeSetup == "CONE" else 8, 15 if CodeSetup == "CONE" else 13):
            if '1' not in indexlist[i]:
                self.history_canvas.itemconfig(self.history_ok[i - 9 if CodeSetup == "CONE" else i - 8], state="normal")
                self.history_canvas.itemconfig(self.history_ng[i - 9 if CodeSetup == "CONE" else i - 8], state="hidden")
                self.historyResultData[i - 9] = True
            elif indexlist[i][0] == '1':#DTT.NoneData or indexlist[i] == "100000000000" or indexlist[i] == "10000000":
                self.history_canvas.itemconfig(self.history_ok[i - 9 if CodeSetup == "CONE" else i - 8], state="hidden")
                self.history_canvas.itemconfig(self.history_ng[i - 9 if CodeSetup == "CONE" else i - 8], state="hidden")
                self.historyResultData[i - 9] = None
            else:
                self.history_canvas.itemconfig(self.history_ok[i - 9 if CodeSetup == "CONE" else i - 8], state="hidden")
                self.history_canvas.itemconfig(self.history_ng[i - 9 if CodeSetup == "CONE" else i - 8], state="normal")
                self.historyResultData[i - 9] = False

    def show_calendar(self, inputdata):

        def print_sel_start():
            split_text = cal.selection_get()
            test = split_text.strftime("%Y-%m-%d")
            # test = test.split(',')
            self.year = test[0]
            self.month = test[1]
            self.day = test[2]
            # input_date = self.year + "년 " + self.month + "월 " + self.day + "일"
            self.history_canvas.itemconfig(self.historyInputText[2], text=test)
            self.historyInputData[2] = test
            top.destroy()

        def print_sel_end():
            split_text = cal.selection_get()
            test = split_text.strftime("%Y-%m-%d")
            # test = test.split(',')
            self.year = test[0]
            self.month = test[1]
            self.day = test[2]
            # input_date = self.year + "년 " + self.month + "월 " + self.day + "일"
            self.history_canvas.itemconfig(self.historyInputText[3], text=test)
            self.historyInputData[3] = test
            top.destroy()

        top = tk.Toplevel(root)

        now = datetime.datetime.now()

        years = now.year
        months = now.month
        days = now.day

        cal = Calendar(top, font="Arial 14", selectmode="day", cursor="hand1", year=years, month=months, day=days,)
        cal.pack(fill="both", expand=True)
        if inputdata == "start":
            tk.Button(top, text="Select", command=print_sel_start).pack()
        elif inputdata == "end":
            tk.Button(top, text="Select", command=print_sel_end).pack()

    def history_btn(self, event):
        x = event.x
        y = event.y

        print(x, y)

        if 31 < x < 278 and 23 < y < 94:
            print('생산 수량 확인 인터페이스 활성화')
            history_produce_frame.tkraise()

        if 1374 < x < 1460 and 125 < y < 187:
            print("left index move")
            self.showingIndex -= 1
            if self.showingIndex == -1:
                self.showingIndex = 0
            self.history_canvas.itemconfig(self.historyIndexView, text=f"{self.selectProduct} - {self.showingIndex}")
            for i in range(6 if CodeSetup == "CONE" else 5):
                try:
                    self.history_canvas.itemconfig(self.history_cam[i], image=self.historyImageData[i][self.showingIndex])
                    if self.historyResultData[i] == True:
                        self.history_canvas.itemconfig(self.history_ok[i], state="normal")
                        self.history_canvas.itemconfig(self.history_ng[i], state="hidden")
                    elif self.historyResultData[i] == None:
                        self.history_canvas.itemconfig(self.history_ok[i], state="hidden")
                        self.history_canvas.itemconfig(self.history_ng[i], state="hidden")
                    else:
                        self.history_canvas.itemconfig(self.history_ok[i], state="hidden")
                        self.history_canvas.itemconfig(self.history_ng[i], state="normal")
                except:
                    self.history_canvas.itemconfig(self.history_cam[i], image=self.NoneDataImage)
                    self.history_canvas.itemconfig(self.history_ok[i], state="hidden")
                    self.history_canvas.itemconfig(self.history_ng[i], state="hidden")

        if 1795 < x < 1877 and 125 < y < 187:
            print("right index move")
            self.showingIndex += 1
            if self.showingIndex == 10:
                self.showingIndex = 9
            self.history_canvas.itemconfig(self.historyIndexView, text=f"{self.selectProduct} - {self.showingIndex}")
            for i in range(6 if CodeSetup == "CONE" else 5):
                try:
                    self.history_canvas.itemconfig(self.history_cam[i], image=self.historyImageData[i][self.showingIndex])
                    if self.historyResultData[i] == True:
                        self.history_canvas.itemconfig(self.history_ok[i], state="normal")
                        self.history_canvas.itemconfig(self.history_ng[i], state="hidden")
                    elif self.historyResultData[i] == None:
                        self.history_canvas.itemconfig(self.history_ok[i], state="hidden")
                        self.history_canvas.itemconfig(self.history_ng[i], state="hidden")
                    else:
                        self.history_canvas.itemconfig(self.history_ok[i], state="hidden")
                        self.history_canvas.itemconfig(self.history_ng[i], state="normal")
                except:
                    self.history_canvas.itemconfig(self.history_cam[i], image=self.NoneDataImage)
                    self.history_canvas.itemconfig(self.history_ok[i], state="hidden")
                    self.history_canvas.itemconfig(self.history_ng[i], state="hidden")

        if 1630 < x < 1882 and 971 < y < 1043:
            print("BACK")
            main_frame.tkraise()

        if 1376 < x < 1625 and 970 < y < 1044:
            print("SEARCH")
            # self.DbRead()
            threading.Thread(target=self.DbRead, daemon=True).start()

        # if 274 < x < 338 and 970 < y < 1040:
        #     self.indexCount1 += 1
        #     index = self.indexCount1 % len(self.historyModelCheck)
        #     self.history_canvas.itemconfig(self.historyInputText[0], text=self.historyModelCheck[index])
        #     self.historyInputData[0] = self.historyModelCheck[index]
        #     print("OPTION1")

        if 497 < x < 571 and 970 < y < 1040:
            self.indexCount2 += 1
            index = self.indexCount2 % 4
            self.history_canvas.itemconfig(self.historyInputText[1], text=self.histroyResultCheck[index])
            self.historyInputData[1] = self.histroyResultCheck[index]
            print(self.historyInputData)
            print("OPTION2")

        if 737 < x < 812 and 970 < y < 1040:
            self.show_calendar("start")
            print("OPTION3")

        if 973 < x < 1048 and 970 < y < 1040:
            self.show_calendar("end")
            print("OPTION4")

class historyProduceFrame(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.history_produce_img = ImageTk.PhotoImage(file=f"bg/{BgSetup}/history_produce.png")

        # get product
        with open('CheckValue/product_info.json') as json_file:
            json_data = json.load(json_file)
        self.ProductName = json_data[LINE][CodeSetup][1]
        self.ProductName.append('ALL')
        self.ProductNameCount = -1
        self.ButtonCount = len(self.ProductName)

        self.SearchValue = ['ALL', 'None', 'None', 0, 23]
        self.SearchValueViewText = [None, None, None, None, None]

        self.SearchViewText = [None, None, None, None, None]

        self.master.attributes("-fullscreen", True)
        self.master.bind("<F11>", lambda event: self.master.attributes("-fullscreen", not self.master.attributes("-fullscreen")))
        self.master.bind("<Escape>", lambda event: self.master.attributes("-fullscreen", False))
        self.create_widgets()

    def create_widgets(self):
        self.grid(row=0, column=0)
        self.history_produce_canvas = tk.Canvas(self, width=1920, height=1080)
        self.hisotry_produce_BG = self.history_produce_canvas.create_image(0, 0, image=self.history_produce_img, anchor="nw")
        self.SearchViewText[0] = self.history_produce_canvas.create_text(440, 247, text="None Search Data", font=("gothic", 25, "bold"), fill="white", anchor="center")
        self.SearchViewText[1] = self.history_produce_canvas.create_text(440, 390, text="None Search Data", font=("gothic", 25, "bold"), fill="white", anchor="center")
        self.SearchViewText[2] = self.history_produce_canvas.create_text(440, 534, text="None Search Data", font=("gothic", 25, "bold"), fill="white", anchor="center")
        self.SearchViewText[3] = self.history_produce_canvas.create_text(440, 676, text="None Search Data", font=("gothic", 25, "bold"), fill="white", anchor="center")
        self.SearchViewText[4] = self.history_produce_canvas.create_text(440, 819, text="None Search Data", font=("gothic", 25, "bold"), fill="white", anchor="center")
        self.SearchValueViewText[0] = self.history_produce_canvas.create_text(170, 1000, text=self.SearchValue[0], font=("gothic", 20, "bold"), fill="white", anchor="center")
        self.SearchValueViewText[1] = self.history_produce_canvas.create_text(465, 1000, text=self.SearchValue[1], font=("gothic", 20, "bold"), fill="white", anchor="center")
        self.SearchValueViewText[2] = self.history_produce_canvas.create_text(722, 1000, text=self.SearchValue[2], font=("gothic", 20, "bold"), fill="white", anchor="center")
        self.SearchValueViewText[3] = self.history_produce_canvas.create_text(1057, 1000, text=self.SearchValue[3], font=("gothic", 25, "bold"), fill="white", anchor="center")
        self.SearchValueViewText[4] = self.history_produce_canvas.create_text(1417, 1000, text=self.SearchValue[4], font=("gothic", 25, "bold"), fill="white", anchor="center")

        # listbox
        color = "#%02x%02X%02x" % (19, 47, 60)
        self.history_produce_log = tk.Listbox(self.history_produce_canvas, fg="white", bg=color, font=("gothic", 15, "bold"))
        self.history_produce_log.place(x=701, y=202, width=1139, height=669)
        self.history_produce_scrollbar = tk.Scrollbar(self.history_produce_canvas, bg=color, orient="vertical")
        self.history_produce_scrollbar.config(command=self.history_produce_log.yview)
        self.history_produce_scrollbar.place(x=701+1139, y=202, width=20, height=669)
        self.history_produce_log.config(yscrollcommand=self.history_produce_scrollbar.set)
        self.history_produce_log.bind("<<ListboxSelect>>", self.on_select)

        self.history_produce_canvas.bind("<Button-1>", self.history_produce_btn)
        self.history_produce_canvas.pack()

    def history_produce_btn(self, event):
        x = event.x
        y = event.y

        print(x,y)

        if 294 < x < 355 and 969 < y < 1032:
            #형번 선택 버튼 클릭
            print('형번 선택')
            self.ProductNameCount += 1
            indexSetup = self.ProductNameCount%self.ButtonCount
            print(self.ProductNameCount , self.ButtonCount, indexSetup)
            self.history_produce_canvas.itemconfig(self.SearchValueViewText[0], text = self.ProductName[indexSetup])
            self.SearchValue[0] = self.ProductName[indexSetup]

            print(self.SearchValue)

        if 552 < x < 609 and 969 < y < 1032:
            #검색 시작 날짜 버튼 클릭
            print('검색 시작 날짜')
            self.show_calendar('start')

        if 812 < x < 873 and 969 < y < 1032:
            #검색 끝 날짜 버튼 클릭
            print('검색 끝 날짜')
            self.show_calendar('end')

        if 891 < x < 966 and 966 < y < 1039:
            #검색 시작 시간 감소 버튼 클릭
            print('검색 시작 시간 down')
            if self.SearchValue[3]-1 < 0:
                messagebox.showerror("에러발생", "검색 시작 시간은 음수가 될 수 없습니다.")
            else:
                self.SearchValue[3] -= 1
                self.history_produce_canvas.itemconfig(self.SearchValueViewText[3], text = self.SearchValue[3])
        
        if 1149 < x < 1216 and 966 < y < 1039:
            #검색 시작 시간 증가 버튼 클릭
            print('검색 시작 시간 up')
            if self.SearchValue[3]+1 >= self.SearchValue[4]:
                messagebox.showerror("에러발생", "검색 시작 시간은 검색 끝 시간보다 같거나 높아질 수 없습니다.")
            else:
                self.SearchValue[3] += 1
                self.history_produce_canvas.itemconfig(self.SearchValueViewText[3], text = self.SearchValue[3])

        if 1252 < x < 1326 and 966 < y < 1039:
            #검색 끝 시간 감소 버튼 클릭
            print('검색 끝 시간 down')
            if self.SearchValue[4]-1 <= self.SearchValue[3]:
                messagebox.showerror("에러발생", "검색 끝 시간은 검색 시작 시간보다 적어질 수 없습니다.")
            else:
                self.SearchValue[4] -= 1
                self.history_produce_canvas.itemconfig(self.SearchValueViewText[4], text = self.SearchValue[4])

        if 1509 < x < 1581 and 966 < y < 1039:
            #검색 끝 시간 증가 버튼 클릭
            print('검색 끝 시간 up')
            if self.SearchValue[4]+1 > 23:
                messagebox.showerror("에러발생", "검색 끝 시간은 23보다 클 수 없습니다.")
            else:
                self.SearchValue[4] += 1
                self.history_produce_canvas.itemconfig(self.SearchValueViewText[4], text = self.SearchValue[4])

        if 1615 < x < 1864 and 931 < y < 1030:
            #검색 버튼 클릭
            print('검색 버튼 클릭')
            #버튼 출력 변수 선언
            TotalSearchValue = ['', 0, 0, 0, 0]
            TotalSearchValue[0] = self.SearchValue[0]

            #함수 진행
            self.history_produce_log.delete(0, "end")
            startDateStr = None
            endDateStr = None
            if self.SearchValue[0] == 'ALL':
                self.SearchValue[0] = None
            if self.SearchValue[1] == 'None':
                self.SearchValue[1] = None
            if self.SearchValue[2] == 'None':
                self.SearchValue[2] = None

            if self.SearchValue[1] != None:
                startDateStr = f'{self.SearchValue[1]} {self.SearchValue[3]}:00:00'
            if self.SearchValue[2] != None:
                endDateStr = f'{self.SearchValue[2]} {self.SearchValue[4]}:00:00'

            returnLow = bearingartDB.readSql_View(self.SearchValue[0], startDateStr, endDateStr)

            listboxUpdateText = [None] * len(returnLow)
            for i, value in enumerate(returnLow):
                listboxUpdateText[i] = value[1].ljust(20, ' ')                                 #model name
                listboxUpdateText[i] += f'|    {str(value[2]).rjust(4," ")}     |'             #total
                listboxUpdateText[i] += f'    {str(value[3]).rjust(4," ")}     |'              #ok
                listboxUpdateText[i] += f'    {str(value[4]).rjust(4," ")}     |'              #normalng
                listboxUpdateText[i] += f'    {str(value[5]).rjust(4," ")}     |'              #criticalng
                listboxUpdateText[i] += f'    {value[7]}'                                      #date

                TotalSearchValue[1] += value[2]
                TotalSearchValue[2] += value[3]
                TotalSearchValue[3] += value[4]
                TotalSearchValue[4] += value[5]

            for i in range(len(listboxUpdateText)):
                self.history_produce_log.insert(0, f'{i+1}  \t'+listboxUpdateText[i])

            self.history_produce_log.insert(0, f'―――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――――')
            self.history_produce_log.insert(0, f'      모델명               전체 수량      양품 수량   일반 불량 수량 중대 불량 수량      생산 시간대')

            for i in range(5):
                self.history_produce_canvas.itemconfig(self.SearchViewText[i], text = TotalSearchValue[i])


        if 30 < x < 278 and 23 < y < 96:
            #제품 검색 화면으로 이동
            print('제품 검색 화면 전환')
            history_frame.tkraise()

    def on_select(self, event):
        w = event.widget
        idx = int(w.curselection()[0])
        value = w.get(idx)

        print(idx, value)

    def show_calendar(self, inputdata):
        from tkcalendar import Calendar

        def print_sel_start():
            split_text = cal.selection_get()
            test = split_text.strftime("%Y-%m-%d")
            # test = test.split(',')
            self.year = test[0]
            self.month = test[1]
            self.day = test[2]
            # input_date = self.year + "년 " + self.month + "월 " + self.day + "일"
            self.history_produce_canvas.itemconfig(self.SearchValueViewText[1], text=test)
            self.SearchValue[1] = test
            top.destroy()

        def print_sel_end():
            split_text = cal.selection_get()
            test = split_text.strftime("%Y-%m-%d")
            # test = test.split(',')
            self.year = test[0]
            self.month = test[1]
            self.day = test[2]
            # input_date = self.year + "년 " + self.month + "월 " + self.day + "일"
            self.history_produce_canvas.itemconfig(self.SearchValueViewText[2], text=test)
            self.SearchValue[2] = test
            top.destroy()

        top = tk.Toplevel(root)

        now = datetime.datetime.now()

        years = now.year
        months = now.month
        days = now.day

        cal = Calendar(top, font="Arial 14", selectmode="day", cursor="hand1", year=years, month=months, day=days,)
        cal.pack(fill="both", expand=True)
        if inputdata == "start":
            tk.Button(top, text="Select", command=print_sel_start).pack()
        elif inputdata == "end":
            tk.Button(top, text="Select", command=print_sel_end).pack()

class AdminManagementFrame(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.ManagementBgImage = ImageTk.PhotoImage(file=f"bg/{BgSetup}/management.png")
        self.ManagementChoiceBgImage = ImageTk.PhotoImage(file=f"bg/{BgSetup}/management_choice.png")
        self.ManagementLoadingImage = ImageTk.PhotoImage(file=f"bg/{BgSetup}/imageLoading.png")

        # 배출 바이패스 버튼
        self.bypassbtn_off = ImageTk.PhotoImage(file=f"bg/{BgSetup}/bypass_off.png")
        self.bypassbtn_on = ImageTk.PhotoImage(file=f"bg/{BgSetup}/bypass_on.png")
        self.bypass_config = configparser.ConfigParser()
        self.bypassINI = 'CountSave/State.INI'

        self.ManagementBtnText = [[[None,None],[None,None],[None,None]],[[None,None],[None,None],[None,None]],[[None,None],[None,None],[None,None]],[[None,None],[None,None],[None,None]]]
        self.ManagementBtnImg = [[[None,None],[None,None],[None,None]],[[None,None],[None,None],[None,None]],[[None,None],[None,None],[None,None]],[[None,None],[None,None],[None,None]]]
        self.ManagementTrigger = ['CAPORI','CAPBAD','CAPINS','BYPASS']
        self.state = [False, False, False]
        # loading gif moduls
        self.loadingFrame = []
        im = Image.open(f"bg/{BgSetup}/loading_main.gif")
        index = 0
        for frame in ImageSequence.Iterator(im):
            # np_frame = np.asarray(frame)
            frame = frame.convert('RGBA')
            frame = frame.resize((150, 150), Image.ANTIALIAS)
            frame_tk = ImageTk.PhotoImage(frame)
            
            self.loadingFrame.append(frame_tk)
            index += 1
        self.loadingProcessTrigger = None
        
        self.ManagementBtn_WH = [95,48]
        if CodeSetup == 'CONE':
            self.BtnTextLabel = [[['111','112'],['113','114'],['115','116']],[['111','112'],['113','114'],['115','116']],[['111','112'],['113','114'],['115','116']],[['111','112'],['113','114'],['115','116']]]
            self.ManagementBtnIndex = [[[69,155,'111'],[199,155,'112'],[69,221,'113'],[199,221,'114'],[69,287,'115'],[199,287,'116']],
                                    [[405,155,'111'],[535,155,'112'],[405,221,'113'],[535,221,'114'],[405,287,'115'],[535,287,'116']],
                                    [[69,440,'111'],[199,440,'112'],[69,506,'113'],[199,506,'114'],[69,572,'115'],[199,572,'116']],
                                    [[405,440,'111'],[535,440,'112'],[405,506,'113'],[535,506,'114'],[405,572,'115'],[535,572,'116']]]
            self.BtnState = [[[False,'111'],[False,'112'],[False,'113'],[False,'114'],[False,'115'],[False,'116']],
                            [[False,'111'],[False,'112'],[False,'113'],[False,'114'],[False,'115'],[False,'116']],
                            [[False,'111'],[False,'112'],[False,'113'],[False,'114'],[False,'115'],[False,'116']],
                            [[False,'111'],[False,'112'],[False,'113'],[False,'114'],[False,'115'],[False,'116']]]
        else:
            self.BtnTextLabel = [[['101','102'],['103','104'],['105',None]],[['101','102'],['103','104'],['105',None]],[['101','102'],['103','104'],['105',None]],[['101','102'],['103','104'],['105',None]]]
            self.ManagementBtnIndex = [[[69,155,'101'],[199,155,'102'],[69,221,'103'],[199,221,'104'],[69,287,'105']],
                                    [[405,155,'101'],[535,155,'102'],[405,221,'103'],[535,221,'104'],[405,287,'105']],
                                    [[69,440,'101'],[199,440,'102'],[69,506,'103'],[199,506,'104'],[69,572,'105']],
                                    [[405,440,'101'],[535,440,'102'],[405,506,'103'],[535,506,'104'],[405,572,'105']]]
            self.BtnState = [[[False,'101'],[False,'102'],[False,'103'],[False,'104'],[False,'105']],
                            [[False,'101'],[False,'102'],[False,'103'],[False,'104'],[False,'105']],
                            [[False,'101'],[False,'102'],[False,'103'],[False,'104'],[False,'105']],
                            [[False,'101'],[False,'102'],[False,'103'],[False,'104'],[False,'105']]]

        self.create_widgets()
    
    def create_widgets(self):
        self.grid(row=0, column=0)
        self.management_canvas = tk.Canvas(self, width=700, height=800)
        self.managementBg = self.management_canvas.create_image(0, 0, image=self.ManagementBgImage, anchor="nw")
        for i in range(len(self.ManagementBtnText[0])):
            for x in range(len(self.ManagementBtnText[0][i])):
                try:
                    self.ManagementBtnImg[0][i][x] = self.management_canvas.create_image(self.ManagementBtnIndex[0][(i*2)+x][0],self.ManagementBtnIndex[0][(i*2)+x][1], image = self.ManagementChoiceBgImage, state = 'hidden', anchor = 'nw')
                except:
                    pass
                self.ManagementBtnText[0][i][x] = self.management_canvas.create_text(117+(130*x), 179+(66*i), text=self.BtnTextLabel[0][i][x], font=("gothic", 15, "bold"), fill="white", anchor="center")

        for i in range(len(self.ManagementBtnText[1])):
            for x in range(len(self.ManagementBtnText[1][i])):
                try:
                    self.ManagementBtnImg[1][i][x] = self.management_canvas.create_image(self.ManagementBtnIndex[1][(i*2)+x][0],self.ManagementBtnIndex[1][(i*2)+x][1], image = self.ManagementChoiceBgImage, state = 'hidden', anchor = 'nw')
                except:
                    pass
                self.ManagementBtnText[1][i][x] = self.management_canvas.create_text(452+(130*x), 179+(66*i), text=self.BtnTextLabel[1][i][x], font=("gothic", 15, "bold"), fill="white", anchor="center")

        for i in range(len(self.ManagementBtnText[2])):
            for x in range(len(self.ManagementBtnText[2][i])):
                try:
                    self.ManagementBtnImg[2][i][x] = self.management_canvas.create_image(self.ManagementBtnIndex[2][(i*2)+x][0],self.ManagementBtnIndex[2][(i*2)+x][1], image = self.ManagementChoiceBgImage, state = 'hidden', anchor = 'nw')
                except:
                    pass
                self.ManagementBtnText[2][i][x] = self.management_canvas.create_text(117+(130*x), 464+(66*i), text=self.BtnTextLabel[2][i][x], font=("gothic", 15, "bold"), fill="white", anchor="center")

        for i in range(len(self.ManagementBtnText[3])):
            for x in range(len(self.ManagementBtnText[3][i])):
                try:
                    self.ManagementBtnImg[3][i][x] = self.management_canvas.create_image(self.ManagementBtnIndex[3][(i*2)+x][0],self.ManagementBtnIndex[3][(i*2)+x][1], image = self.ManagementChoiceBgImage, state = 'hidden', anchor = 'nw')
                except:
                    pass
                self.ManagementBtnText[3][i][x] = self.management_canvas.create_text(452+(130*x), 464+(66*i), text=self.BtnTextLabel[3][i][x], font=("gothic", 15, "bold"), fill="white", anchor="center")

        # 배출 바이패스 정보 가져오기
        self.bypass_state_config = configparser.ConfigParser()
        self.initbypassINI = 'CountSave/State.INI'

        self.bypass_state_config.read('CountSave/State.INI', encoding='euc-kr')
        self.bypass_state = self.bypass_state_config['BYPASS']['bypass']
        print(f'▶ 현재 바이패스 상태 : {self.bypass_state}')

        if self.bypass_state == 'on' :
            main_frame.bypass_mode = True
            self.pass_mode = self.management_canvas.create_image(483, 32, image=self.bypassbtn_on, anchor="nw")
        elif self.bypass_state == 'off' :
            main_frame.bypass_mode = False
            self.pass_mode = self.management_canvas.create_image(483, 32, image=self.bypassbtn_off, anchor="nw")

        self.imageLoading = self.management_canvas.create_image(350,400, image = self.ManagementLoadingImage, anchor = 'center', state = 'hidden')
        self.loadingGifImage = self.management_canvas.create_image(350,500, image = '', anchor = 'center', state = 'hidden')
        self.management_canvas.bind("<Button-1>", self.management_btn)
        self.management_canvas.pack()

    def make_safe_dir(self, dir):
        if not os.path.exists(dir):
            os.makedirs(dir)

    def git_clone(self, git_url):
        target_dir = os.getcwd()  # 현재 경로
        self.make_safe_dir(target_dir)
        git.Git(target_dir).clone(git_url) # 저장될 공간

    def state_set(self, state):
        self.state_config = configparser.ConfigParser()
        self.stateINI = 'CountSave/State.INI'

        if os.path.isfile(self.stateINI) == False: 
            self.state_config['STATE'] = 0
                
            # INI 파일 생성 
            with open(self.stateINI, 'w') as f:
                self.state_config.write(f)

        self.state_config.read('CountSave/State.INI', encoding='euc-kr')

        self.state_config.set('STATE', 'index', state)

        with open(self.stateINI, 'w') as f:
            self.state_config.write(f)

    def management_btn(self, event):
        global ConnectClientCheck
        x = event.x
        y = event.y

        print(x, y)
        for j in range(4):
            for i in range(len(self.ManagementBtnIndex[j])):
                if self.ManagementBtnIndex[j][i][0] < x < self.ManagementBtnIndex[j][i][0]+self.ManagementBtn_WH[0] and self.ManagementBtnIndex[j][i][1] < y < self.ManagementBtnIndex[j][i][1]+self.ManagementBtn_WH[1]:
                    # print(self.ManagementBtnIndex[j][i][2])
                    # print(j, i)
                    if j == 3:
                        if STH.ManualMode == True:
                            result = messagebox.askyesno("최종 확인", "해당 PC를 바이패스 시키겠습니까?\n바이패스시 PC가 종료됩니다.")
                            if result == True:
                                if self.BtnState[j][i][0] == True:
                                    messagebox.showinfo("설정 완료","바이패스가 해제됩니다.\n시스템을 재가동 해주세요.")
                                else:
                                    messagebox.showinfo("설정 완료","바이패스가 설정됩니다.\n해당 PC가 종료됩니다.")
                            else:
                                main_frame.update_signal(f'Not ManualMode / Change to ManualMode Please')
                                return
                        else:
                            main_frame.update_signal(f'Not ManualMode / Change to ManualMode Please')
                            return
                        
                    self.BtnState[j][i][0] = not self.BtnState[j][i][0]
                    self.management_canvas.itemconfig(self.ManagementBtnImg[j][int(i/2)][int(i%2)], state = 'normal' if self.BtnState[j][i][0] else 'hidden')
                    sendMsg = f'{self.ManagementTrigger[j]}ON{self.BtnState[j][i][1]}' if self.BtnState[j][i][0] else f'{self.ManagementTrigger[j]}OFF{self.BtnState[j][i][1]}'
                    if 'CAPORI' in sendMsg:
                        if 'ON' in sendMsg:
                            main_frame.update_signal(f'Origin Image Capture On - {self.BtnState[j][i][1]}')
                        else:
                            main_frame.update_signal(f'Origin Image Capture Off - {self.BtnState[j][i][1]}')
                    elif 'CAPBAD' in sendMsg:
                        if 'ON' in sendMsg:
                            main_frame.update_signal(f'Bad Image Capture On - {self.BtnState[j][i][1]}')
                        else:
                            main_frame.update_signal(f'Bad Image Capture Off - {self.BtnState[j][i][1]}')
                    elif 'CAPINS' in sendMsg:
                        if 'ON' in sendMsg:
                            main_frame.update_signal(f'Ins Image Capture On - {self.BtnState[j][i][1]}')
                        else:
                            main_frame.update_signal(f'Ins Image Capture Off - {self.BtnState[j][i][1]}')
                    elif 'BYPASS' in sendMsg:
                        if 'ON' in sendMsg:
                            main_frame.update_signal(f'Bypass On - {self.BtnState[j][i][1]}')
                        else:
                            main_frame.update_signal(f'Bypass Off - {self.BtnState[j][i][1]}')
                        ConnectClientCheck = 0
                        for i in range(len(self.BtnState[3])):
                            if self.BtnState[3][i][0] == True:
                                pass
                            else:
                                ConnectClientCheck += 1
                        TSD.connect_count = ConnectClientCheck
                    print(f'SendMsg - {sendMsg}')
                    ######################################데이터 송신 필요
                    TSD.sendAllClient(sendMsg)
                    self.management_State_Save()

        # 상태 표시 선택
        if 37 < x < 37+113 and 13 < y < 13+23:
            print('▶ 상태 표시 1. 비전 검사 적용 완료')
            self.state_set('1')
            main_frame.updatestate('1')

        if 37 < x < 37+113 and 42 < y < 42+23:
            print('▶ 상태 표시 2. 비전 검사 미적용 (세팅중)')
            self.state_set('2')
            main_frame.updatestate('2')

        if 37 < x < 37+113 and 71 < y < 71+23:
            print('▶ 상태 표시 3. 유효성 평가 대상')
            self.state_set('3')
            main_frame.updatestate('3')

        if 71 < x < 163 and 714 < y < 764:
            # self.management_State_SendAll()
            model_frame.tkraise()
            print('USB 모델 업로드')

        if 406 < x < 502 and 714 < y < 763:
            result = messagebox.askyesno("최종 확인", "검증 파라메터를 가져오시겠습니까?")
            if result == True:
                print('검증 파라메터 가져오기')
                TSD.sendAllClient('PARAMETER')
            else:
                print('취소하셨습니다.')
                # main_frame.management_mode = False

        if 200 < x < 296 and 714 < y < 762:
            print('촬영사진 가져오기')
            threading.Thread(target=self.imageLoadingThread, daemon= True).start()
            # result = transfer.stealdate()
        
         # 배출 바이패스
        if 484 < x < 484 + 83 and 33 < y < 33 + 25 :
            if main_frame.bypass_mode == False :
                bypass_result = messagebox.askyesno("최종 확인", "배출 바이패스 하시겠습니까?")
                if bypass_result :
                    print('배출 바이패스 설정')
                    main_frame.bypass_mode = True
                    self.management_canvas.itemconfig(self.pass_mode, image=self.bypassbtn_on, state = 'normal')

                    self.bypass_config.read('CountSave/State.INI', encoding='euc-kr')
                    self.bypass_config.set('BYPASS', 'bypass', 'on')
                    with open(self.bypassINI, 'w') as f:
                        self.bypass_config.write(f)
                else :
                    print('배출 바이패스 취소')
            elif main_frame.bypass_mode == True :
                bypass_result = messagebox.askyesno("최종 확인", "배출 바이패스 해제하시겠습니까?")
                if bypass_result :
                    print('배출 바이패스 해제')
                    main_frame.bypass_mode = False
                    self.management_canvas.itemconfig(self.pass_mode, image=self.bypassbtn_off, state = 'normal')

                    self.bypass_config.read('CountSave/State.INI', encoding='euc-kr')
                    self.bypass_config.set('BYPASS', 'bypass', 'off')
                    with open(self.bypassINI, 'w') as f:
                        self.bypass_config.write(f)

                else :
                    print('배출 바이패스 취소')

        if 483 < x < 483 + 85 and 63 < y < 63 + 26: # 프로그램 재실행 버튼
            print(f'프로그램 재실행 선택')

            result = messagebox.askyesno("최종 확인", "프로그램 재실행하시겠습니까?")
            if result == True:
                logger.info('[Signal] Admin program restart order')
                main_frame.update_signal("[ ★ ] 프로그램 재실행")
                CountDataSave()
                TSD.sendAllClient("REBOOT")
                STH.retry_procedure()
            else :
                print(f'프로그램 재실행 취소')

        if 579 < x < 579 + 85 and 63 < y < 63 + 26: # 코드 업데이트 버튼(메인, 클라이언트 전부 업데이트)
            print(f'코드 업데이트 선택')

            result = messagebox.askyesno("최종 확인", "코드를 업데이트하시겠습니까?")
            if result == True:
                logger.info('[Signal] Admin code update order')
                main_frame.update_signal("[ ★ ] 코드 업데이트")
                TSD.sendAllClient("UPDATE")

                repo_name = f'Main_{CodeSetup}'

                try :
                    if os.path.exists('Main_.py'):         
                        os.remove("Main_.py")     # 기존 백업 메인 코드 삭제

                    if os.path.exists('Main.py'): # 기존 메인 코드 백업         
                        os.rename('Main.py', 'Main_.py') 
                    git_url = f'https://github.com/KRThor/{repo_name}.git' # 코드 업데이트할 깃허브 주소
                    self.git_clone(git_url) # 다운로드
                    print("[★] 코드 다운로드")
                    time.sleep(0.2)
                    os.rename(f'{repo_name}/Main.py', 'Main.py') # 다운로드 받은 코드 경로 수정
                    time.sleep(0.2)
                    os.system(f'rmdir /S /Q {repo_name}') # 다운로드 받은 파일 삭제 (파일 있으면 충돌로 에러나서 업데이트 후 삭제)
                    print("[★] 폴더 제거")
                    time.sleep(0.2)
                    os.system('python compile_M.py')
                    print("[★] 코드 컴파일")
                    print('[★] 코드 업데이트 성공')
                    logger.info(f"[Notice] 코드 업데이트 성공")
                except :
                    print("[★] 코드 다운로드 실패")
                    logger.info(f"[Notice] 코드 업데이트 실패")
                    print(traceback.format_exc())

            else :
                print(f'코드 업데이트 취소')

    def imageLoadingThread(self):
        try:
            self.management_canvas.itemconfig(self.imageLoading, state = 'normal')
            self.loadingProcessTrigger = True
            threading.Thread(target=self.loadingProcessGIF, daemon=True).start()
            result = transfer.stealdate()
            print(f"[★] {result}")     # 성공"시 : "SUCCESS" // 실패시 : "FAIL"
            if result == "SUCCESS":
                messagebox.showinfo("완료", "검증 이미지 로딩을 완료하였습니다.")
            else:
                messagebox.showerror("에러발생", "검증 이미지를 로딩하지 못하였습니다.")
            self.loadingProcessTrigger = False
            self.management_canvas.itemconfig(self.imageLoading, state = 'hidden')
        except:
            print(traceback.format_exc())

    def loadingProcessGIF(self):
        while True:
            for i in range(len(self.loadingFrame)):
                self.management_canvas.itemconfig(self.loadingGifImage, image=self.loadingFrame[i], state = 'normal')
                time.sleep(0.1)
                if self.loadingProcessTrigger == False:
                    self.management_canvas.itemconfig(self.loadingGifImage, image='', state = 'hidden')
                    return

    def management_State_Save(self):
        WriteJsonData = dict()
        WriteJsonData["BtnState"] = self.BtnState
        if CodeSetup == 'CONE':
            with open('CheckValue/management.json', 'w', encoding='utf-8') as make_file:
                json.dump(WriteJsonData, make_file, indent="\t")
        else:
            with open('CheckValue/management.json', 'w', encoding='utf-8') as make_file:
                json.dump(WriteJsonData, make_file, indent="\t")

    def management_State_Load(self):
        global ConnectClientCheck
        if CodeSetup == 'CONE':
            if os.path.isfile('CheckValue/management.json'):
                with open('CheckValue/management.json', 'r') as read_file:
                    ReadJsonData = json.load(read_file)
                self.BtnState = ReadJsonData['BtnState'].copy()
                ConnectClientCheck = 0
                for i in range(len(self.BtnState[3])):
                    if self.BtnState[3][i][0] == True:
                        pass
                    else:
                        ConnectClientCheck += 1
                print(ConnectClientCheck)
            else:
                pass
        else:
            if os.path.isfile('CheckValue/management.json'):
                with open('CheckValue/management.json', 'r') as read_file:
                    ReadJsonData = json.load(read_file)
                self.BtnState = ReadJsonData['BtnState'].copy()
                ConnectClientCheck = 0
                for i in range(len(self.BtnState[3])):
                    if self.BtnState[3][i][0] == True:
                        pass
                    else:
                        ConnectClientCheck += 1
                print(ConnectClientCheck)
            else:
                pass

    def management_State_SendAll(self):
        for j in range(4):
            for i in range(len(self.ManagementBtnIndex[j])):
                self.management_canvas.itemconfig(self.ManagementBtnImg[j][int(i/2)][int(i%2)], state = 'normal' if self.BtnState[j][i][0] else 'hidden')
                sendMsg = f'{self.ManagementTrigger[j]}ON{self.BtnState[j][i][1]}' if self.BtnState[j][i][0] else f'{self.ManagementTrigger[j]}OFF{self.BtnState[j][i][1]}'
                print(sendMsg)
                TSD.sendAllClient(sendMsg)


class ModelManageFrame(tk.Frame):
    def __init__(self, master=None, db_inst=''):
        super().__init__(master)
        self.master = master

        # image sources
        self.model_bg_img = ImageTk.PhotoImage(file = f'bg/{BgSetup}/bg_managing_model.png')
        self.confirm_img = ImageTk.PhotoImage(file = f'bg/{BgSetup}/btn_confirm.png')
        self.update_img = ImageTk.PhotoImage(file = f'bg/{BgSetup}/btn_update.png')
        self.change_img = ImageTk.PhotoImage(file = f'bg/{BgSetup}/btn_change.png')
        
        self.db = db_inst

        # image sources
        self.model_bg_img = ImageTk.PhotoImage(file = f'bg/{BgSetup}/bg_managing_model.png')
        self.confirm_img = ImageTk.PhotoImage(file = f'bg/{BgSetup}/btn_confirm.png')
        self.update_img = ImageTk.PhotoImage(file = f'bg/{BgSetup}/btn_update.png')
        self.change_img = ImageTk.PhotoImage(file = f'bg/{BgSetup}/btn_change.png')

        # loading gif
        self.loading_img_frame = []
        loading_gif = Image.open(f"bg/{BgSetup}/loading.gif")
        index = 0
        for frame in ImageSequence.Iterator(loading_gif):
            np_frame = np.asarray(frame)
            frame_tk = ImageTk.PhotoImage(frame)
            self.loading_img_frame.append(frame_tk)
            # print('\n\n', len(self.loading_img_frame)) #24
            index += 1
        
        # variables
        self.model_update_mode = False
        self.model_change_mode = False

        self.create_widgets()

    def create_widgets(self):
        self.grid(row = 0, column = 0)
        self.model_canvas = tk.Canvas(self, width = 600, height = 750)
        self.model_bg_image = self.model_canvas.create_image(0, 0, image = self.model_bg_img, anchor = 'nw')
        self.confirm_img_on = self.model_canvas.create_image(298, 674, image = self.confirm_img, anchor = 'nw', state = 'hidden')
        self.update_img_on = self.model_canvas.create_image(35, 91, image = self.update_img, anchor = 'nw', state = 'hidden')
        self.change_img_on = self.model_canvas.create_image(169, 91, image = self.change_img, anchor = 'nw', state = 'hidden')

        # model list up
        # self.model_list = tk.Listbox(self.model_canvas, fg = 'black',  font=('gothic', 10), selectmode='extended')
        self.model_list = tk.Listbox(self.model_canvas, fg = 'black',  font=('gothic', 10) )
        self.model_list.place(x=68, y=195, width=440, height=425)

        self.model_list_scrollbar_y = tk.Scrollbar(self.model_canvas, orient="vertical")
        self.model_list_scrollbar_y.config(command=self.model_list.yview)
        self.model_list_scrollbar_y.place(x=68+440, y=195, width=15, height=425)
        self.model_list.config(yscrollcommand=self.model_list_scrollbar_y.set)

        self.model_list_scrollbar_x = tk.Scrollbar(self.model_canvas, orient="horizontal")
        self.model_list_scrollbar_x.config(command=self.model_list.xview)
        self.model_list_scrollbar_x.place(x=68, y=195+425, width=440+15, height=15)
        self.model_list.config(xscrollcommand=self.model_list_scrollbar_x.set)  
        
        self.model_list.place_forget()
        self.model_list_scrollbar_y.place_forget()
        self.model_list_scrollbar_x.place_forget()

        # databse list up
        # self.db_list = tk.Listbox(self.model_canvas, fg = 'black', font=('gothic', 10), selectmode='extended')
        self.db_list = tk.Listbox(self.model_canvas, fg = 'black', font=('gothic', 10))
        self.db_list.place(x=68, y=195, width=450, height=435)

        self.db_list_scrollbar_y = tk.Scrollbar(self.model_canvas, orient="vertical")
        self.db_list_scrollbar_y.config(command=self.db_list.yview)
        self.db_list_scrollbar_y.place(x=68+440, y=195, width=15, height=425)
        self.db_list.config(yscrollcommand=self.db_list_scrollbar_y.set)

        self.db_list_scrollbar_x = tk.Scrollbar(self.model_canvas, orient="horizontal")
        self.db_list_scrollbar_x.config(command=self.db_list.xview)
        self.db_list_scrollbar_x.place(x=68, y=195+425, width=440+15, height=15)
        self.db_list.config(xscrollcommand=self.db_list_scrollbar_x.set) 

        self.db_list.place_forget()
        self.db_list_scrollbar_y.place_forget()
        self.db_list_scrollbar_x.place_forget()

        font_main = ("Calibri", 14)
        # get product
        with open('CheckValue/product_info.json') as json_file:
            json_data = json.load(json_file)
        self.common_model_name = json_data[LINE][CodeSetup][0]

        values_product = ['ALL']
        values_product = values_product + json_data[LINE][CodeSetup][1]            
        self.select_product = ttk.Combobox(self, width = 8, justify = 'center', values= values_product, font= font_main, state = "readonly")
        self.select_product.current(0)
        self.show_product = self.model_canvas.create_window(310, 110, anchor='nw', window='')

        # get cam_name : 내륜 111~116 / 외륜 101~105
        if CodeSetup == 'CUP':
            values_cam = ['ALL']
            values_cam = values_cam + list(range(101, 106))
        else:
            values_cam = ['ALL']
            values_cam = values_cam + list(range(111, 117))
        self.select_cam = ttk.Combobox(self, width = 8, justify = 'center', values= values_cam, font= font_main, state = "readonly")
        self.select_cam.current(0)
        self.show_cam = self.model_canvas.create_window(415, 110, anchor='nw', window='')

        self.search_button = tk.Button(self, text=' V ', fg = 'black', command=self.get_db_values)
        self.search_button['font'] = ("Calibri", 10, 'bold')
        self.search_button.place_forget()

        # loading imgs
        self.loading_on = self.model_canvas.create_image(300, 375, image=self.loading_img_frame[0], anchor="center", state="hidden")

        self.model_canvas.bind('<Button-1>', self.model_btn)
        self.model_canvas.pack()      

   
    def animate_loading(self, mode):
        if mode == 'UPLOAD':
            self.model_list.place_forget()
            self.model_list_scrollbar_y.place_forget()
            self.model_list_scrollbar_x.place_forget()
            self.master.update()

            while True:
                if transfer.done_check == 'LOADING':
                    for f_num in range(len(self.loading_img_frame)):
                        time.sleep(0.1)
                        self.model_canvas.itemconfig(self.loading_on, image=self.loading_img_frame[f_num], state = 'normal')
                        self.master.update()
                else:
                    self.model_canvas.itemconfig(self.loading_on, image='', state = 'hidden')
                    self.master.update()
                    if transfer.done_check == 'SUCCESS':
                        self.db.update_sql()
                        answer = messagebox.showinfo('INFO', '모델 업데이트 완료')
                        self.model_canvas.itemconfig(self.loading_on, image='', state = 'hidden')
                    elif transfer.done_check == 'FAIL':
                        answer = messagebox.showerror('ERROR', '모델 업데이트 에러')
                        self.model_canvas.itemconfig(self.loading_on, image='', state = 'hidden')
                    break

        elif mode == 'CHANGE':
            self.db_list.place_forget()
            self.db_list_scrollbar_y.place_forget()
            self.db_list_scrollbar_x.place_forget()
            self.master.update()

            while True:
                if transfer.modelbackup_result == 'LOADING':
                    for f_num in range(len(self.loading_img_frame)):
                        time.sleep(0.1)
                        self.model_canvas.itemconfig(self.loading_on, image=self.loading_img_frame[f_num], state = 'normal')
                        self.master.update()
                else:
                    self.model_canvas.itemconfig(self.loading_on, image='', state = 'hidden')
                    self.master.update()
                    if transfer.modelbackup_result == 'SUCCESS':
                        answer = messagebox.showinfo('INFO', '모델 변경 완료')
                        self.model_canvas.itemconfig(self.loading_on, image='', state = 'hidden')
                    elif transfer.modelbackup_result == 'FAIL':
                        answer = messagebox.showerror('ERROR', '모델 변경 에러')
                        self.model_canvas.itemconfig(self.loading_on, image='', state = 'hidden')
                    break
                # time check > break
     

    def showup_model_list(self):
        try:
            self.model_list.delete(0, 'end')
            self.db_list.delete(0, 'end')
        except:
            pass

        self.model_canvas.itemconfig(self.show_product, window=self.select_product, state = 'hidden')
        self.model_canvas.itemconfig(self.show_cam, window=self.select_cam, state = 'hidden')
        self.db_list.place_forget()
        self.db_list_scrollbar_x.place_forget()
        self.model_list_scrollbar_x.place_forget()
        self.search_button.place_forget()

        self.model_list.place(x=68, y=195, width=440, height=425)
        self.model_list_scrollbar_y.place(x=68+440, y=195, width=15, height=425)
        self.model_list_scrollbar_x.place(x=68, y=195+425, width=440+15, height=15)

        # json 파싱 후 해당 라인 업로드 정보 이중리스트로 저장
        self.modelinfo_list = jsonparser.USBjsonparser() 
        # print(self.modelinfo_list,'\n\n\n')

        self.update_model_list = [None]*len(self.modelinfo_list)
 
        self.model_list.insert(0, '[ 구분 ]  [ 라인 ]        [ 기종 ]        [ 카메라 ]  [ 모델 LOT ]    [ 코멘트 ]')
        self.model_list.insert(1, ' ')

        for idx, row in enumerate(self.modelinfo_list[::-1]):
            # print('\n', row)
            self.update_model_list[idx] = f"   {idx+1:02d}      " # idx
            self.update_model_list[idx] += "%-12s" % LINE     # line

            # ip, username, passwd, modelname, camnum, modelpath, modelversion, comment
            self.update_model_list[idx] += "%-20s" % row[3]  # MODEL_NAME
            self.update_model_list[idx] += "%-12s" % row[4]  # CAM_NUM
            self.update_model_list[idx] += "%-12s" % row[6]  # modelversion
            self.update_model_list[idx] += "%-20s" % row[7]  # comment

            self.model_list.insert(idx+2, self.update_model_list[idx])


    def get_db_values(self):
        # select catergory
        self.select_product_now = self.select_product.get()
        self.select_cam_now = self.select_cam.get()
        # print('1: ', self.select_product_now, '2: ', self.select_cam_now)
        
        # show db list
        try:
            self.db_list.delete(0, 'end')
        except:
            pass
        
        self.showup_db_list()
        
        # show confirm button
        self.model_canvas.itemconfig(self.confirm_img_on, state = 'normal')
        

    def get_db_list(self):
        try:
            self.model_list.delete(0, 'end')
            self.db_list.delete(0, 'end')
        except:
            pass

        self.model_canvas.itemconfig(self.show_product, window=self.select_product, state= 'normal')
        self.model_canvas.itemconfig(self.show_cam, window=self.select_cam, state= 'normal')
        self.search_button.place(x=525, y=112)	


    def showup_db_list(self):
        self.model_list.place_forget()
        self.model_list_scrollbar_y.place_forget()
        self.model_list_scrollbar_x.place_forget() 

        self.db_list.place(x=68, y=195, width=450, height=435)
        self.db_list_scrollbar_y.place(x=68+440, y=195, width=15, height=425)
        self.db_list_scrollbar_x.place(x=68, y=195+425, width=440+15, height=15)

        # read sql
        self.sql_values = self.db.read_sql(Product=self.select_product_now, Cam=self.select_cam_now)
        self.update_db_list = [None]*len(self.sql_values)

        self.db_list.insert(0, '[ 구분 ]     [ 업데이트날짜 ]      [ 라인 ]         [ 기종 ]        [ 카메라 ] [ 모델 LOT ]    [ 코멘트 ]')
        self.db_list.insert(1, ' ')

        for idx, row in enumerate(self.sql_values[::-1]):
            # print('\n', row)
            self.update_db_list[idx] = f"   {idx+1:02d}      " #idx
            self.update_db_list[idx] += f"{row[6]}      " #date
            self.update_db_list[idx] += "%-12s" % row[1] # line
            self.update_db_list[idx] += "%-20s" % row[2] # product name
            self.update_db_list[idx] += "%-12s" % row[3] # camera
            self.update_db_list[idx] += "%-12s" % row[4] # model lot
            self.update_db_list[idx] += "%-20s" % row[5] # comment
            
            self.db_list.insert(idx+2, self.update_db_list[idx])
        # print('showup_db_list')


    def model_btn(self, event):
        x, y = event.x, event.y
        print(f'model btn clicked! x: {x}, y: {y}')

        # 모델 업데이트 버튼
        if 40 < x < (40+115) and 95 < y < (95+50):
            self.model_canvas.itemconfig(self.confirm_img_on, state = 'normal')
            self.model_canvas.itemconfig(self.update_img_on, state = 'normal')
            self.model_canvas.itemconfig(self.change_img_on, state = 'hidden')

            self.model_update_mode = True
            self.model_change_mode = False  

            try:
                self.db_list.delete(0, 'end')
            except:
                pass

            self.showup_model_list()
            # print('모델 업데이트')
           
        # 모델 변경 버튼
        if 178 < x < (178+115) and 95 < y < (95+50):
            self.model_canvas.itemconfig(self.change_img_on, state = 'normal')
            self.model_canvas.itemconfig(self.update_img_on, state = 'hidden')

            self.model_update_mode = False
            self.model_change_mode = True  

            self.get_db_list() 
            # print('모델 변경')

        # 적용 버튼
        if 300 < x < (300+115) and 680 < y < (680+50):
            # 모델 업데이트 적용
            if self.model_update_mode:
                self.model_canvas.itemconfig(self.update_img, state = 'hidden')
                self.model_canvas.itemconfig(self.change_img, state = 'hidden')
                selected_1 = self.model_list.curselection()
                # print(selected_1)
                if selected_1 == ():
                    answer = messagebox.showinfo('INFO', '모델을 선택해주세요')
                else:
                    # loading gif
                    transfer.done_check = 'LOADING'
                    Thread(target = self.animate_loading, args=['UPLOAD'], daemon=True).start()
                    
                    for idx_1 in selected_1[::-1]:
                        # 다중 선택 값 가져오기
                        selected_list_1 = self.model_list.get(idx_1)
                        # 리스트 내 공백 제거
                        values_1 = selected_list_1.split('  ')
                        values_1 = [s1 for s1 in values_1 if s1]       
                        print('\n\n\n', values_1)               
                        
                        # 모델 업로드
                        ip = self.modelinfo_list[::-1][idx_1-2][0]
                        username = self.modelinfo_list[::-1][idx_1-2][1]
                        passwd = self.modelinfo_list[::-1][idx_1-2][2]
                        modelname = self.modelinfo_list[::-1][idx_1-2][3] # ST32451
                        camnum = self.modelinfo_list[::-1][idx_1-2][4] # CAM101
                        modelpath = self.modelinfo_list[::-1][idx_1-2][5]
                        modelversion = self.modelinfo_list[::-1][idx_1-2][6]
                        comment = self.modelinfo_list[::-1][idx_1-2][7]

                        #################################################################### update 22.05.10
                        # 형번 확인
                        DTT.product_infoJson_Load()
                        if self.common_model_name != []:
                            for common_model_name_value in self.common_model_name:
                                if modelname in common_model_name_value:
                                    update_modelname = common_model_name_value
                                else:
                                    update_modelname = [modelname]
                        else:
                            update_modelname = [modelname]
  
                        print('[INFO] update_model: ', update_modelname)

                        Thread(target = transfer.modelupload, args=[ip, username, passwd, update_modelname, camnum, modelpath, modelversion, comment], daemon=True).start()
                        # print(f"[★] {transfer.done_check}")  # 성공시 : "SUCCESS" // 실패시 : "FAIL"

                        # DB 입력
                        self.db.update_data[1] = values_1[1].strip() # line
                        self.db.update_data[2] = values_1[2].strip() # product_name
                        self.db.update_data[3] = values_1[3].strip() # cam_name
                        self.db.update_data[4] = values_1[4].strip() # model_lot
                        self.db.update_data[5] = values_1[5].strip() # notes 
                        ####################################################################                                                   

            # 모델 변경 적용
            if self.model_change_mode:
                # print('0. 모델 변경 적용 클릭')
                self.model_canvas.itemconfig(self.update_img, state = 'hidden')
                self.model_canvas.itemconfig(self.change_img, state = 'hidden')

                selected_2 = self.db_list.curselection()
                if selected_2 == ():
                    answer = messagebox.showinfo('INFO', '모델을 선택해주세요')
                else:
                    # loading gif
                    transfer.modelbackup_result = 'LOADING'
                    Thread(target = self.animate_loading, args=['CHANGE'], daemon=True).start()
                    
                    for idx_2 in selected_2[::-1]:                        
                        # 다중 선택 값 가져오기
                        selected_list_2 = self.db_list.get(idx_2)
                        # print('\nselected_list_2', selected_list_2)
                        # 리스트 나누기
                        values_2 = selected_list_2.split('  ')
                        str_match_2 = [s2 for s2 in values_2 if s2]
                        # print('\n\nstr_match_2', str_match_2)

                        modelname = str_match_2[3].strip()
                        camnum = str_match_2[4].strip()
                        modelversion =  str_match_2[5].strip()

                        # print('\n\ncamnum', camnum)
                        ip, username, passwd = jsonparser.getUSERinfo(camnum)
                        
                        #################################################################### update 22.05.10
                        # 형번 확인
                        DTT.product_infoJson_Load()
                        if self.common_model_name != []:
                            for common_model_name_value in self.common_model_name:
                                if modelname in common_model_name_value:
                                    change_modelname = common_model_name_value
                                else:
                                    change_modelname = [modelname]
                        else:
                            change_modelname = [modelname]                                    
                        ####################################################################   
                             
                        print('[INFO] change_model: ', change_modelname)                           
                        Thread(target = transfer.modelbackup, args=[ip, username, passwd, change_modelname, camnum, modelversion], daemon=True).start()
                        # print(f"[★] {transfer.modelbackup_result}")      # 성공시 : "SUCCESS" // 실패시 : "FAIL"

        # 취소 버튼
        if 440 < x < (440+115) and 680 < y < (680+50):
            # init
            self.model_update_mode = False
            self.model_change_mode = False

            self.model_list.delete(0, 'end')
            self.db_list.delete(0, 'end')

            self.model_canvas.itemconfig(self.confirm_img, state = 'hidden')
            self.model_canvas.itemconfig(self.update_img, state = 'hidden')
            self.model_canvas.itemconfig(self.change_img, state = 'hidden')

            main_frame.tkraise()

  
            
class TransferSocketData(threading.Thread):
    def __init__(self):
        threading.Thread.__init__(self)

        # HOST = '192.168.222.114'
        # HOST = '192.168.43.149'
        PORT = 9999
        # 소켓 객체를 생성합니다.
        # 주소 체계(address family)로 IPv4, 소켓 타입으로 TCP 사용합니다.
        self.server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        # 포트 사용중이라 연결할 수 없다는
        # WinError 10048 에러 해결를 위해 필요합니다.
        self.server_socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        # HOST는 hostname, ip address, 빈 문자열 ""이 될 수 있습니다.
        # 빈 문자열이면 모든 네트워크 인터페이스로부터의 접속을 허용합니다.
        # PORT는 1-65535 사이의 숫자를 사용할 수 있습니다.
        self.server_socket.bind((MAINHOST, PORT))
        self.server_socket.listen()

        self.allClient = []

        self.connect_count = 0
        self.connect_check = False
        self.allClientAddr = []

        self.cilentIP = list(None for i in range(5))

        print("Server Start")

    def recvall(self, sock, count):
        buf = b""
        while count:
            newbuf = sock.recv(count)
            if not newbuf:
                return None
            buf += newbuf
            count -= len(newbuf)
        return buf

    def threaded(self, client_socket, addr):
        print("Connected by :", addr[0], ":", addr[1])
        logger.info(f"[Notice] Connected by : {addr[0]} : {addr[1]}")
        my_client = 0

        while True:
            try:
                data = self.recvall(client_socket, 20)

                if data == None:
                    print("Disconnected by " + addr[0], ":", addr[1], "(no data)")
                    break

                try:
                    recvData = data.decode().strip()
                    print("socket recv : ", recvData)
                except:
                    recvData = ""
                    logger.info(f'[Check] Recv Data is Different')
                # print(recvData)

                if recvData == "TOPCAPCOMP":
                    print("[INFO] TOPCAPCOMP 수신 : ", time.time() - STH.startTime)
                    logger.info(f'[INFO] TOPCAPCOMP 수신')
                    # while True:
                    #     if CTH.CapCompMain == True:
                    #         break
                    #     time.sleep(0.1)
                    #     print("메인 상부 촬영 대기중")

                    # self.sendAllClient('SIDEINSSTART')

                if "COMP" in recvData:
                    # DataIndex = int(recvData[4:5])
                    print(f"Notice : [Recive Data {recvData}]")
                    logger.info(f'[INFO] Model Loading Complete by Client - {recvData}')
                    main_frame.update_signal(f'Loading Complete by Client - {recvData}')
                    DTT.AllCompRecv += 1

                # GPU 확인 텍스트 추가 2022-12-09 OH
                if "GPUNG" in recvData: 
                    camidx = recvData[6:7]
                    if camidx == '1' :
                        main_frame.main_canvas.itemconfig(main_frame.gputext1, state = 'normal')
                    elif camidx == '2':
                        main_frame.main_canvas.itemconfig(main_frame.gputext2, state = 'normal')
                    elif camidx == '3':
                        main_frame.main_canvas.itemconfig(main_frame.gputext3, state = 'normal')
                    elif camidx == '4':
                        main_frame.main_canvas.itemconfig(main_frame.gputext4, state = 'normal')
                    elif camidx == '5':
                        main_frame.main_canvas.itemconfig(main_frame.gputext5, state = 'normal')
                    elif camidx == '6':
                        main_frame.main_canvas.itemconfig(main_frame.gputext6, state = 'normal')

                # GPU 확인 텍스트 추가 2022-12-09 OH
                if "GPUOK" in recvData: 
                    camidx = recvData[6:7]
                    if camidx == '1' :
                        main_frame.main_canvas.itemconfig(main_frame.gputext1, state = 'hidden')
                    elif camidx == '2':
                        main_frame.main_canvas.itemconfig(main_frame.gputext2, state = 'hidden')
                    elif camidx == '3':
                        main_frame.main_canvas.itemconfig(main_frame.gputext3, state = 'hidden')
                    elif camidx == '4':
                        main_frame.main_canvas.itemconfig(main_frame.gputext4, state = 'hidden')
                    elif camidx == '5':
                        main_frame.main_canvas.itemconfig(main_frame.gputext5, state = 'hidden')
                    elif camidx == '6':
                        main_frame.main_canvas.itemconfig(main_frame.gputext6, state = 'hidden')
                        
                if "RESULT" in recvData:
                    DataIndex = int(recvData[6:7])
                    print(DataIndex)
                    if STH.ManualMode == True:
                        print(f"Notice : [Receive Data RESULT{DataIndex}]-menual")
                        result_value_data = self.recvall(client_socket, 100)
                        result_value = result_value_data.decode().strip()
                        print(result_value)
                        print(DTT.CompResultList)
                        DTT.CompResultList[DataIndex] = result_value
                        DTT.AllDataRecv += 1

                        # length_Blur = self.recvall(client_socket, 20)
                        # binData_Blur = self.recvall(client_socket, int(length_Blur))
                        # decBlur = pickle.loads(binData_Blur)
                        # DTT.CompBlurDataList[DataIndex] = decBlur.copy()

                        length = self.recvall(client_socket, 20)
                        binData = self.recvall(client_socket, int(length))
                        # data = np.frombuffer(stringData, dtype='uint8')
                        # decimg=cv2.imdecode(data,1)
                        decimg = pickle.loads(binData)
                        # type(decimg)
                        DTT.CompImageList[DataIndex] = decimg.copy()
                        # (W, H) = DTT.CompImageList[DataIndex][0].shape[:2]
                        # print(W,H)
                        print(f"RESULT{DataIndex} : {DTT.CompResultList}")
                        print(
                            f"[INFO]ResultImage Recv, Client{DataIndex} : ", time.time() - STH.startTime,
                        )
                        DTT.AllImageRecv += 1
                    else:
                        if (0 < DataIndex < 4) if CodeSetup == "CONE" else (0 < DataIndex < 3):
                            print(f"Notice : [Receive Data RESULT{DataIndex}]")
                            result_value_data = self.recvall(client_socket, 100)
                            result_value = result_value_data.decode().strip()
                            DTT.frontDataList[0][DataIndex - 1] = result_value
                            DTT.AllDataRecv += 1

                            # length_Blur = self.recvall(client_socket, 20)
                            # binData_Blur = self.recvall(client_socket, int(length_Blur))
                            # decBlur = pickle.loads(binData_Blur)
                            # DTT.frontBlurDataList[0][DataIndex - 1] = decBlur.copy()

                            length = self.recvall(client_socket, 20)
                            binData = self.recvall(client_socket, int(length))
                            # data = np.frombuffer(stringData, dtype='uint8')
                            # decimg=cv2.imdecode(data,1)
                            decimg = pickle.loads(binData)
                            DTT.frontImageList[0][DataIndex - 1] = decimg.copy()
                            print(f"RESULT{DataIndex} : {DTT.frontDataList}")
                            print(
                                f"[INFO]ResultImage Recv, Client{DataIndex} : ", time.time() - STH.startTime,
                            )
                            DTT.AllImageRecv += 1

                        elif (DataIndex >= 4 if CodeSetup == "CONE" else DataIndex >= 3) or DataIndex == 0:
                            print(f"Notice : [Receive Data RESULT{DataIndex}]")
                            result_value_data = self.recvall(client_socket, 100)
                            result_value = result_value_data.decode().strip()
                            DTT.CompResultList[DataIndex] = result_value
                            DTT.AllDataRecv += 1

                            # length_Blur = self.recvall(client_socket, 20)
                            # binData_Blur = self.recvall(client_socket, int(length_Blur))
                            # decBlur = pickle.loads(binData_Blur)
                            # DTT.CompBlurDataList[DataIndex] = decBlur.copy()

                            length = self.recvall(client_socket, 20)
                            binData = self.recvall(client_socket, int(length))
                            # data = np.frombuffer(stringData, dtype='uint8')
                            # decimg=cv2.imdecode(data,1)
                            decimg = pickle.loads(binData)
                            # type(decimg)
                            DTT.CompImageList[DataIndex] = decimg.copy()
                            # (W, H) = DTT.CompImageList[DataIndex][0].shape[:2]
                            # print(W,H)
                            print(f"RESULT{DataIndex} : {DTT.CompResultList}")
                            print(
                                f"[INFO]ResultImage Recv, Client{DataIndex} : ", time.time() - STH.startTime,
                            )
                            DTT.AllImageRecv += 1
                    logger.info(f"Notice - RESULT {DataIndex} Recv / DataRecvCount : {DTT.AllDataRecv} / ImageRecvCount : {DTT.AllImageRecv}")
                    print(DTT.AllDataRecv, DTT.AllImageRecv)

                if "PICKLE" in recvData:
                    try:
                        DataIndex = recvData[6:7]
                        updateX = DTT.PickleDict[recvData][0]
                        updateY = DTT.PickleDict[recvData][1]
                        strListData = self.recvall(client_socket, 500)
                        strListData = strListData.decode("utf-8")
                        DTT.pickleListData[int(DataIndex)] = eval(strListData)
                        print(DTT.pickleListData[int(DataIndex)])
                        main_frame.makeLabelWindow(updateX, updateY, DTT.pickleListData[int(DataIndex)], (int(DataIndex)))
                    except Exception as ex:
                        logger.info(f"[NOTICE] Pickle Data Recv Error : {traceback.format_exc()}")
                        print(traceback.format_exc())

                if "DATAREQUEST" in recvData:
                    DataIndex = int(recvData[11:12])
                    sendData = str(DTT.pickleListData[DataIndex])
                    client_socket.send(sendData.ljust(500).encode())

                if recvData == "CLIENT0":
                    # self.cilentIP[0] = str(addr[0])
                    my_client = "0"
                    main_frame.main_canvas.itemconfig(main_frame.connectionCheck[0], state="normal")
                    ngcage_frame.ngcage_canvas.itemconfig(ngcage_frame.connectionCheck[0], state="normal")
                    logger.info(f'[Notice] {recvData} Connect Check')
                elif recvData == "CLIENT1":
                    self.cilentIP[0] = str(addr[0])
                    my_client = "1"
                    main_frame.main_canvas.itemconfig(main_frame.connectionCheck[1], state="normal")
                    ngcage_frame.ngcage_canvas.itemconfig(ngcage_frame.connectionCheck[1], state="normal")
                    logger.info(f'[Notice] {recvData} Connect Check')
                elif recvData == "CLIENT2":
                    self.cilentIP[1] = str(addr[0])
                    my_client = "2"
                    main_frame.main_canvas.itemconfig(main_frame.connectionCheck[2], state="normal")
                    ngcage_frame.ngcage_canvas.itemconfig(ngcage_frame.connectionCheck[2], state="normal")
                    logger.info(f'[Notice] {recvData} Connect Check')
                elif recvData == "CLIENT3":
                    self.cilentIP[2] = str(addr[0])
                    my_client = "3"
                    main_frame.main_canvas.itemconfig(main_frame.connectionCheck[3], state="normal")
                    ngcage_frame.ngcage_canvas.itemconfig(ngcage_frame.connectionCheck[3], state="normal")
                    logger.info(f'[Notice] {recvData} Connect Check')
                elif recvData == "CLIENT4":
                    self.cilentIP[3] = str(addr[0])
                    my_client = "4"
                    main_frame.main_canvas.itemconfig(main_frame.connectionCheck[4], state="normal")
                    ngcage_frame.ngcage_canvas.itemconfig(ngcage_frame.connectionCheck[4], state="normal")
                    logger.info(f'[Notice] {recvData} Connect Check')
                elif recvData == "CLIENT5":
                    self.cilentIP[4] = str(addr[0])
                    my_client = "5"
                    main_frame.main_canvas.itemconfig(main_frame.connectionCheck[5], state="normal")
                    ngcage_frame.ngcage_canvas.itemconfig(ngcage_frame.connectionCheck[5], state="normal")
                    logger.info(f'[Notice] {recvData} Connect Check')

                data = ""

                if 'NGFRAME' in recvData:
                    ng_client_num = recvData[:3]
                    ng_part_num = recvData[10]
                    ngname = recvData[11:]

                    length = self.recvall(client_socket, 20)
                    binData = self.recvall(client_socket, int(length))
                    decimg = pickle.loads(binData)
                    
                    NF.save_new_image_and_remove_oldest('NG', decimg, ng_client_num, ng_part_num, ngname)

                if 'OKFRAME' in recvData:
                    ok_client_num = recvData[:3]
                    ok_part_num = recvData[10]
                    okname = recvData[11:]

                    length = self.recvall(client_socket, 20)
                    binData = self.recvall(client_socket, int(length))
                    decimg = pickle.loads(binData)
                    
                    NF.save_new_image_and_remove_oldest('OK', decimg, ok_client_num, ok_part_num, okname)

            except:
                print("Disconnected by " + addr[0], ":", addr[1], "(error)", traceback.format_exc())
                logger.info(f'[NOTICE] Disconnected by  + {addr[0]} : {addr[1]} \n {traceback.format_exc()}')
                break

        logger.info(f"[NOTICE] Disconnected by  + {addr[0]} : {addr[1]} \n {traceback.format_exc()}")
        if my_client == "0":
            main_frame.main_canvas.itemconfig(main_frame.connectionCheck[0], state="hidden")
            ngcage_frame.ngcage_canvas.itemconfig(ngcage_frame.connectionCheck[0], state="hidden")
        elif my_client == "1":
            main_frame.main_canvas.itemconfig(main_frame.connectionCheck[1], state="hidden")
            ngcage_frame.ngcage_canvas.itemconfig(ngcage_frame.connectionCheck[1], state="hidden")
        elif my_client == "2":
            main_frame.main_canvas.itemconfig(main_frame.connectionCheck[2], state="hidden")
            ngcage_frame.ngcage_canvas.itemconfig(ngcage_frame.connectionCheck[2], state="hidden")
        elif my_client == "3":
            main_frame.main_canvas.itemconfig(main_frame.connectionCheck[3], state="hidden")
            ngcage_frame.ngcage_canvas.itemconfig(ngcage_frame.connectionCheck[3], state="hidden")
        elif my_client == "4":
            main_frame.main_canvas.itemconfig(main_frame.connectionCheck[4], state="hidden")
            ngcage_frame.ngcage_canvas.itemconfig(ngcage_frame.connectionCheck[4], state="hidden")
        elif my_client == "5":
            main_frame.main_canvas.itemconfig(main_frame.connectionCheck[5], state="hidden")
            ngcage_frame.ngcage_canvas.itemconfig(ngcage_frame.connectionCheck[5], state="hidden")

        self.allClient.remove(client_socket)
        self.allClientAddr.remove(addr[0])
        self.connect_count -= 1
        if self.connect_count < 0:
            self.connect_count = 0

        if self.connect_count != ConnectClientCheck:
            self.connect_check = False

        client_socket.close()
        print("Client Closed")
        logger.info(f"Warning : Client Thread Closed - {my_client}")
        main_frame.update_signal(f"Client{my_client} Disconnect")

    def sendAllClient(self, msg, length=20):
        if isinstance(msg, str):
            for client in self.allClient:
                # print('send ip : ',client, 'msg : ',msg.ljust(length))
                client.send(msg.ljust(length).encode())
        else:
            for client in self.allClient:
                client.send(msg)

    def run(self):
        time.sleep(5)
        while True:
            # 서버소켓은 accept() 함수에서 기다리다가 클라이언트 소켓으로부터 신호가 오면 메시지와 주소를 받는다.
            client_socket, addr = self.server_socket.accept()
            # time.sleep(1)
            addrData = addr[0]
            self.allClient += [client_socket]
            self.allClientAddr += [addrData]
            # 각각의 클라이언트 소켓들을 새로운 쓰레드로 할당한다.
            start_new_thread(self.threaded, (client_socket, addr))
            self.connect_count += 1
            print("Notice : [Now Connecting Member = {}]".format(self.connect_count))
            if self.connect_count >= ConnectClientCheck:
                self.connect_check = True
                main_frame.update_signal("Client All Connect")
            # self.recvAllClient()
            time.sleep(0.01)


class DataTrimThread(threading.Thread):
    def __init__(self):
        threading.Thread.__init__(self)
        self.inspectionSession = False
        self.CompImageList = [[], [], [], [], [], []]
        self.CompResultList = ["10000000000000"] * (6 if CodeSetup == "CONE" else 5)
        # self.CompTypeList = [None] * 6
        self.AllDataRecv = 0
        self.AllImageRecv = 0
        self.AllCompRecv = 0
        self.imgSaveCount = 0
        self.modelLoadingCheck = False
        self.OKdata = "00000000000000"
        self.NoneData = "10000000000000"
        self.NgCage1Setup = [1,2,3,4,5]

        #카메라 블러처리 리미트 셋업
        if CodeSetup == 'CONE':
            self.CameraBlurLimit = {
                'CLIENT0' : [1,2],
                'CLIENT1' : [1,2],
                'CLIENT2' : [1,2],
                'CLIENT3' : [1,2],
                'CLIENT4' : [1,2],
                'CLIENT5' : [1,2],
            }
        else:
            self.CameraBlurLimit = {
                'CLIENT0' : [300,400],
                'CLIENT1' : [300,400],
                'CLIENT2' : [50,100],
                'CLIENT3' : [300,400],
                'CLIENT4' : [300,400],
            }
        #카메라 블러처리 결과데이터
        self.CompBlurDataList = [[],[],[],[],[],[]] if CodeSetup == "CONE" else [[], [], [], [], []]
        self.frontBlurDataList = [[[],[],[]], [[],[],[]], [[],[],[]], [[],[],[]]]

        self.forceClientNgData = {
            'CLIENT0' : '00000000100000',
            'CLIENT1' : '00000001000000',
            'CLIENT2' : '00000000010000' if CodeSetup == 'CUP' else '00000000001000',
            'CLIENT3' : '00000000000001' if CodeSetup == 'CUP' else '00000000010000',
            'CLIENT4' : '00000000000010' if CodeSetup == 'CUP' else '00000000000100',
            'CLIENT5' : '10000000000000' if CodeSetup == 'CUP' else '00000000000100'
        }

        self.frontDataList = [
            [self.NoneData, self.NoneData, self.NoneData],
            [self.NoneData, self.NoneData, self.NoneData],
            [self.NoneData, self.NoneData, self.NoneData],
            [self.NoneData, self.NoneData, self.NoneData],
        ]
        self.frontImageList = [[[], [], []], [[], [], []], [[], [], []], [[], [], []]]
        # self.frontTypeList = [[None, None, None], [None, None, None], [None, None, None]]

        if CodeSetup == "CONE":
            self.PickleDict = {
                "PICKLE0": [36, 548],
                "PICKLE1": [36, 131],
                "PICKLE2": [484, 131],
                "PICKLE3": [932, 131],
                "PICKLE4": [484, 548],
                "PICKLE5": [932, 548],
            }
        else:
            self.PickleDict = {
                "PICKLE0": [36, 548],
                "PICKLE1": [36, 131],
                "PICKLE2": [484, 131],
                "PICKLE3": [484, 548],
                "PICKLE4": [932, 548],
            }
        # self.badTypeList = ['TYPE00'] * 20
        self.nowbadTypeIndex1 = 10
        self.nowbadTypeIndex2 = 10
        self.inputngCage = 0

        self.SendingBadType = {
            0: 0,
            1: 1,
            2: 2,
            3: 3,
            4: 4,
            5: 5,
            6: 7,
            7: 8,
            8: 9,
            9: 10,
            10: 11,
            11: 12,
            12: 14,
            13: 16,
            14: 18,
            15: 19,
            16: 20,
            17: 23,
            18: 26,
            19: 29,
            20: 33,
            21: 39,
            22: 42,
        }
        self.badType = {
            0: "데이터 부족", 
            1: "S/F 누락", 
            2: "롤러 누락", 
            3: "각인 누락", 
            4: "식별각인 누락",
            5: '케이지 깨짐',
            6: "이종 혼입",
            7: "소단면 불량", 
            8: "대단면 불량", 
            9: "소단 내경 불량", 
            10: "소단 외경 불량", 
            11: "대단 외경 불량", 
            12: "대단 내경 불량", 
            13: "외경 불량",
        }
        self.pickleListData = [None] * (6 if CodeSetup == "CONE" else 5)
        self.DbupdatePath = [None] * (6 if CodeSetup == "CONE" else 5)

        self.NgContinuityCheck = [[] for _ in range(20)]
        self.NgContinuityCount = 0

        if CodeSetup == 'CONE':
            self.UpdateCameraName = {0: "114 CAMERA", 1: "111 CAMERA", 2: "112 CAMERA", 3: "113 CAMERA", 4: "115 CAMERA", 5: "116 CAMERA"}
        else:
            self.UpdateCameraName = {0: "104 CAMERA", 1: "101 CAMERA", 2: "102 CAMERA", 3: "103 CAMERA", 4: "105 CAMERA"}
        
        self.ProductData = []
        self.product_infoJson_Load()

    def product_infoJson_Load(self):
        if os.path.isfile('CheckValue/product_info.json'):
            with open('CheckValue/product_info.json', 'r') as read_file:
                ReadJsonData = json.load(read_file)
            self.ProductData = ReadJsonData[LINE][CodeSetup][1]
            self.common_model_name = ReadJsonData[LINE][CodeSetup][0]
            # print(self.ProductData)

    def run(self):
        sendData1 = "NO"
        sendData2 = "NO"
        NoneDataImg_ori = cv2.imread(f"bg/{BgSetup}/NoneData.png")
        NoneDataImg = []
        NoneDataImg.append(NoneDataImg_ori)
        BypassDataImg_ori = cv2.imread(f"bg/{BgSetup}/bypassImage.png")
        BypassDataImg = []
        BypassDataImg.append(BypassDataImg_ori)
        NoneDataBlur = ['CLIENT', 0]
        while True:
            try:
                time.sleep(0.1)
                if self.AllCompRecv >= ConnectClientCheck:
                    self.AllCompRecv = 0
                    # STH.sendMsg('COMP', 'COMP')
                    # COMP Signal Send
                    main_frame.update_signal("Model Loading Comp")
                    print("Model Loading Comp")
                    main_frame.main_canvas.itemconfig(main_frame.modelloadingView, state="hidden")
                    main_frame.loadingProcessTrigger = False
                    # STH.sendPlcSocketData(f'R')
                    # STH.SendDataTrim("W", "D3003", "1")
                    STH.ModbusSignalSend(Mode = "READY", Value = 1)
                    STH.ModbusSignalSend(Mode = "BUSY", Value = 0)
                    management_frame.management_State_SendAll()
                    self.modelLoadingCheck = True

                if self.AllDataRecv >= ConnectClientCheck or (STH.DataTrimCheck == True and (time.time() - STH.startTime) > 10):
                    if (STH.DataTrimCheck == True and (time.time() - STH.startTime) > 10):
                        logger.info(f'[CHECK] Time Over Datatrim Session - DataTrim')
                        main_frame.update_signal(f'[CHECK] Time Over Datatrim Session_D')
                    STH.DataTrimCheck = None
                    # if STH.trigerCheck == False:
                    #     main_frame.update_signal('RESULT Signal Not Research / auto datatrim start')
                    #     logger.info('[Notice] 검사 결과 강제 요청')
                    # # STH.ModbusSignalCheck[4] = True
                    # # STH.Modbusclient.write_coils(address = 31088, count = 2, values = [True,False])
                    # print("결과신호 강제 요청")
                    # # TSD.sendAllClient('RESULTREQUEST')
                    # print("[INFO] RESULTREQUEST 송신 : ", time.time() - STH.startTime)

                    # main_frame.update_signal('Light OFF')
                    # time.sleep(0.5)
                    # checkSum = LST.check_sum(LST.lightValue2, LST.lightValue2, LST.lightValue2, LST.lightValue2, LST.lightValue2, LST.lightValue2, LST.lightValue2, LST.lightValue2)
                    # values = bytearray([58, 58, 0, LST.lightValue2, LST.lightValue2, LST.lightValue2, LST.lightValue2, LST.lightValue2, LST.lightValue2, LST.lightValue2, LST.lightValue2, checkSum, 238, 238])
                    # LST.sendMsg(values)

                    self.AllDataRecv = 0
                    if STH.ManualMode == True:
                        print("메뉴얼모드 진행중")
                        print("결과 데이터 수신 완료")
                        print("결과 데이터 front > Comp 업데이트")

                        # for i in range(5 if CodeSetup == 'CUP' else 6):
                        #     try:
                        #         if self.CompBlurDataList[i][1] > self.CameraBlurLimit[self.CompBlurDataList[i][0]][1]:
                        #             print(f'{self.CompBlurDataList[i][0]} Blur Data Session OK - {self.CompBlurDataList[i][1]}')
                        #             logger.info(f'{self.CompBlurDataList[i][0]} Blur Data Session OK - {self.CompBlurDataList[i][1]}')
                        #             main_frame.main_canvas.itemconfig(main_frame.CameraBlur[i], state = 'hidden')
                        #             pass
                        #         elif self.CameraBlurLimit[self.CompBlurDataList[i][0]][0] < self.CompBlurDataList[i][1] <= self.CameraBlurLimit[self.CompBlurDataList[i][0]][1]:
                        #             print(f'{self.CompBlurDataList[i][0]} Blur Data Session Check - {self.CompBlurDataList[i][1]}')
                        #             logger.info(f'{self.CompBlurDataList[i][0]} Blur Data Session Check - {self.CompBlurDataList[i][1]}')
                        #             main_frame.main_canvas.itemconfig(main_frame.CameraBlur[i], state = 'normal')
                        #             pass
                        #         elif self.CameraBlurLimit[self.CompBlurDataList[i][0]][0] >= self.CompBlurDataList[i][1]:
                        #             print(f'{self.CompBlurDataList[i][0]} Blur Data Session Trouble - {self.CompBlurDataList[i][1]}')
                        #             logger.info(f'{self.CompBlurDataList[i][0]} Blur Data Session Trouble - {self.CompBlurDataList[i][1]}')
                        #             main_frame.main_canvas.itemconfig(main_frame.CameraBlur[i], state = 'normal')
                        #             pass
                        #         else:
                        #             print(f'{self.CompBlurDataList[i][0]} Blur Data Session Index Error - {self.CompBlurDataList[i][1]}')
                        #             logger.info(f'{self.CompBlurDataList[i][0]} Blur Data Session Index Error - {self.CompBlurDataList[i][1]}')
                        #             main_frame.main_canvas.itemconfig(main_frame.CameraBlur[i], state = 'hidden')
                        #             self.CompResultList[i] = self.forceClientNgData[self.CompBlurDataList[i][0]]
                        #     except:
                        #         print(f'CLIENT{i} Blur Data Session Error')
                        #         logger.info(f'CLIENT{i} Blur Data Session Error')

                        # print(self.CompResultList)
                        
                        if CodeSetup == "CONE":
                            if self.CompResultList[1] == self.OKdata and self.CompResultList[2] == self.OKdata and self.CompResultList[3] == self.OKdata:
                                sendData1 = "OK"
                            elif self.CompResultList[1] == self.NoneData or self.CompResultList[2] == self.NoneData or self.CompResultList[3] == self.NoneData:
                                sendData1 = "NO"
                            else:
                                omissionCheck = False
                                for i in range(len(DTT.NgCage1Setup)):
                                    if self.CompResultList[1][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client1\n\n\n')
                                        omissionCheck = True
                                    if self.CompResultList[2][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client2\n\n\n')
                                        omissionCheck = True
                                    if self.CompResultList[3][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client3\n\n\n')
                                        omissionCheck = True

                                
                                if omissionCheck == False:
                                    sendData1 = "NG"
                                else:
                                    sendData1 = "NG_Omisson"

                            if self.CompResultList[0] == self.OKdata and self.CompResultList[4] == self.OKdata and self.CompResultList[5] == self.OKdata:
                                sendData2 = "OK"
                            elif self.CompResultList[0] == self.NoneData or self.CompResultList[4] == self.NoneData or self.CompResultList[5] == self.NoneData:
                                sendData2 = "NO"
                            else:
                                omissionCheck = False
                                for i in range(len(DTT.NgCage1Setup)):
                                    if self.CompResultList[0][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client0\n\n\n')
                                        omissionCheck = True
                                    if self.CompResultList[4][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client4\n\n\n')
                                        omissionCheck = True
                                    if self.CompResultList[5][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client5\n\n\n')
                                        omissionCheck = True

                                
                                if omissionCheck == False:
                                    sendData2 = "NG"
                                else:
                                    sendData2 = "NG_Omisson"
                        else:
                            if self.CompResultList[1] == self.OKdata and self.CompResultList[2] == self.OKdata:
                                sendData1 = "OK"
                            elif self.CompResultList[1] == self.NoneData or self.CompResultList[2] == self.NoneData:
                                sendData1 = "NO"
                            else:
                                omissionCheck = False
                                for i in range(len(DTT.NgCage1Setup)):
                                    if self.CompResultList[1][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client1\n\n\n')
                                        omissionCheck = True
                                    if self.CompResultList[2][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client2\n\n\n')
                                        omissionCheck = True

                                
                                if omissionCheck == False:
                                    sendData1 = "NG"
                                else:
                                    sendData1 = "NG_Omisson"
                            if self.CompResultList[0] == self.OKdata and self.CompResultList[3] == self.OKdata and self.CompResultList[4] == self.OKdata:
                                sendData2 = "OK"
                            elif self.CompResultList[0] == self.NoneData or self.CompResultList[3] == self.NoneData or self.CompResultList[4] == self.NoneData:
                                sendData2 = "NO"
                            else:
                                omissionCheck = False
                                for i in range(len(DTT.NgCage1Setup)):
                                    if self.CompResultList[0][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client0\n\n\n')
                                        omissionCheck = True
                                    if self.CompResultList[3][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client3\n\n\n')
                                        omissionCheck = True
                                    if self.CompResultList[4][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client4\n\n\n')
                                        omissionCheck = True
                                
                                if omissionCheck == False:
                                    sendData2 = "NG"
                                else:
                                    sendData2 = "NG_Omisson"

                        if STH.StartMode[0] == "1":
                            print("Station 3 - Product On")
                        else:
                            print("Station 3 - Product Off")
                            sendData1 = "NO"

                        if STH.StartMode[1] == "1":
                            print("Station 6 - Product On")
                        else:
                            print("Station 6 - Product Off")
                            sendData2 = "NO"

                        print(self.CompResultList)
                        print("sendData1 : ", sendData1, "sendData2 : ", sendData2)

                        if main_frame.bypass_mode == True:
                            if sendData1 != "NO":
                                sendData1 = "OK"
                            if sendData2 != "NO":
                                sendData2 = "OK"

                        if main_frame.test_mode == False:
                            if sendData1 != "NO":
                                sendData1 = "NG_Omisson"
                            if sendData2 != "NO":
                                sendData2 = "NG_Omisson"
                        if main_frame.test_mode == True:
                            if sendData1 != "NO":
                                sendData1 = "NG"
                            if sendData2 != "NO":
                                sendData2 = "NG"
                            
                        if sendData1 == "OK":
                            # STH.SendDataTrim("W", "D3006", "3")
                            STH.ModbusSignalSend(Mode = "RESULT1", Value = 3)
                            logger.info("[Notice] Result Send - Station3 > OK")
                            print("[Notice] Result Send - Station3 > OK")
                        elif sendData1 == "NG":
                            # STH.SendDataTrim("W", "D3006", "2")  # ORIGIN
                            STH.ModbusSignalSend(Mode = "RESULT1", Value = 2)
                            logger.info("[Notice] Result Send - Station3 > NG")
                            print("[Notice] Result Send - Station3 > NG")
                        elif sendData1 == 'NG_Omisson':
                            # STH.SendDataTrim("W", "D3006", "1")  # ORIGIN
                            STH.ModbusSignalSend(Mode = "RESULT1", Value = 1)
                            logger.info('[Notice] Result Send - Station3 > Omission NG')
                            print('[Notice] Result Send - Station3 > Omission NG')
                        else:
                            # STH.SendDataTrim("W", "D3006", "4")
                            STH.ModbusSignalSend(Mode = "RESULT1", Value = 4)
                            logger.info("[Notice] Result Send - Station3 > None")

                        # time.sleep(0.5)

                        if sendData2 == "OK":
                            # STH.SendDataTrim("W", "D3007", "3")
                            STH.ModbusSignalSend(Mode = "RESULT2", Value = 3)
                            logger.info("[Notice] Result Send - Station6 > OK")
                            print("[Notice] Result Send - Station6 > OK")
                        elif sendData2 == "NG":
                            # STH.SendDataTrim('W','D3007','2') #ORIGIN
                            STH.ModbusSignalSend(Mode = "RESULT2", Value = 2)
                            logger.info("[Notice] Result Send - Station6 > OK")
                            print("[Notice] Result Send - Station6 > OK")
                        elif sendData2 == 'NG_Omisson':
                            # STH.SendDataTrim("W", "D3007", "1")  # ORIGIN
                            STH.ModbusSignalSend(Mode = "RESULT2", Value = 1)
                            logger.info('[Notice] Result Send - Station6 > Omission NG')
                            print('[Notice] Result Send - Station6 > Omission NG')
                        else:
                            # STH.SendDataTrim("W", "D3007", "4")
                            STH.ModbusSignalSend(Mode = "RESULT2", Value = 4)
                            logger.info("[Notice] Result Send - Station6 > None")
                            print("[Notice] Result Send - Station6 > None")
                        STH.ModbusSignalSend(Mode = "RESULTSESSION", Value = 2)
                        main_frame.update_signal(f"RESULT send {sendData1}{sendData2}")
                    else:
                        # print("[INFO] Result 결과 송신 데이터 종합 시작 : ", time.time() - STH.startTime)

                        print("결과 데이터 수신 완료")
                        print("결과 데이터 front > Comp 업데이트")
                        for i in range(len(management_frame.ManagementBtnIndex[3])):
                            if management_frame.BtnState[3][i][0] == True:
                                if CodeSetup == 'CONE':
                                    if i == 0:
                                        #101 or 111 bypass
                                        DTT.frontDataList[0][0] = self.OKdata
                                        # DTT.frontBlurDataList[0][0] = ['CLIENT1', 100000]
                                        DTT.frontImageList[0][0] = BypassDataImg.copy()
                                        pass
                                    elif i == 1:
                                        #102 or 112 bypass
                                        DTT.frontDataList[0][1] = self.OKdata
                                        # DTT.frontBlurDataList[0][1] = ['CLIENT2', 100000]
                                        DTT.frontImageList[0][1] = BypassDataImg.copy()
                                        pass
                                    elif i == 2:
                                        #103 or 113 
                                        DTT.frontDataList[0][2] = self.OKdata
                                        # DTT.frontBlurDataList[0][2] = ['CLIENT3', 100000]
                                        DTT.frontImageList[0][2] = BypassDataImg.copy()
                                        pass
                                    elif i == 3:
                                        #104 or 114 bypass
                                        self.CompResultList[0] = self.OKdata 
                                        # self.CompBlurDataList[0] = ['CLIENT0', 100000]
                                        self.CompImageList[0] = BypassDataImg.copy()
                                        pass
                                    elif i == 4:
                                        #105 or 115 bypass
                                        self.CompResultList[4] = self.OKdata 
                                        # self.CompBlurDataList[4] = ['CLIENT0', 100000]
                                        self.CompImageList[4] = BypassDataImg.copy()
                                        pass
                                    elif i == 5:
                                        #106 or 115 bypass
                                        self.CompResultList[5] = self.OKdata 
                                        # self.CompBlurDataList[5] = ['CLIENT0', 100000]
                                        self.CompImageList[5] = BypassDataImg.copy()
                                        pass
                                else:
                                    if i == 0:
                                        #101 or 111 bypass
                                        DTT.frontDataList[0][0] = self.OKdata
                                        DTT.frontBlurDataList[0][0] = ['CLIENT1', 100000]
                                        DTT.frontImageList[0][0] = BypassDataImg.copy()
                                        print('CLIENT1 bypass data update')
                                        pass
                                    elif i == 1:
                                        #102 or 112 bypass
                                        DTT.frontDataList[0][1] = self.OKdata
                                        DTT.frontBlurDataList[0][1] = ['CLIENT2', 100000]
                                        DTT.frontImageList[0][1] = BypassDataImg.copy()
                                        print('CLIENT2 bypass data update')
                                        pass
                                    elif i == 2:
                                        #103 or 113 bypass
                                        self.CompResultList[3] = self.OKdata
                                        # self.CompBlurDataList[3] = ['CLIENT3', 100000]
                                        self.CompImageList[3] = BypassDataImg.copy()
                                        print('CLIENT3 bypass data update')
                                        pass
                                    elif i == 3:
                                        #104 or 114 bypass
                                        self.CompResultList[0] = self.OKdata 
                                        # self.CompBlurDataList[0] = ['CLIENT0', 100000]
                                        self.CompImageList[0] = BypassDataImg.copy()
                                        print('CLIENT0 bypass data update')
                                        pass
                                    elif i == 4:
                                        #105 or 115 bypass
                                        self.CompResultList[4] = self.OKdata
                                        # self.CompBlurDataList[4] = ['CLIENT4', 100000]
                                        self.CompImageList[4] = BypassDataImg.copy()
                                        print('CLIENT4 bypass data update')
                                        pass
                                pass
                        for i in range(3 if CodeSetup == "CONE" else 2):
                            try:
                                self.CompResultList[i + 1] = self.frontDataList[3][i]
                                # self.CompBlurDataList[i + 1] = self.frontBlurDataList[3][i]
                            except Exception as ex:
                                self.CompResultList[i + 1] = self.NoneData
                                # self.CompBlurDataList[i + 1] = NoneDataBlur
                                logger.info(f"Warning : 해당 문제로 인하여 에러가 발생 - {ex} \n {traceback.format_exc()}")
                                print(ex)
                        # for i in range(5 if CodeSetup == 'CUP' else 6):
                        #     try:
                        #         if self.CompBlurDataList[i][1] > self.CameraBlurLimit[self.CompBlurDataList[i][0]][1]:
                        #             print(f'{self.CompBlurDataList[i][0]} Blur Data Session OK - {self.CompBlurDataList[i][1]}')
                        #             logger.info(f'{self.CompBlurDataList[i][0]} Blur Data Session OK - {self.CompBlurDataList[i][1]}')
                        #             main_frame.main_canvas.itemconfig(main_frame.CameraBlur[i], state = 'hidden')
                        #             pass
                        #         elif self.CameraBlurLimit[self.CompBlurDataList[i][0]][0] < self.CompBlurDataList[i][1] <= self.CameraBlurLimit[self.CompBlurDataList[i][0]][1]:
                        #             print(f'{self.CompBlurDataList[i][0]} Blur Data Session Check - {self.CompBlurDataList[i][1]}')
                        #             logger.info(f'{self.CompBlurDataList[i][0]} Blur Data Session Check - {self.CompBlurDataList[i][1]}')
                        #             main_frame.main_canvas.itemconfig(main_frame.CameraBlur[i], state = 'normal')
                        #             pass
                        #         elif self.CameraBlurLimit[self.CompBlurDataList[i][0]][0] >= self.CompBlurDataList[i][1]:
                        #             print(f'{self.CompBlurDataList[i][0]} Blur Data Session Trouble - {self.CompBlurDataList[i][1]}')
                        #             logger.info(f'{self.CompBlurDataList[i][0]} Blur Data Session Trouble - {self.CompBlurDataList[i][1]}')
                        #             main_frame.main_canvas.itemconfig(main_frame.CameraBlur[i], state = 'normal')
                        #             pass
                        #         else:
                        #             print(f'{self.CompBlurDataList[i][0]} Blur Data Session Index Error - {self.CompBlurDataList[i][1]}')
                        #             logger.info(f'{self.CompBlurDataList[i][0]} Blur Data Session Index Error - {self.CompBlurDataList[i][1]}')
                        #             main_frame.main_canvas.itemconfig(main_frame.CameraBlur[i], state = 'hidden')
                        #             self.CompResultList[i] = self.forceClientNgData[self.CompBlurDataList[i][0]]
                        #     except:
                        #         print(f'CLIENT{i} Blur Data Session Error')
                        #         logger.info(f'CLIENT{i} Blur Data Session Error')

                        print("데이터 화면 업데이트")
                        if CodeSetup == "CONE":
                            if self.frontDataList[0][0] == self.OKdata and self.frontDataList[0][1] == self.OKdata and self.frontDataList[0][2] == self.OKdata:
                                sendData1 = "OK"
                            elif self.frontDataList[0][0] == self.NoneData or self.frontDataList[0][1] == self.NoneData or self.frontDataList[0][2] == self.NoneData:
                                sendData1 = "NO"
                            else:
                                omissionCheck = False
                                for i in range(len(DTT.NgCage1Setup)):
                                    if self.frontDataList[0][0][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client1\n\n\n')
                                        omissionCheck = True
                                    if self.frontDataList[0][1][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client2\n\n\n')
                                        omissionCheck = True
                                    if self.frontDataList[0][2][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client3\n\n\n')
                                        omissionCheck = True

                                
                                if omissionCheck == False:
                                    sendData1 = "NG"
                                else:
                                    sendData1 = "NG_Omisson"
                            if self.CompResultList[0] == self.OKdata and self.CompResultList[4] == self.OKdata and self.CompResultList[5] == self.OKdata:
                                sendData2 = "OK"
                            elif self.CompResultList[0] == self.NoneData or self.CompResultList[4] == self.NoneData or self.CompResultList[5] == self.NoneData:
                                sendData2 = "NO"
                            else:
                                omissionCheck = False
                                for i in range(len(DTT.NgCage1Setup)):
                                    if self.CompResultList[0][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client0\n\n\n')
                                        omissionCheck = True
                                    if self.CompResultList[4][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client4\n\n\n')
                                        omissionCheck = True
                                    if self.CompResultList[5][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client5\n\n\n')
                                        omissionCheck = True

                                if omissionCheck == False:
                                    sendData2 = "NG"
                                else:
                                    sendData2 = "NG_Omisson"
                        else:
                            if self.frontDataList[0][0] == self.OKdata and self.frontDataList[0][1] == self.OKdata:
                                sendData1 = "OK"
                            elif self.frontDataList[0][0] == self.NoneData or self.frontDataList[0][1] == self.NoneData:
                                sendData1 = "NO"
                            else:
                                omissionCheck = False
                                for i in range(len(DTT.NgCage1Setup)):
                                    if self.frontDataList[0][0][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client1\n\n\n')
                                        omissionCheck = True
                                    if self.frontDataList[0][1][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client2\n\n\n')
                                        omissionCheck = True

                                
                                if omissionCheck == False:
                                    sendData1 = "NG"
                                else:
                                    sendData1 = "NG_Omisson"
                            if self.CompResultList[0] == self.OKdata and self.CompResultList[3] == self.OKdata and self.CompResultList[4] == self.OKdata:
                                sendData2 = "OK"
                            elif self.CompResultList[0] == self.NoneData or self.CompResultList[3] == self.NoneData or self.CompResultList[4] == self.NoneData:
                                sendData2 = "NO"
                            else:
                                omissionCheck = False
                                for i in range(len(DTT.NgCage1Setup)):
                                    if self.CompResultList[0][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client0\n\n\n')
                                        omissionCheck = True
                                    if self.CompResultList[3][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client3\n\n\n')
                                        omissionCheck = True
                                    if self.CompResultList[4][i+1] == '1':
                                        print('\n\n\n[Notice] Omission Data Check - Client4\n\n\n')
                                        omissionCheck = True

                                
                                if omissionCheck == False:
                                    sendData2 = "NG"
                                else:
                                    sendData2 = "NG_Omisson"

                        if STH.StartMode[0] == "1":
                            print("Station 3 - Product On")
                        else:
                            print("Station 3 - Product Off")
                            sendData1 = "NO"

                        if STH.StartMode[1] == "1":
                            print("Station 6 - Product On")
                        else:
                            print("Station 6 - Product Off")
                            sendData2 = "NO"

                        # print(self.CompResultList)
                        print("sendData1 : ", sendData1, "sendData2 : ", sendData2)
                        # print("[INFO] 스타트 모드에 따른 sendData 초기화 : ", time.time() - STH.startTime)

                        if main_frame.bypass_mode == True:
                            if sendData1 != "NO":
                                sendData1 = "OK"
                            if sendData2 != "NO":
                                sendData2 = "OK"

                        if main_frame.test_mode == False:
                            if sendData1 != "NO":
                                sendData1 = "NG_Omisson"
                            if sendData2 != "NO":
                                sendData2 = "NG_Omisson"
                        if main_frame.test_mode == True:
                            if sendData1 != "NO":
                                sendData1 = "NG"
                            if sendData2 != "NO":
                                sendData2 = "NG"

                        if sendData1 == "OK":
                            # STH.SendDataTrim("W", "D3006", "3")
                            STH.ModbusSignalSend(Mode = "RESULT1", Value = 3)
                            logger.info("[Notice] Result Send - Station3 > OK")
                            print("[Notice] Result Send - Station3 > OK")
                        elif sendData1 == "NG":
                            # STH.SendDataTrim("W", "D3006", "2")  # ORIGIN
                            STH.ModbusSignalSend(Mode = "RESULT1", Value = 2)
                            logger.info(f"[Notice] Result Send - Station3 > NG")
                            print(f"[Notice] Result Send - Station3 > NG")
                        elif sendData1 == "NG_Omisson":
                            # STH.SendDataTrim("W", "D3006", "1")  # ORIGIN
                            STH.ModbusSignalSend(Mode = "RESULT1", Value = 1)
                            logger.info(f"[Notice] Result Send - Station3 > NG_Omisson")
                            print(f"[Notice] Result Send - Station3 > NG_Omisson")
                        else:
                            # STH.SendDataTrim("W", "D3006", "4")
                            STH.ModbusSignalSend(Mode = "RESULT1", Value = 4)
                            logger.info("[Notice] Result Send - Station3 > None")
                            print("[Notice] Result Send - Station3 > None")

                        # time.sleep(0.5)

                        if sendData2 == "OK":
                            # STH.SendDataTrim("W", "D3007", "3")
                            STH.ModbusSignalSend(Mode = "RESULT2", Value = 3)
                            logger.info("[Notice] Result Send - Station6 > OK")
                            print("[Notice] Result Send - Station6 > OK")
                        elif sendData2 == "NG":
                            # STH.SendDataTrim("W", "D3007", "2")  # ORIGIN
                            STH.ModbusSignalSend(Mode = "RESULT2", Value = 2)
                            logger.info(f"[Notice] Result Send - Station6 > NG")
                            print(f"[Notice] Result Send - Station6 > NG")
                        elif sendData2 == "NG_Omisson":
                            # STH.SendDataTrim("W", "D3007", "1")  # ORIGIN
                            STH.ModbusSignalSend(Mode = "RESULT2", Value = 1)
                            logger.info(f"[Notice] Result Send - Station6 > NG_Omisson")
                            print(f"[Notice] Result Send - Station6 > NG_Omisson")
                        else:
                            # STH.SendDataTrim("W", "D3007", "4")
                            STH.ModbusSignalSend(Mode = "RESULT2", Value = 4)
                            logger.info("[Notice] Result Send - Station6 > None")
                            print("[Notice] Result Send - Station6 > None")
                        main_frame.update_signal(f"RESULT send {sendData1}{sendData2}")
                        logger.info(f"RESULT send {sendData1}{sendData2}")
                        STH.ModbusSignalSend(Mode = "RESULTSESSION", Value = 2)

                        ########################################################################################### 1211 추가 (화면에 보여질 불량 갯수체크 데이터 종합)
                        for i in range(len(self.CompResultList)):
                            try:
                                for x in range(len(DTT.badType.keys())):
                                    try:
                                        if self.CompResultList[i][x] == "1":
                                            main_frame.mainbadCheckList[i][x] += 1
                                    except:
                                        # print('except break')
                                        break
                            except:
                                # print('Data Check Error', traceback.format_exc())
                                pass
                        print(f'[Check] mainbadCheckList - {main_frame.mainbadCheckList}')

                if self.AllImageRecv >= ConnectClientCheck or (STH.DataTrimCheck == None and (time.time() - STH.startTime) > 10):
                    if (STH.DataTrimCheck == None and (time.time() - STH.startTime) > 10):
                        logger.info(f'[CHECK] Time Over Datatrim Session - ImageTrim')
                        main_frame.update_signal(f'[CHECK] Time Over Datatrim Session_I')
                    STH.DataTrimCheck = False
                    ngCheckList = []
                    if STH.ManualMode == True:
                        print("menual Mode Image Data Check Start")
                        print("[INFO] Result 이미지 송신 데이터 종합 시작 : ", time.time() - STH.startTime)
                        self.AllImageRecv = 0

                        print("이미지 데이터 수신 완료")

                        self.DbupdatePath = [None] * (6 if CodeSetup == "CONE" else 5)
                        for i in range(6 if CodeSetup == "CONE" else 5):
                            UpdateCheck = False
                            for x in range(len(self.CompImageList[i])):
                                self.imgSaveCount += 1
                                year_s = datetime.datetime.today().year
                                month_s = datetime.datetime.today().month
                                day_s = datetime.datetime.today().day
                                path = f"ResultData/{str(year_s)}/{str(month_s)}/{str(day_s)}/{i}"
                                inputNum = "%05d" % self.imgSaveCount

                                try:
                                    self.CompImageList[i][x] = cv2.resize(self.CompImageList[i][x], dsize=(391, 290))
                                except:
                                    self.CompImageList[i][x] = NoneDataImg_ori.copy()
                                    print(traceback.format_exc())

                                # 수동모드 사진 저장 안함
                                # MQ.put((path, 'SaveImg_{}.jpg'.format(inputNum), self.CompImageList[i][x]))

                                if x == 0:
                                    main_frame.show_img(self.CompImageList[i][x], i)
                                    UpdateCheck = True
                                else:
                                    print(traceback.format_exc())
                            if UpdateCheck == False:
                                main_frame.show_img(NoneDataImg_ori, i)
                            else:
                                pass

                        for i in range(6 if CodeSetup == "CONE" else 5):
                            if self.CompResultList[i] == self.OKdata:
                                main_frame.main_canvas.itemconfig(main_frame.main_ok[i], state="normal")
                                main_frame.main_canvas.itemconfig(main_frame.main_ng[i], state="hidden")
                            elif self.CompResultList[i] == self.NoneData or self.CompResultList[i] == None:
                                main_frame.main_canvas.itemconfig(main_frame.main_ok[i], state="hidden")
                                main_frame.main_canvas.itemconfig(main_frame.main_ng[i], state="hidden")
                            elif "1" in self.CompResultList[i]:
                                main_frame.main_canvas.itemconfig(main_frame.main_ok[i], state="hidden")
                                main_frame.main_canvas.itemconfig(main_frame.main_ng[i], state="normal")

                        print("[INFO] 수동모드 데이터베이스 저장 안함 : ", time.time() - STH.startTime)
                    else:
                        print("[INFO] Result 이미지 송신 데이터 종합 시작 : ", time.time() - STH.startTime)
                        STH.ngContinuousCount += 1
                        self.AllImageRecv = 0
                        print("이미지 데이터 front > Comp 업데이트")
                        for i in range(3 if CodeSetup == "CONE" else 2):
                            try:
                                if self.frontImageList[3][i]:
                                    self.CompImageList[i + 1] = self.frontImageList[3][i].copy()
                                else:
                                    self.CompImageList[i + 1] = NoneDataImg.copy()
                            except Exception as ex:
                                logger.info(f"Warning : 해당 문제로 인하여 에러가 발생 - {ex} \n {traceback.format_exc()}")
                                self.CompImageList[i + 1] = NoneDataImg.copy()

                        self.DbupdatePath = [None] * (6 if CodeSetup == "CONE" else 5)
                        for i in range(6 if CodeSetup == "CONE" else 5):
                            UpdateCheck = False
                            for x in range(len(self.CompImageList[i])):
                                self.imgSaveCount += 1
                                year_s = datetime.datetime.today().year
                                month_s = datetime.datetime.today().month
                                day_s = datetime.datetime.today().day
                                path = f"ResultData/{str(year_s)}/{str(month_s)}/{str(day_s)}/{i}"
                                inputNum = datetime.datetime.now().strftime("%Y_%m_%d_%H_%M_%S")

                                try:
                                    self.CompImageList[i][x] = cv2.resize(self.CompImageList[i][x], dsize=(391, 290))
                                except:
                                    self.CompImageList[i].append(cv2.imread(f"bg/{BgSetup}/NoneData.png"))
                                if CodeSetup == 'CUP':
                                    # if "1" in self.CompResultList[0] or "1" in self.CompResultList[1] or "1" in self.CompResultList[2] or "1" in self.CompResultList[3] or "1" in self.CompResultList[4]:
                                    MQ.put((path, "SaveImg_{}.jpg".format(inputNum), self.CompImageList[i][x],))
                                else:
                                    # if "1" in self.CompResultList[0] or "1" in self.CompResultList[1] or "1" in self.CompResultList[2] or "1" in self.CompResultList[3] or "1" in self.CompResultList[4] or "1" in self.CompResultList[5]:
                                    MQ.put((path, "SaveImg_{}.jpg".format(inputNum), self.CompImageList[i][x],))

                                if x == 0:
                                    self.DbupdatePath[i] = path + f"/SaveImg_{inputNum}.jpg"
                                    main_frame.show_img(self.CompImageList[i][x], i)
                                    UpdateCheck = True
                                else:
                                    self.DbupdatePath[i] = self.DbupdatePath[i] + "||" + path + f"/SaveImg_{inputNum}.jpg"
                            if UpdateCheck == False:
                                main_frame.show_img(NoneDataImg_ori, i)
                            else:
                                pass
                                # print(self.DbupdatePath)

                        if CodeSetup == "CONE":
                            if self.CompResultList[0] == self.NoneData or self.CompResultList[1] == self.NoneData or self.CompResultList[2] == self.NoneData or self.CompResultList[3] == self.NoneData or self.CompResultList[4] == self.NoneData or self.CompResultList[0] == None or self.CompResultList[1] == None or self.CompResultList[2] == None or self.CompResultList[3] == None or self.CompResultList[4] == None or self.CompResultList[5] == self.NoneData or self.CompResultList[5] == None:
                                logger.info(f"[Notice] : None Data Check {self.CompResultList}")
                                pass
                            elif self.CompResultList[0] == self.OKdata and self.CompResultList[1] == self.OKdata and self.CompResultList[2] == self.OKdata and self.CompResultList[3] == self.OKdata and self.CompResultList[4] == self.OKdata and self.CompResultList[5] == self.OKdata:
                                main_frame.ok_count += 1
                                main_frame.total_count += 1
                                main_frame.main_canvas.itemconfig(main_frame.okText, text=main_frame.ok_count)
                                main_frame.main_canvas.itemconfig(main_frame.totalText, text=main_frame.total_count)
                                logger.info(f"[Notice] : Ok Data Check {self.CompResultList}")
                            elif "1" in self.CompResultList[0] or "1" in self.CompResultList[1] or "1" in self.CompResultList[2] or "1" in self.CompResultList[3] or "1" in self.CompResultList[4] or "1" in self.CompResultList[5]:
                                # check
                                main_frame.ng_count += 1
                                main_frame.total_count += 1
                                main_frame.main_canvas.itemconfig(main_frame.ngText, text=main_frame.ng_count)
                                main_frame.main_canvas.itemconfig(main_frame.totalText, text=main_frame.total_count)

                                logger.info(f"[Notice] : Ng Data Check {self.CompResultList}")

                                if "1" in self.CompResultList[0]:
                                    STH.ngContinuousList[0] += 1
                                if "1" in self.CompResultList[1]:
                                    STH.ngContinuousList[1] += 1
                                if "1" in self.CompResultList[2]:
                                    STH.ngContinuousList[2] += 1
                                if "1" in self.CompResultList[3]:
                                    STH.ngContinuousList[3] += 1
                                if "1" in self.CompResultList[4]:
                                    STH.ngContinuousList[4] += 1
                                if "1" in self.CompResultList[5]:
                                    STH.ngContinuousList[5] += 1

                        else:
                            if self.CompResultList[0] == self.NoneData or self.CompResultList[1] == self.NoneData or self.CompResultList[2] == self.NoneData or self.CompResultList[3] == self.NoneData or self.CompResultList[4] == self.NoneData or self.CompResultList[0] == None or self.CompResultList[1] == None or self.CompResultList[2] == None or self.CompResultList[3] == None or self.CompResultList[4] == None:  # or self.CompResultList[5] == self.NoneData  or self.CompResultList[5] == None
                                logger.info(f"[Notice] : None Data Check {self.CompResultList}")
                                pass
                            elif self.CompResultList[0] == self.OKdata and self.CompResultList[1] == self.OKdata and self.CompResultList[2] == self.OKdata and self.CompResultList[3] == self.OKdata and self.CompResultList[4] == self.OKdata:  # and self.CompResultList[5] == self.OKdata:
                                main_frame.ok_count += 1
                                main_frame.total_count += 1
                                main_frame.main_canvas.itemconfig(main_frame.okText, text=main_frame.ok_count)
                                main_frame.main_canvas.itemconfig(main_frame.totalText, text=main_frame.total_count)
                                logger.info(f"[Notice] : Ok Data Check {self.CompResultList}")
                            elif "1" in self.CompResultList[0] or "1" in self.CompResultList[1] or "1" in self.CompResultList[2] or "1" in self.CompResultList[3] or "1" in self.CompResultList[4]:  # or '1' in self.CompResultList[5]:
                                # check
                                main_frame.ng_count += 1
                                main_frame.total_count += 1
                                main_frame.main_canvas.itemconfig(main_frame.ngText, text=main_frame.ng_count)
                                main_frame.main_canvas.itemconfig(main_frame.totalText, text=main_frame.total_count)

                                logger.info(f"[Notice] : Ng Data Check {self.CompResultList}")

                                if "1" in self.CompResultList[0]:
                                    STH.ngContinuousList[0] += 1
                                if "1" in self.CompResultList[1]:
                                    STH.ngContinuousList[1] += 1
                                if "1" in self.CompResultList[2]:
                                    STH.ngContinuousList[2] += 1
                                if "1" in self.CompResultList[3]:
                                    STH.ngContinuousList[3] += 1
                                if "1" in self.CompResultList[4]:
                                    STH.ngContinuousList[4] += 1

                        # OH 85% 디스크 차면 팝업창 # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
                        drive = 'c:'
                        c = dict([x for x in zip(['total','used','free'], shutil.disk_usage(drive))])
                        result = c['used']/c['total']*100

                        if result > 85 :
                            print("[INFO] 디스크 85% 용량 초과")
                            logger.info(f"[INFO] disk error")
                            main_frame.main_canvas.itemconfig(main_frame.show_pop, image = main_frame.disk_error_pop, state="normal")
                        else : 
                            main_frame.main_canvas.itemconfig(main_frame.show_pop, image = '', state="hidden")
                        # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 

                        for i in range(6 if CodeSetup == "CONE" else 5):
                            if self.CompResultList[i] == self.OKdata:
                                main_frame.main_canvas.itemconfig(main_frame.main_ok[i], state="normal")
                                main_frame.main_canvas.itemconfig(main_frame.main_ng[i], state="hidden")
                            elif self.CompResultList[i] == self.NoneData or self.CompResultList[i] == None:
                                main_frame.main_canvas.itemconfig(main_frame.main_ok[i], state="hidden")
                                main_frame.main_canvas.itemconfig(main_frame.main_ng[i], state="hidden")
                            elif "1" in self.CompResultList[i]:
                                main_frame.main_canvas.itemconfig(main_frame.main_ok[i], state="hidden")
                                main_frame.main_canvas.itemconfig(main_frame.main_ng[i], state="normal")
                        # (H, W) = self.CompImageList[0][0].shape[:2]
                        # print(W, H)

                        print("[INFO] 데이터 종합 저장 및 갯수 카운트 : ", time.time() - STH.startTime)

                        if (self.CompResultList[0] == self.OKdata and self.CompResultList[1] == self.OKdata and self.CompResultList[2] == self.OKdata and self.CompResultList[3] == self.OKdata and self.CompResultList[4] == self.OKdata and self.CompResultList[5] == self.OKdata) if CodeSetup == "CONE" else (self.CompResultList[0] == self.OKdata and self.CompResultList[1] == self.OKdata and self.CompResultList[2] == self.OKdata and self.CompResultList[3] == self.OKdata and self.CompResultList[4] == self.OKdata):
                            bearingartDB.dbCount_Process(STH.nowModel, 'OK')
                            print(self.CompResultList)
                        else:
                            print('CompTypeList ori: ', self.CompResultList)
                            ngCheckList = []
                            updateMsg = ""
                            checkCount = 0
                            for i in range(6 if CodeSetup == "CONE" else 5):
                                for x in range(0, len(DTT.badType.keys())):
                                    if x in ngCheckList:
                                        print("데이터 이미 업데이트 완료")
                                    else:
                                        try:
                                            if self.CompResultList[i][x] == "1":
                                                ngCheckList.append(x)
                                                print("ng데이터 추가완료")
                                        except:
                                            pass
                            ngCheckList.sort()
                            for i in range(len(DTT.badType.keys())):
                                if i in ngCheckList:
                                    print("ngList : ", ngCheckList)
                                    print("setting : ", DTT.NgCage1Setup)
                                    if updateMsg == "":
                                        updateMsg = updateMsg + self.badType[i]
                                    else:
                                        checkCount += 1
                                    if i in DTT.NgCage1Setup:
                                        print("NG케이지 1에 업데이트해야함", i)

                            if checkCount == 0:
                                ngcageText = updateMsg
                            else:
                                ngcageText = updateMsg + f" 외 {checkCount}건"
                            print("최종 메시지 : ", ngcageText)

                            if "데이터 부족" in ngcageText:
                                # bearingartDB.dbCount_Process(STH.nowModel, 'NG')      #데이터 없음시에도 db 카운터 늘리려면 주석 풀기 / 아니면 주석상태로 방치
                                pass
                            else:
                                for i in range(len(DTT.badType.keys())):
                                    if i in ngCheckList:
                                        # if DTT.ngCageSetting[i] == True:
                                        if i in DTT.NgCage1Setup:
                                            self.inputngCage = 1
                                            print("ngCage1 update check")
                                            break
                                        else:
                                            self.inputngCage = 2
                                            print("ngCage2 update check")
                                    else:
                                        self.inputngCage = 2
                                        print("ngCage2 update check")

                                # print(self.CompResultList)
                                # print(ngCheckList)
                                if self.inputngCage == 1:
                                    bearingartDB.dbCount_Process(STH.nowModel, 'CRI')
                                    print("ngcage1 update")

                                    if self.nowbadTypeIndex1 == 0:
                                        for i in reversed(range(9)):
                                            ngcage_frame.ngCageUpdate1[i + 1] = ngcage_frame.ngCageUpdate1[i].copy()
                                            main_frame.bad_text1_inputData[i + 1] = main_frame.bad_text1_inputData[i]
                                            main_frame.main_canvas.itemconfig(
                                                main_frame.bad_text1[i + 1], text=main_frame.bad_text1_inputData[i + 1],
                                            )
                                            if i == 0:
                                                ngcage_frame.ngCageUpdate1[0] = [
                                                    False,
                                                    [None, None, None, None, None],
                                                    [None, None, None, None, None],
                                                    [None],
                                                ]
                                                main_frame.bad_text1_inputData[0] = None

                                    self.nowbadTypeIndex1 -= 1
                                    if self.nowbadTypeIndex1 == -1:
                                        self.nowbadTypeIndex1 = 0

                                    main_frame.main_canvas.itemconfig(
                                        main_frame.bad_image1[self.nowbadTypeIndex1], image=main_frame.main_frame_bad,
                                    )
                                    main_frame.bad_text1_inputData[self.nowbadTypeIndex1] = ngcageText
                                    main_frame.main_canvas.itemconfig(
                                        main_frame.bad_text1[self.nowbadTypeIndex1], text=main_frame.bad_text1_inputData[self.nowbadTypeIndex1],
                                    )

                                    ngcage_frame.ngCageUpdate1[self.nowbadTypeIndex1][0] = True
                                    ngcage_frame.ngCageUpdate1[self.nowbadTypeIndex1][1] = copy.deepcopy(self.CompResultList)
                                    ngcage_frame.ngCageUpdate1[self.nowbadTypeIndex1][2] = copy.deepcopy(self.CompImageList)
                                    ngcage_frame.ngCageUpdate1[self.nowbadTypeIndex1][3] = copy.deepcopy(ngCheckList)

                                elif self.inputngCage == 2:
                                    print("ngcage2 update")
                                    bearingartDB.dbCount_Process(STH.nowModel, 'NG')

                                    if self.nowbadTypeIndex2 == 0:
                                        for i in reversed(range(9)):
                                            ngcage_frame.ngCageUpdate2[i + 1] = ngcage_frame.ngCageUpdate2[i].copy()
                                            main_frame.bad_text2_inputData[i + 1] = main_frame.bad_text2_inputData[i]
                                            main_frame.main_canvas.itemconfig(
                                                main_frame.bad_text2[i + 1], text=main_frame.bad_text2_inputData[i + 1],
                                            )
                                            if i == 0:
                                                ngcage_frame.ngCageUpdate2[0] = [
                                                    False,
                                                    [None, None, None, None, None],
                                                    [None, None, None, None, None],
                                                    [None],
                                                ]
                                                main_frame.bad_text2_inputData[0] = None

                                    self.nowbadTypeIndex2 -= 1
                                    if self.nowbadTypeIndex2 == -1:
                                        self.nowbadTypeIndex2 = 0

                                    main_frame.main_canvas.itemconfig(
                                        main_frame.bad_image2[self.nowbadTypeIndex2], image=main_frame.main_frame_bad,
                                    )
                                    main_frame.bad_text2_inputData[self.nowbadTypeIndex2] = ngcageText
                                    main_frame.main_canvas.itemconfig(
                                        main_frame.bad_text2[self.nowbadTypeIndex2], text=main_frame.bad_text2_inputData[self.nowbadTypeIndex2],
                                    )

                                    ngcage_frame.ngCageUpdate2[self.nowbadTypeIndex2][0] = True
                                    ngcage_frame.ngCageUpdate2[self.nowbadTypeIndex2][1] = copy.deepcopy(self.CompResultList)
                                    ngcage_frame.ngCageUpdate2[self.nowbadTypeIndex2][2] = copy.deepcopy(self.CompImageList)
                                    ngcage_frame.ngCageUpdate2[self.nowbadTypeIndex2][3] = copy.deepcopy(ngCheckList)

                                    

                        print("[INFO] NG cage update : ", time.time() - STH.startTime)

                        # DATABASE UPDATE
                        if CodeSetup == "CONE":
                            # if "1" in self.CompResultList[0] or "1" in self.CompResultList[1] or "1" in self.CompResultList[2] or "1" in self.CompResultList[3] or "1" in self.CompResultList[4] or "1" in self.CompResultList[5]:
                            inputDBData = {
                                ("ModelName", f"'{STH.nowModel}'"),
                                ("Result1", f"'{self.CompResultList[0]}'"),
                                ("Result2", f"'{self.CompResultList[1]}'"),
                                ("Result3", f"'{self.CompResultList[2]}'"),
                                ("Result4", f"'{self.CompResultList[3]}'"),
                                ("Result5", f"'{self.CompResultList[4]}'"),
                                ("Result6", f"'{self.CompResultList[5]}'"),
                                ("ImagePath1", f"'{self.DbupdatePath[0]}'"),
                                ("ImagePath2", f"'{self.DbupdatePath[1]}'"),
                                ("ImagePath3", f"'{self.DbupdatePath[2]}'"),
                                ("ImagePath4", f"'{self.DbupdatePath[3]}'"),
                                ("ImagePath5", f"'{self.DbupdatePath[4]}'"),
                                ("ImagePath6", f"'{self.DbupdatePath[5]}'"),
                                ("InputTime", "now()"),
                            }

                            bearingartDB.writeSql(inputDBData)
                        else:
                            # if "1" in self.CompResultList[0] or "1" in self.CompResultList[1] or "1" in self.CompResultList[2] or "1" in self.CompResultList[3] or "1" in self.CompResultList[4]:
                            inputDBData = {
                                ("ModelName", f"'{STH.nowModel}'"),
                                ("Result1", f"'{self.CompResultList[0]}'"),
                                ("Result2", f"'{self.CompResultList[1]}'"),
                                ("Result3", f"'{self.CompResultList[2]}'"),
                                ("Result4", f"'{self.CompResultList[3]}'"),
                                ("Result5", f"'{self.CompResultList[4]}'"),
                                ("ImagePath1", f"'{self.DbupdatePath[0]}'"),
                                ("ImagePath2", f"'{self.DbupdatePath[1]}'"),
                                ("ImagePath3", f"'{self.DbupdatePath[2]}'"),
                                ("ImagePath4", f"'{self.DbupdatePath[3]}'"),
                                ("ImagePath5", f"'{self.DbupdatePath[4]}'"),
                                ("InputTime", "now()"),
                            }

                            bearingartDB.writeSql(inputDBData)

                        # 마무리 데이터 정리
                        print("데이터 순차적 정리(front data)")
                        # print(self.frontDataList)
                        for i in range(2, -1, -1):
                            try:
                                self.frontDataList[i + 1] = self.frontDataList[i].copy()
                                self.frontImageList[i + 1] = self.frontImageList[i].copy()
                                self.frontBlurDataList[i + 1] = self.frontBlurDataList[i].copy()
                                # self.frontTypeList[i+1] = self.frontTypeList[i].copy()
                            except:
                                print("error")
                        # print(self.frontDataList)
                        self.frontBlurDataList[0] = [[],[],[]]
                        print("[INFO] DB 저장 완료 : ", time.time() - STH.startTime)

                    # 불량 5개 카운트하여 신호 추가송신
                    try:
                        for i in range(5 if CodeSetup == 'CUP' else 6):
                            if main_frame.ngSettingList[i][2] == True:
                                if STH.ngContinuousCount >= int(main_frame.ngSettingList[i][0]):
                                    print(f"Ng Check Limit Over - {STH.ngContinuousCount}")
                                    SignalSendCheck = False
                                    for i in range(6 if CodeSetup == "CONE" else 5):
                                        if STH.ngContinuousList[i] >= int(main_frame.ngSettingList[i][1]):
                                            SignalSendCheck = True
                                            print(f"Ng Check List Over - Camera : {self.UpdateCameraName[i]}, Count : {STH.ngContinuousList[i]}, Limit : {main_frame.ngSettingList[i][1]}")
                                            logger.info(f"Ng Check List Over - Camera : {self.UpdateCameraName[i]}, Count : {STH.ngContinuousList[i]}, Limit : {main_frame.ngSettingList[i][1]}")
                                        else:
                                            print(f"Ng Check List Pass - Camera : {self.UpdateCameraName[i]}, Count : {STH.ngContinuousList[i]}, Limit : {main_frame.ngSettingList[i][1]}")
                                            logger.info(f"Ng Check List Pass - Camera : {self.UpdateCameraName[i]}, Count : {STH.ngContinuousList[i]}, Limit : {main_frame.ngSettingList[i][1]}")
                                            pass
                                    if SignalSendCheck == True:
                                        SignalSendCheck = False
                                        # STH.SendDataTrim("W", "D3016", "1")
                                        STH.ModbusSignalSend(Mode = "NGCONTINUITY", Value = 1)
                                    STH.ngContinuousCount = 0
                                    STH.ngContinuousList = [0] * (6 if CodeSetup == "CONE" else 5)
                                else:
                                    print(f"Ng Check Limit Count Lack")
                            else:
                                print("Ng Continuous Check Not Activate")
                                logger.info("Ng Continuous Check Not Activate")
                    except:
                        print(f"Ng Continuous Check Process Error : {traceback.format_exc()}")
                        logger.info(f"Ng Continuous Check Process Error : {traceback.format_exc()}")

                    if main_frame.bypass_mode == True:
                        pass
                    else:
                        if len(ngCheckList) < 1:
                            pass
                        else:
                            for x, value in enumerate(ngCheckList):
                                # STH.Modbusclient.write_register(address=41606+x, value=int(STH.ModbusCageSetting[self.SendingBadType[value]]))
                                pass
                    self.CompImageList = [[], [], [], [], [], []] if CodeSetup == "CONE" else [[], [], [], [], []]
                    # self.CompBlurDataList = [[],[],[],[],[],[]] if CodeSetup == "CONE" else [[], [], [], [], []]
                    self.CompResultList = ["10000000000000"] * (6 if CodeSetup == "CONE" else 5)
                    CountDataSave()

            except Exception as ex:
                logger.info(f"Warning : 데이터 후처리 스레드 에러 - {ex} - {traceback.format_exc()}")


class lightSignalTRAN(threading.Thread):
    def __init__(self):
        threading.Thread.__init__(self)
        self.lightConnectRetry()
        self.lightValue1 = [170, 170, 170, 255, 255, 255, 150, 150]
        self.lightValue2 = 0

    def lightConnectRetry(self):
        countingNum = 9

        for i in range(9):
            # countingEnd = 0
            try:
                countingNum = countingNum - 1
                osCommend = "sudo chmod a+rw /dev/ttyUSB" + str(countingNum)
                os.system(osCommend)
                self.SerialPort, self.SerialBaud = "/dev/ttyUSB" + str(countingNum), 19200
                self.ser = serial.Serial(self.SerialPort, self.SerialBaud, timeout=0)

            except Exception as ex:
                print(ex)
                print("[INFO] Serial Connecting Error - Number : ", self.SerialPort)

    def run(self):
        while True:
            self.read = self.ser.readline()  # .decode("ascii").strip()

            # print(self.read)

            time.sleep(0.08)

    def sendMsg(self, data):
        self.ser.write(data)

    def check_sum(self, data1, data2, data3, data4, data5, data6, data7, data8):
        result_data = data1 ^ data2 ^ data3 ^ data4 ^ data5 ^ data6 ^ data7 ^ data8
        return result_data


class serialTRAN(threading.Thread):
    def __init__(self):
        threading.Thread.__init__(self)
        ################# 시리얼 포트 변경 필요 #################
        # os.system('sudo chmod a+rw /dev/ttyUSB0')
        self.nowModel = ""
        self.MDdict = {
            1: "MODEL1",
            2: "MODEL2",
            3: "MODEL3",
            4: "MODEL4",  ########################################### need to change
            5: "MODEL5",
            6: "MODEL6",
            7: "MODEL7",
            8: "MODEL8",
            9: "MODEL9",
            10: "MODEL10",
            11: "MODEL11",
            12: "MODEL12",
            13: "MODEL13",
            14: "MODEL14",
            15: "MODEL15",
            15: "MODEL15",
            16: "MODEL16",
            17: "MODEL17",
            18: "MODEL18",
            19: "MODEL19",
            20: "MODEL20",
            21: "MODEL21",
            22: "MODEL22",
            23: "MODEL23",
            24: "MODEL24",
            25: "MODEL25",
            26: "MODEL26",
            27: "MODEL27",
            28: "MODEL28",
            29: "MODEL29",
            30: "MODEL30"
        }

        # self.connectRetry()
        self.CheckFirstTrigger = False
        self.ConnectModbus()
        
        self.startTime = time.time()
        self.requastTime = time.time()
        self.DataTrimCheck = False
        self.StartCheck = False
        self.ManualMode = True
        self.StartMode = "00"
        

        self.recvModbusChecker = [0,0,0,0,0,0,0,0,0,0]
        self.ModbusSessionChecker = [False,False,False,False,False,False,False,False,False,False]
        self.ModbusDefaultRecvAddress = 0
        self.DisposeMode = None
        self.ReworkMode = False
        self.EmergencyMode = False

        self.ngContinuousCount = 0
        self.ngContinuousList = [0, 0, 0, 0, 0, 0]

    def ConnectModbus(self):
        while True:
            try:
                try:
                    self.Modbusclient.close()
                except:
                    pass
                self.Modbusclient = ModbusTcpClient(host = ModebusHostIp, port = ModebusHostPort, timeout = 5)
                # self.Modbusclient.write_register(address=int(data1), value=int(data2), unit=0x01)
                # self.Modbusclient.write_registers(address=200, values=[1,0,0,0,0,0,0,0,0,0,0], unit=0x01)
                if self.CheckFirstTrigger == False:
                    self.ModbusSignalSend(Mode = "RESET")
                    self.CheckFirstTrigger = True
                self.ModbusSignalSend(Mode = "RECONNECT")
                self.ModbusSignalSend(Mode = "ALIVE", Value = 1)
                print(f'Modbus Signal Send > Mode - ALIVE, SendValue - 1')
                logger.info(f'Modbus Signal Send > Mode - ALIVE, SendValue - 1')
                return
            except:
                print(f'Modbus Connect Error - {traceback.format_exc()}')
                logger.info(f'Modbus Connect Error - {traceback.format_exc()}')
                time.sleep(1)

    def ModbusSignalSend(self, Mode, Value = 0):
        try:
            idx = int(Value)
            if Mode == "RESET":
                self.Modbusclient.write_registers(address=200, values=[0,0,0,0,0,0,0,0,0,0,0], unit=0x01)
            elif Mode == "REBOOT":
                self.Modbusclient.write_registers(address=200, values=[0,0,0,0,0,0,0,0,0,2,0], unit=0x01)
            elif Mode == "ALIVE":
                self.Modbusclient.write_register(address=200, value=idx, unit=0x01)
            elif Mode == "MODEL":
                self.Modbusclient.write_register(address=201, value=idx, unit=0x01)
            elif Mode == "LINEMODE":
                self.Modbusclient.write_register(address=202, value=idx, unit=0x01)
            elif Mode == "READY":
                self.Modbusclient.write_register(address=203, value=idx, unit=0x01)
            elif Mode == "BUSY":
                self.Modbusclient.write_register(address=204, value=idx, unit=0x01)
            elif Mode == "RESULTSESSION":
                self.Modbusclient.write_register(address=205, value=idx, unit=0x01)
            elif Mode == "RESULT1":
                self.Modbusclient.write_register(address=206, value=idx, unit=0x01)
            elif Mode == "RESULT2":
                self.Modbusclient.write_register(address=207, value=idx, unit=0x01)
            elif Mode == "NGCONTINUITY":
                self.Modbusclient.write_register(address=208, value=idx, unit=0x01)
            elif Mode == "STATE":
                self.Modbusclient.write_register(address=209, value=idx, unit=0x01)
            elif Mode == "PLCSETUP":
                self.Modbusclient.write_register(address=210, value=idx, unit=0x01)
            elif Mode == "RECONNECT":
                self.Modbusclient.write_registers(address=203, values=[1,0,0,0,0,0,0,0], unit=0x01)
            if Mode in ["RESET", "REBOOT"]:
                print(f'Modbus Signal Send > Mode - {Mode}')
                logger.info(f'Modbus Signal Send > Mode - {Mode}')
            else:
                if Mode == 'ALIVE':
                    pass
                else:
                    print(f'Modbus Signal Send > Mode - {Mode}, SendValue - {idx}')
                    logger.info(f'Modbus Signal Send > Mode - {Mode}, SendValue - {idx}')
        except:
            print(f'Modbus Send Error - {traceback.format_exc()}')
            logger.info(f'Modbus Send Error - {traceback.format_exc()}')
            time.sleep(1)

    def recvall(self, sock, count):
        buf = b""
        while count:
            newbuf = sock.recv(count)
            if not newbuf:
                return None
            buf += newbuf
            count -= len(newbuf)
        return buf

    def ReadDataCheck(self, Data):
        try:
            recvData = Data.split("/")

            checkDevice = recvData[1]
            checkValue = recvData[2]
            return checkDevice, checkValue
        except:
            print(traceback.format_exc())
            return None, None

    def SendDataTrim(self, Mode, Device, Data):
        SubData = f"{Mode}/{Device}/{Data}/"
        WriteData = SubData.ljust(20, "0")
        self.client_socket.send(WriteData.encode())
        print(f"send : [{WriteData}]")
        logger.info(f"[SIGNAL SEND] Mode : {Mode}, Device : {Device}, Data : {Data}")

    def retry_procedure(self):  # 프로그램 재실행 OH
        text = '[ ★ ] 프로그램 재시작'
        logger.info(text)
        print(text)
        time.sleep(1)
        python = sys.executable
        os.execl(python, python, * sys.argv)

    def run(self):
        while True:
            if TSD.connect_count == ConnectClientCheck:
                break
            time.sleep(1)
            print("Client All Connect Wait")
            main_frame.update_signal("Client All Connect Wait")
        self.ModbusSignalSend(Mode = "ALIVE", Value = 1)
        print(f'Modbus Signal Send > Mode - ALIVE, SendValue - 1')
        logger.info(f'Modbus Signal Send > Mode - ALIVE, SendValue - 1')
        cycleCheckCount = 0
        while True:
            try:
                startTime = time.time()
                time.sleep(0.1)
                result = self.Modbusclient.read_holding_registers(self.ModbusDefaultRecvAddress, count=10, unit=0x01)
                checkData = result.registers
                
                # print(checkData)

                for i in range(len(self.recvModbusChecker)):
                    if self.recvModbusChecker[i] != checkData[i]:
                        self.recvModbusChecker[i] = checkData[i]
                        self.ModbusSessionChecker[i] = True

                #PC Alive Cycle Upload
                cycleCheckCount += 1
                if cycleCheckCount >= 10:
                    cycleCheckCount = 0
                    self.ModbusSignalSend(Mode = "ALIVE", Value = 1)
                    # print(f'Modbus Signal Send > Mode - ALIVE, SendValue - 1')
                    # logger.info(f'Modbus Signal Send > Mode - ALIVE, SendValue - 1')

                #PLC Alive Signal Check
                if self.ModbusSessionChecker[0] == True:
                    if self.recvModbusChecker[0] == 1:
                        print(f'[Signal] PLC Alive On Check')
                        logger.info(f'[Signal] PLC Alive On Check')
                    elif self.recvModbusChecker[0] == 0:
                        print(f'[Signal] PLC Alive Off Check')
                        logger.info(f'[Signal] PLC Alive Off Check')
                    else:
                        print(f'[Signal] PLC Alive Signal Error - {self.ModbusDefaultRecvAddress}Address RecvData - {self.recvModbusChecker[0]}')
                        logger.info(f'[Signal] PLC Alive Off Check')
                    self.ModbusSessionChecker[0] = False

                #Model Signal Check
                if self.ModbusSessionChecker[1] == True:
                    try:
                        logger.info("[Signal] MODELLOADING Signal Recv")
                        
                        main_frame.main_canvas.itemconfig(main_frame.modelloadingView, state="normal")
                        main_frame.loadingProcessTrigger = None
                        threading.Thread(target=main_frame.loadingProcessGIF, daemon=True).start()
                        while True:
                            print(TSD.connect_count ,  ConnectClientCheck)
                            if TSD.connect_count == ConnectClientCheck:
                                break
                            time.sleep(1)
                            print("Client All Connect Wait")
                        indexData = int(self.recvModbusChecker[1])
                        self.nowModel = DTT.ProductData[indexData-1]

                        main_frame.update_signal(f"{self.MDdict[indexData]} recv")
                        TSD.sendAllClient(self.MDdict[indexData])
                        print(f"모델신호 {self.MDdict[indexData]} 송신")
                        main_frame.update_signal("Model Loading")
                        main_frame.main_canvas.itemconfig(main_frame.productName, text=DTT.ProductData[indexData-1])

                        #modbus Signal Send (Model Session)
                        self.ModbusSignalSend(Mode = "MODEL", Value = indexData)
                        self.ModbusSignalSend(Mode = "READY", Value = 0)
                        self.ModbusSignalSend(Mode = "BUSY", Value = 1)
                    except:
                        print(traceback.format_exc())
                        logger.info(f'Modbus Model Process Error - {traceback.format_exc()}')

                    self.ModbusSessionChecker[1] = False

                #Auto OR Manual Signal Check
                if self.ModbusSessionChecker[2] == True:
                    if self.recvModbusChecker[2] == 1:
                        print(f'[Signal] Modbus Auto Signal Recv')
                        logger.info(f'[Signal] Modbus Auto Signal Recv')
                        print("자동 모드 실행")
                        main_frame.update_signal("AUTO recv")
                        logger.info("[SIGNAL] AUTO Signal Recv")
                        self.ManualMode = False
                        # self.SendDataTrim("W", "D3002", f"{value}")
                        self.ModbusSignalSend(Mode = "LINEMODE", Value = int(self.recvModbusChecker[2]))
                    elif self.recvModbusChecker[2] == 2:
                        print(f'[Signal] Modbus Manual Signal Recv')
                        logger.info(f'[Signal] Modbus Manual Signal Recv')
                        main_frame.update_signal(f'[Signal] Modbus Manual Signal Recv')
                        STH.ngContinuousCount = 0
                        STH.ngContinuousList = [0] * (6 if CodeSetup == "CONE" else 5)
                        logger.info("[SIGNAL] NG Continuity Inspection Count Reset")
                        main_frame.update_signal("NG 연속검출 횟수 초기화")
                        self.ManualMode = True
                        self.ModbusSignalSend(Mode = "LINEMODE", Value = int(self.recvModbusChecker[2]))
                        self.ModbusSignalSend(Mode = "NGCONTINUITY", Value = 0)

                        # if DTT.modelLoadingCheck == True: # oh 수동 시 준비완료 제거
                        #     self.ModbusSignalSend(Mode = "READY", Value = 1)

                        # self.SendDataTrim("W", "D3002", f"{value}")
                        # self.SendDataTrim("W", "D3003", "1")
                        # self.SendDataTrim("W", "D3016", "0")
                    self.ModbusSessionChecker[2] = False

                #Start Signal Check
                if self.ModbusSessionChecker[3] == True:
                    if self.recvModbusChecker[3] == 1:
                        self.startTime = time.time()
                        self.StartCheck = True
                        self.DataTrimCheck = True
                        self.requastTime = time.time()
                        print(f'[Signal] Modbus Inspection Start Signal ON Recv')
                        logger.info(f'[Signal] Modbus Inspection Start Signal ON Recv')

                        logger.info("[SIGNAL] Start Signal Recv")
                        main_frame.update_signal("Start Recv")

                        # self.SendDataTrim("R", "D3109", "1")
                        # data = self.recvall(self.client_socket, 20).decode().strip()
                        # Pin1device, Pin1value = self.ReadDataCheck(data)

                        # self.SendDataTrim("R", "D3110", "1")
                        # data = self.recvall(self.client_socket, 20).decode().strip()
                        # Pin2device, Pin2value = self.ReadDataCheck(data)

                        pinData = self.Modbusclient.read_holding_registers(self.ModbusDefaultRecvAddress+4, count=2, unit=0x01)
                        pinCheckData = pinData.registers

                        self.StartMode = f"{pinCheckData[0]}{pinCheckData[1]}"
                        print(f"Start Mode - {self.StartMode}")
                        logger.info(f"[SIGNAL Check] StartMode - {self.StartMode}")

                        DTT.AllDataRecv = 0
                        DTT.AllImageRecv = 0

                        print("[INFO] START 수신 : ", time.time() - STH.startTime)

                        TSD.sendAllClient("START")

                        DTT.inspectionSession = True
                        print("[INFO] Start 시그널 처리 완료 : ", time.time() - STH.startTime)
                        # self.SendDataTrim("W", "D3003", "0")
                        self.ModbusSignalSend(Mode = "READY", Value = 0)
                        self.ModbusSignalSend(Mode = "BUSY", Value = 1)
                        self.ModbusSignalSend(Mode = "STATE", Value = 1)

                    elif self.recvModbusChecker[3] == 0:
                        print(f'[Signal] Modbus Inspection Start Signal OFF Recv')
                        logger.info(f'[Signal] Modbus Inspection Start Signal OFF Recv')

                    else:
                        print(f'[Signal] Modbus Inspection Start Signal Error')
                        logger.info(f'[Signal] Modbus Inspection Start Signal Error - {self.ModbusDefaultRecvAddress+3}Address RecvData - {self.recvModbusChecker[3]}')
                    self.ModbusSessionChecker[3] = False

                #InsPart1 Product Signal Check
                if self.ModbusSessionChecker[4] == True:
                    if self.recvModbusChecker[4] == 1:
                        print('[Signal] Modbus InspectionPart1 Product ON Signal Recv')
                        logger.info('[Signal] Modbus InspectionPart1 Product ON Signal Recv')
                    elif self.recvModbusChecker[4] == 0:
                        print('[Signal] Modbus InspectionPart1 Product OFF Signal Recv')
                        logger.info('[Signal] Modbus InspectionPart1 Product OFF Signal Recv')
                    else:
                        print('[Signal] Modbus InspectionPart1 Product Signal Error')
                        logger.info(f'[Signal] Modbus InspectionPart1 Product Signal Error - {self.ModbusDefaultRecvAddress+4}Address RecvData - {self.recvModbusChecker[4]}')
                    self.ModbusSessionChecker[4] = False

                #InsPart2 Product Signal Check
                if self.ModbusSessionChecker[5] == True:
                    if self.recvModbusChecker[5] == 1:
                        print('[Signal] Modbus InspectionPart2 Product ON Signal Recv')
                        logger.info('[Signal] Modbus InspectionPart2 Product ON Signal Recv')
                    elif self.recvModbusChecker[5] == 0:
                        print('[Signal] Modbus InspectionPart2 Product OFF Signal Recv')
                        logger.info('[Signal] Modbus InspectionPart2 Product OFF Signal Recv')
                    else:
                        print('[Signal] Modbus InspectionPart2 Product Signal Error')
                        logger.info(f'[Signal] Modbus InspectionPart2 Product Signal Error - {self.ModbusDefaultRecvAddress+5}Address RecvData - {self.recvModbusChecker[5]}')
                    self.ModbusSessionChecker[5] = False                    

                if (time.time()-self.requastTime) > 5 and self.StartCheck == True:
                    self.StartCheck = None
                    print('[Signal] Modbus ResultRequest Forced ON Signal recv')
                    logger.info('[Signal] Modbus ResultRequest Forced ON Signal recv')
                    main_frame.update_signal("Result Requast Forced recv")
                    TSD.sendAllClient("RESULTREQUEST")
                    self.ModbusSignalSend(Mode = "RESULTSESSION", Value = 1)
                    print("[INFO] RESULTREQUEST 송신 : ", time.time() - STH.startTime)

                #ResultRequest Signal Check
                if self.ModbusSessionChecker[6] == True:
                    if self.recvModbusChecker[6] == 1:
                        if self.StartCheck == True:
                            self.StartCheck = False
                            print('[Signal] Modbus ResultRequest ON Signal recv')
                            logger.info('[Signal] Modbus ResultRequest ON Signal recv')
                            main_frame.update_signal("Result Requast recv")
                            TSD.sendAllClient("RESULTREQUEST")
                            self.ModbusSignalSend(Mode = "RESULTSESSION", Value = 1)
                            print("[INFO] RESULTREQUEST 송신 : ", time.time() - STH.startTime)
                        else:
                            logger.info('[Notice] Forced Request Session - PASS')
                    elif self.recvModbusChecker[6] == 0:
                        print('[Signal] Modbus ResultRequest OFF Signal recv')
                        logger.info('[Signal] Modbus ResultRequest OFF Signal recv')
                    else:
                        print('[Signal] Modbus ResultRequest Signal Error')
                        logger.info(f'[Signal] Modbus ResultRequest Signal Error - {self.ModbusDefaultRecvAddress+6}Address RecvData - {self.recvModbusChecker[6]}')
                    self.ModbusSessionChecker[6] = False

                #Ins Result PLC Check Complete Signal Check
                if self.ModbusSessionChecker[7] == True:
                    if self.recvModbusChecker[7] == 1:
                        print('[Signal] Modbus PLC Complete ON Signal recv')
                        main_frame.update_signal('Modbus PLC Complete')
                        logger.info('[Signal] Modbus PLC Complete ON Signal recv')
                        self.ModbusSignalSend(Mode = "READY", Value = 1)
                        self.ModbusSignalSend(Mode = "BUSY", Value = 0)
                        self.ModbusSignalSend(Mode = "RESULTSESSION", Value = 0)
                        self.ModbusSignalSend(Mode = "RESULT1", Value = 0)
                        self.ModbusSignalSend(Mode = "RESULT2", Value = 0)
                        self.ModbusSignalSend(Mode = "STATE", Value = 0)
                        # self.SendDataTrim("W", "D3006", "0")
                        # self.SendDataTrim("W", "D3007", "0")
                        # self.SendDataTrim("W", "D3003", "1")
                    elif self.recvModbusChecker[7] == 0:
                        print('[Signal] Modbus PLC Complete OFF Signal recv')
                        logger.info('[Signal] Modbus PLC Complete OFF Signal recv')
                    else:
                        print('[Signal] Modbus PLC Complete Signal Error')
                        logger.info(f'[Signal] Modbus PLC Complete Signal Error - {self.ModbusDefaultRecvAddress+7}Address RecvData - {self.recvModbusChecker[7]}')
                    self.ModbusSessionChecker[7] = False

                #Vision Modify Signal Check
                if self.ModbusSessionChecker[8] == True:
                    if self.recvModbusChecker[8] == 0:
                        print('[Signal] Modbus Vision Modify Signal OFF Check')
                        logger.info('[Signal] Modbus Vision Modify Signal OFF Check')
                        if self.DisposeMode == True:
                            self.DisposeMode = False
                            logger.info("[SIGNAL] Dispose Session Complete - Process Reset")
                            main_frame.update_signal('Dispose Session Complete')
                            DTT.frontDataList = [
                                [DTT.NoneData, DTT.NoneData, DTT.NoneData],
                                [DTT.NoneData, DTT.NoneData, DTT.NoneData],
                                [DTT.NoneData, DTT.NoneData, DTT.NoneData],
                                [DTT.NoneData, DTT.NoneData, DTT.NoneData],
                            ]
                            DTT.frontImageList = [[[], [], []], [[], [], []], [[], [], []], [[], [], []]]
                            DTT.frontBlurDataList = [[[],[],[]], [[],[],[]], [[],[],[]], [[],[],[]]]
                            DTT.CompImageList = [[], [], [], [], [], []] if CodeSetup == "CONE" else [[], [], [], [], []]
                            # DTT.CompBlurDataList = [[],[],[],[],[],[]] if CodeSetup == "CONE" else [[], [], [], [], []]
                            DTT.CompResultList = ["10000000000000"] * (6 if CodeSetup == "CONE" else 5)
                            self.ModbusSignalSend(Mode = "STATE", Value = 0)
                        if self.ReworkMode == True:
                            logger.info("[SIGNAL] Rework Session Complete - Process Reset")
                            main_frame.update_signal('Rework Session Complete')
                            self.ReworkMode = False
                            self.ManualMode = False
                            self.ModbusSignalSend(Mode = "STATE", Value = 0)
                        if self.EmergencyMode == True:
                            logger.info("[SIGNAL] Emergency Session Complete - Process Reset")
                            main_frame.update_signal('Emergency Session Complete')
                            self.EmergencyMode = False
                            self.ModbusSignalSend(Mode = "RECONNECT")
                            self.ModbusSignalSend(Mode = "ALIVE", Value = 1)
                            self.ModbusSignalSend(Mode = "READY", Value = 1)
                    elif self.recvModbusChecker[8] == 1:
                        print('[Signal] Modbus Vision Modify Signal ON(Vision Process Reset - 1) Check')
                        logger.info('[Signal] Modbus Vision Modify Signal ON(Vision Process Reset - 1) Check')
                        main_frame.update_signal("Vision Process Reset recv")
                        self.ModbusSignalSend(Mode = "READY", Value = 0)
                        self.ModbusSignalSend(Mode = "BUSY", Value = 1)
                        self.ModbusSignalSend(Mode = "STATE", Value = 4)
                        TSD.sendAllClient("CLIENTRESET")
                        DTT.frontDataList = [
                            [DTT.NoneData, DTT.NoneData, DTT.NoneData],
                            [DTT.NoneData, DTT.NoneData, DTT.NoneData],
                            [DTT.NoneData, DTT.NoneData, DTT.NoneData],
                            [DTT.NoneData, DTT.NoneData, DTT.NoneData],
                        ]
                        DTT.frontImageList = [[[], [], []], [[], [], []], [[], [], []], [[], [], []]]
                        DTT.frontBlurDataList = [[[],[],[]], [[],[],[]], [[],[],[]], [[],[],[]]]
                        DTT.CompImageList = [[], [], [], [], [], []] if CodeSetup == "CONE" else [[], [], [], [], []]
                        # DTT.CompBlurDataList = [[],[],[],[],[],[]] if CodeSetup == "CONE" else [[], [], [], [], []]
                        DTT.CompResultList = ["10000000000000"] * (6 if CodeSetup == "CONE" else 5)

                        for i in range(6 if CodeSetup == "CONE" else 5):
                            main_frame.main_canvas.itemconfig(main_frame.main_cam[i], image="")
                            main_frame.main_canvas.itemconfig(main_frame.main_ok[i], state="hidden")
                            main_frame.main_canvas.itemconfig(main_frame.main_ng[i], state="hidden")

                        for i in range(10):
                            main_frame.main_canvas.itemconfig(main_frame.bad_image1[i], image="")
                            main_frame.main_canvas.itemconfig(main_frame.bad_text1[i], text="")
                            main_frame.main_canvas.itemconfig(main_frame.bad_image2[i], image="")
                            main_frame.main_canvas.itemconfig(main_frame.bad_text2[i], text="")
                        DTT.nowbadTypeIndex1 = 10
                        DTT.nowbadTypeIndex2 = 10
                        ngcage_frame.ngCageUpdate1 = [
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                        ]
                        ngcage_frame.ngCageUpdate2 = [
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                            [False, [None, None, None, None, None], [None, None, None, None, None], [None],],
                        ]
                        main_frame.bad_text1_inputData = [None] * 10
                        main_frame.bad_text2_inputData = [None] * 10

                        main_frame.mainbadCheckList = [[0] * 25, [0] * 25, [0] * 25, [0] * 25, [0] * 25, [0] * 25] if CodeSetup == "CONE" else [[0] * 25, [0] * 25, [0] * 25, [0] * 25, [0] * 25]
                        main_frame.mainbadCheckListboxData = [[], [], [], [], [], []] if CodeSetup == "CONE" else [[], [], [], [], []]
                        main_frame.mainbadCheckListTotalcount = [0, 0, 0, 0, 0, 0] if CodeSetup == "CONE" else [0, 0, 0, 0, 0]
                        DTT.AllDataRecv = 0
                        DTT.AllImageRecv = 0

                        for i in range(6 if CodeSetup == "CONE" else 5):
                            main_frame.main_canvas.itemconfig(main_frame.CameraBlur[i], state = 'hidden')
                        CountDataSave()
                        time.sleep(1)
                        self.ModbusSignalSend(Mode = "READY", Value = 1)
                        self.ModbusSignalSend(Mode = "BUSY", Value = 0)
                        self.ModbusSignalSend(Mode = "STATE", Value = 0)
                        # self.SendDataTrim("W", "D3003", "1")

                    elif self.recvModbusChecker[8] == 2:
                        print('[Signal] Modbus Vision Modify Signal ON(Vision Count Reset - 2) Check')
                        logger.info('[Signal] Modbus Vision Modify Signal ON(Vision Count Reset - 2) Check')
                        self.ModbusSignalSend(Mode = "READY", Value = 0)
                        self.ModbusSignalSend(Mode = "BUSY", Value = 1)
                        self.ModbusSignalSend(Mode = "STATE", Value = 4)
                        main_frame.update_signal("Vision Count Reset recv")
                        main_frame.total_count = main_frame.ok_count = main_frame.ng_count = 0
                        main_frame.main_canvas.itemconfig(main_frame.totalText, text=main_frame.total_count)
                        main_frame.main_canvas.itemconfig(main_frame.okText, text=main_frame.ok_count)
                        main_frame.main_canvas.itemconfig(main_frame.ngText, text=main_frame.ng_count)
                        main_frame.mainbadCheckList = [[0] * 25, [0] * 25, [0] * 25, [0] * 25, [0] * 25, [0] * 25] if CodeSetup == "CONE" else [[0] * 25, [0] * 25, [0] * 25, [0] * 25, [0] * 25]
                        main_frame.mainbadCheckListboxData = [[], [], [], [], [], []] if CodeSetup == "CONE" else [[], [], [], [], []]
                        main_frame.mainbadCheckListTotalcount = [0, 0, 0, 0, 0, 0] if CodeSetup == "CONE" else [0, 0, 0, 0, 0]
                        DTT.AllDataRecv = 0
                        DTT.AllImageRecv = 0
                        
                        CountDataSave()
                        time.sleep(1)
                        self.ModbusSignalSend(Mode = "READY", Value = 1)
                        self.ModbusSignalSend(Mode = "BUSY", Value = 0)
                        self.ModbusSignalSend(Mode = "STATE", Value = 0)
                        # self.SendDataTrim("W", "D3003", "1")

                    elif self.recvModbusChecker[8] == 3:  # 리부트 : 컴퓨터 재실행
                        print('[Signal] Modbus Vision Modify Signal ON(Vision Reboot - 3) Check')
                        logger.info('[Signal] Modbus Vision Modify Signal ON(Vision Reboot - 3) Check')
                        main_frame.update_signal("Vision Reboot recv")
                        CountDataSave()
                        TSD.sendAllClient("REBOOT")
                        self.ModbusSignalSend(Mode = 'REBOOT')

                        os.system("shutdown -r") 

                    elif self.recvModbusChecker[8] == 4:
                        print('[Signal] Modbus Vision Modify Signal ON(Dispose Work - 4) Check')
                        main_frame.update_signal("Modbus Vision Dispose Work On")
                        logger.info('[Signal] Modbus Vision Modify Signal ON(Dispose Work - 4) Check')
                        self.DisposeMode = True
                        self.ModbusSignalSend(Mode = "STATE", Value = 3)

                    elif self.recvModbusChecker[8] == 5:
                        print('[Signal] Modbus Vision Modify Signal ON(NGCAGERESET Work - 5) Check')
                        main_frame.update_signal("Modbus Vision NGCAGERESET Work On")
                        logger.info('[Signal] Modbus Vision Modify Signal ON(NGCAGERESET Work - 5) Check')
                        self.ModbusSignalSend(Mode = "READY", Value = 0)
                        self.ModbusSignalSend(Mode = "BUSY", Value = 1)
                        self.ModbusSignalSend(Mode = "STATE", Value = 4)
                        # self.MxSignalCheck[8] = True
                        # self.Modbusclient.write_coils(address = 31088, count = 2, values = [False,True])

                        for i in range(6 if CodeSetup == "CONE" else 5):
                            main_frame.main_canvas.itemconfig(main_frame.main_cam[i], image="")
                            main_frame.main_canvas.itemconfig(main_frame.main_ok[i], state="hidden")
                            main_frame.main_canvas.itemconfig(main_frame.main_ng[i], state="hidden")
                        for i in range(10):
                            main_frame.main_canvas.itemconfig(main_frame.bad_image1[i], image="")
                            main_frame.main_canvas.itemconfig(main_frame.bad_text1[i], text="")
                            main_frame.main_canvas.itemconfig(main_frame.bad_image2[i], image="")
                            main_frame.main_canvas.itemconfig(main_frame.bad_text2[i], text="")
                        DTT.nowbadTypeIndex1 = 10
                        DTT.nowbadTypeIndex2 = 10
                        ngcage_frame.ngCageUpdate1 = [[False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]]]
                        ngcage_frame.ngCageUpdate2 = [[False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]]]
                        main_frame.bad_text1_inputData = [None] * 10
                        main_frame.bad_text2_inputData = [None] * 10

                        main_frame.total_count -= main_frame.ng_count
                        main_frame.ng_count = 0
                        main_frame.main_canvas.itemconfig(main_frame.totalText, text=main_frame.total_count)
                        main_frame.main_canvas.itemconfig(main_frame.okText, text=main_frame.ok_count)
                        main_frame.main_canvas.itemconfig(main_frame.ngText, text=main_frame.ng_count)
                        main_frame.mainbadCheckList = [[0] * 25, [0] * 25, [0] * 25, [0] * 25, [0] * 25, [0] * 25] if CodeSetup == "CONE" else [[0] * 25, [0] * 25, [0] * 25, [0] * 25, [0] * 25]
                        main_frame.mainbadCheckListboxData = [[], [], [], [], [], []] if CodeSetup == "CONE" else [[], [], [], [], []]
                        main_frame.mainbadCheckListTotalcount = [0, 0, 0, 0, 0, 0] if CodeSetup == "CONE" else [0, 0, 0, 0, 0]
                        DTT.AllDataRecv = 0
                        DTT.AllImageRecv = 0
                        # for i in range(6 if CodeSetup == "CONE" else 5):
                        #     main_frame.main_canvas.itemconfig(main_frame.CameraBlur[i], state = 'hidden')
                        CountDataSave()
                        time.sleep(1)
                        # self.SendDataTrim("W", "D3003", "1")
                        self.ModbusSignalSend(Mode = "READY", Value = 1)
                        self.ModbusSignalSend(Mode = "BUSY", Value = 0)
                        self.ModbusSignalSend(Mode = "STATE", Value = 0)

                    elif self.recvModbusChecker[8] == 6:
                        print('[Signal] Modbus Vision Modify Signal ON(NGCAGEBUFFRESET Work - 6) Check')
                        main_frame.update_signal("Modbus Vision NGCAGEBUFFRESET Work On")
                        logger.info('[Signal] Modbus Vision Modify Signal ON(NGCAGEBUFFRESET Work - 6) Check')
                        self.ModbusSignalSend(Mode = "READY", Value = 0)
                        self.ModbusSignalSend(Mode = "BUSY", Value = 1)
                        self.ModbusSignalSend(Mode = "STATE", Value = 6)
                        # self.MxSignalCheck[8] = True
                        # self.Modbusclient.write_coils(address = 31088, count = 2, values = [False,True])

                        for i in range(6 if CodeSetup == "CONE" else 5):
                            main_frame.main_canvas.itemconfig(main_frame.main_cam[i], image="")
                            main_frame.main_canvas.itemconfig(main_frame.main_ok[i], state="hidden")
                            main_frame.main_canvas.itemconfig(main_frame.main_ng[i], state="hidden")
                        for i in range(10):
                            main_frame.main_canvas.itemconfig(main_frame.bad_image1[i], image="")
                            main_frame.main_canvas.itemconfig(main_frame.bad_text1[i], text="")
                            main_frame.main_canvas.itemconfig(main_frame.bad_image2[i], image="")
                            main_frame.main_canvas.itemconfig(main_frame.bad_text2[i], text="")
                        DTT.nowbadTypeIndex1 = 10
                        DTT.nowbadTypeIndex2 = 10
                        ngcage_frame.ngCageUpdate1 = [[False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]]]
                        ngcage_frame.ngCageUpdate2 = [[False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]]]
                        main_frame.bad_text1_inputData = [None] * 10
                        main_frame.bad_text2_inputData = [None] * 10

                        # main_frame.total_count -= main_frame.ng_count
                        # main_frame.ng_count = 0
                        # main_frame.main_canvas.itemconfig(main_frame.totalText, text=main_frame.total_count)
                        # main_frame.main_canvas.itemconfig(main_frame.okText, text=main_frame.ok_count)
                        # main_frame.main_canvas.itemconfig(main_frame.ngText, text=main_frame.ng_count)
                        # main_frame.mainbadCheckList = [[0] * 25, [0] * 25, [0] * 25, [0] * 25, [0] * 25, [0] * 25] if CodeSetup == "CONE" else [[0] * 25, [0] * 25, [0] * 25, [0] * 25, [0] * 25]
                        # main_frame.mainbadCheckListboxData = [[], [], [], [], [], []] if CodeSetup == "CONE" else [[], [], [], [], []]
                        # main_frame.mainbadCheckListTotalcount = [0, 0, 0, 0, 0, 0] if CodeSetup == "CONE" else [0, 0, 0, 0, 0]
                        DTT.AllDataRecv = 0
                        DTT.AllImageRecv = 0
                        # for i in range(6 if CodeSetup == "CONE" else 5):
                        #     main_frame.main_canvas.itemconfig(main_frame.CameraBlur[i], state = 'hidden')
                        CountDataSave()
                        time.sleep(1)
                        # self.SendDataTrim("W", "D3003", "1")
                        self.ModbusSignalSend(Mode = "READY", Value = 1)
                        self.ModbusSignalSend(Mode = "BUSY", Value = 0)
                        self.ModbusSignalSend(Mode = "STATE", Value = 0)
                    elif self.recvModbusChecker[8] == 7:
                        print('[Signal] Modbus Vision Modify Signal ON(Rework - 7) Check')
                        main_frame.update_signal("Modbus Vision Rework On")
                        logger.info('[Signal] Modbus Vision Modify Signal ON(Rework - 7) Check')
                        self.ManualMode = True
                        self.ReworkMode = True
                        self.ModbusSignalSend(Mode = "STATE", Value = 7)
                    elif self.recvModbusChecker[8] == 8:
                        print('[Signal] Modbus Vision Modify Signal ON(Emergency - 8) Check')
                        main_frame.update_signal("Modbus Vision Emergency On")
                        logger.info('[Signal] Modbus Vision Modify Signal ON(Emergency - 8) Check')
                        DTT.frontDataList = [
                            [DTT.NoneData, DTT.NoneData, DTT.NoneData],
                            [DTT.NoneData, DTT.NoneData, DTT.NoneData],
                            [DTT.NoneData, DTT.NoneData, DTT.NoneData],
                            [DTT.NoneData, DTT.NoneData, DTT.NoneData],
                        ]
                        DTT.frontImageList = [[[], [], []], [[], [], []], [[], [], []], [[], [], []]]
                        DTT.frontBlurDataList = [[[],[],[]], [[],[],[]], [[],[],[]], [[],[],[]]]
                        DTT.CompImageList = [[], [], [], [], [], []] if CodeSetup == "CONE" else [[], [], [], [], []]
                        # DTT.CompBlurDataList = [[],[],[],[],[],[]] if CodeSetup == "CONE" else [[], [], [], [], []]
                        DTT.CompResultList = ["10000000000000"] * (6 if CodeSetup == "CONE" else 5)
                        DTT.AllDataRecv = 0
                        DTT.AllImageRecv = 0
                        self.EmergencyMode = True
                        self.ModbusSignalSend(Mode = "STATE", Value = 8)
                    else:
                        print('[Signal] Modbus Vision Modify Signal Error')
                        logger.info(f'[Signal] Modbus Vision Modify Signal Error - {self.ModbusDefaultRecvAddress+8}Address RecvData - {self.recvModbusChecker[8]}')
                    self.ModbusSessionChecker[8] = False
                # print(f'One Check Cycle Time - {time.time() - startTime}')

                if TE.reset == True: # OH
                    print('[TRIGGER] 08:00 Ngcage Reset')
                    main_frame.update_signal("[TRIGGER] 08:00 Ngcage Reset")
                    logger.info('[TRIGGER] 08:00 Ngcage Reset')

                    print("ng list delete")
                    for i in range(10):
                        main_frame.main_canvas.itemconfig(main_frame.bad_image1[i], image="")
                        main_frame.main_canvas.itemconfig(main_frame.bad_text1[i], text="")
                        main_frame.main_canvas.itemconfig(main_frame.bad_image2[i], image="")
                        main_frame.main_canvas.itemconfig(main_frame.bad_text2[i], text="")
                    DTT.nowbadTypeIndex1 = 10
                    DTT.nowbadTypeIndex2 = 10
                    ngcage_frame.ngCageUpdate1 = [[False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]]]
                    ngcage_frame.ngCageUpdate2 = [[False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]], [False, [None, None, None, None, None], [None, None, None, None, None], [None]]]
                    main_frame.bad_text1_inputData = [None] * 10
                    main_frame.bad_text2_inputData = [None] * 10

                    main_frame.mainbadCheckList = [[0] * 25, [0] * 25, [0] * 25, [0] * 25, [0] * 25, [0] * 25] if CodeSetup == "CONE" else [[0] * 25, [0] * 25, [0] * 25, [0] * 25, [0] * 25]
                    main_frame.mainbadCheckListboxData = [[], [], [], [], [], []] if CodeSetup == "CONE" else [[], [], [], [], []]
                    main_frame.mainbadCheckListTotalcount = [0, 0, 0, 0, 0, 0] if CodeSetup == "CONE" else [0, 0, 0, 0, 0]

                    CountDataSave()
                    TE.reset = False

            except:
                print(f'Modbus Communication Error - {traceback.format_exc()}')
                logger.info(f'Modbus Communication Error - {traceback.format_exc()}')
                self.ConnectModbus()

        
    def connectRetry(self):
        while True:
            try:
                try:
                    self.client_socket.close()
                except:
                    pass
                print("Notice : [Socket Connecting. wait please]")
                logger.info("Notice : [Socket Connecting. wait please]")
                self.client_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                self.client_socket.connect((HOST, PORT))
                print("Notice : [Socket Connected]")
                logger.info("Notice : [Socket Connect Complete]")
                # self.SendDataTrim("W", "D3000", "0")
                # self.SendDataTrim("W", "D3001", "0")
                # self.SendDataTrim("W", "D3002", "0")
                # self.SendDataTrim("W", "D3003", "0")
                # self.SendDataTrim("W", "D3006", "0")
                # self.SendDataTrim("W", "D3007", "0")
                # self.SendDataTrim("W", "D3016", "0")
                self.ModbusSignalSend(Mode = "RESET")
                self.ModbusSignalSend(Mode = "ALIVE", Value = 1)
                print(f'Modbus Signal Send > Mode - ALIVE, SendValue - 1')
                logger.info(f'Modbus Signal Send > Mode - ALIVE, SendValue - 1')
                break
            except:
                time.sleep(1)
                print(f"Notice : [Socket Connect Error] {traceback.format_exc()}")
                logger.info("Notice : [Socket Connect Error]")


from multiprocessing import Process
from multiprocessing import Queue as MQueue


def SaveImages(MQ):
    while True:
        if not MQ.empty():
            try:
                # startTime = time.time()
                data = MQ.get()
                foldername = data[0]
                filename = data[1]

                if not (os.path.isdir(foldername)):
                    os.makedirs(os.path.join(foldername))

                cv2.imwrite(foldername + "/" + filename, data[2])
                # print('{} image saved'.format(filename))
                # print("[INFO] Elapesed time : ", time.time() - startTime)

            except Exception as ex:
                # print('M - ', ex)
                pass
        else:
            pass
            time.sleep(1)


class MysqlDB:
    def __init__(self):
        self.host = "localhost"
        self.user = "root"
        self.password = "yous7051"
        self.dbname = "bearingart"

    def readSql_View(self, ModelName = None, StartSearch = None, EndSearch = None):
        db = pymysql.connect(host=self.host, user=self.user, password=self.password, db=self.dbname, charset="utf8")
        curs = db.cursor()
        sql = "select * from production_status"
        addvalue = False

        if ModelName is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "model='{}'".format(ModelName)

        if StartSearch is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "DateHour >= '{}'".format(StartSearch)

        if EndSearch is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "DateHour <= '{}'".format(EndSearch)
        
        print(sql)

        curs.execute(sql)
        rows = curs.fetchall()
        
        # self.writeExcel(rows, text, dirName=dirs)
        db.close()

        # print(rows)

        return rows

    def readSql_Load(self, ModelName = None, CheckDate=None, CheckHour=None):
        db = pymysql.connect(host=self.host, user=self.user, password=self.password, db=self.dbname, charset="utf8")
        curs = db.cursor()
        sql = "select * from production_status"
        addvalue = False

        if ModelName is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "model='{}'".format(ModelName)

        if CheckDate is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "UpdateDate >= '{}'".format(CheckDate)

        if CheckDate is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "DateHour='{}'".format(CheckHour)

        print(sql)

        curs.execute(sql)
        rows = curs.fetchall()
        
        # self.writeExcel(rows, text, dirName=dirs)
        db.close()

        # print(rows)

        return rows

    def updateSql(self, updateValue):
        db = pymysql.connect(host=self.host, user=self.user, password=self.password, db=self.dbname, charset="utf8")
        param = values = ""
        curs = db.cursor()

        sql = f"UPDATE production_status SET Total = {updateValue[0]}, Ok = {updateValue[1]}, Ng = {updateValue[2]}, CriticalNg = {updateValue[3]} WHERE UpdateDate = '{updateValue[4]}' AND model = '{updateValue[5]}' AND DateHour = '{updateValue[6]}'"
        print(sql,'\n\n\n')
        curs.execute(sql)
        db.commit()
        db.close()
        pass

    #'model','total','ok','ng','Date',

    def writeSql_count(self, valueList):
        db = pymysql.connect(host=self.host, user=self.user, password=self.password, db=self.dbname, charset="utf8")
        param = values = ""
        curs = db.cursor()
        for name, val in valueList:
            param = param + name + ", "
            values = values + val + ", "

        sql = "INSERT INTO production_status ({}) VALUES ({})".format(param[:-2], values[:-2])
                #값을 넣겠다 / ~~에 / ELENTEC_COUNT테이블에 / ~~필드에 / 무슨 값을 / ~~값
        print(sql)
        curs.execute(sql)
        db.commit()
        db.close()

    def dbCount_Process(self, model, result):

        today = str(datetime.date.today())
        today_Hour = str(datetime.datetime.now())[0:13]

        checkSql = self.readSql_Load(model, today, today_Hour)

        try:
            #db 데이터 존재 o
            for i, index in enumerate(checkSql[0]):
                if i == 2:
                    totalCount = int(index)
                if i == 3:
                    okCount = int(index)
                if i == 4:
                    ngCount = int(index)
                if i == 5:
                    CriticalNg = int(index)

            if result == 'NG':
                totalCount += 1
                ngCount += 1
            elif result == 'CRI':
                totalCount += 1
                CriticalNg += 1
            else:
                totalCount += 1
                okCount += 1
            
            updateValue = [totalCount, okCount, ngCount, CriticalNg, today, model, today_Hour]
            self.updateSql(updateValue)
        except:
            #db 데이터 존재 x
            print('None Data')
            totalCount = 0
            okCount = 0
            ngCount = 0
            CriticalNg = 0
            inputDBData = {
                ("model", f"'{model}'"),
                ("Total", f"'{totalCount}'"),
                ("Ok", f"'{okCount}'"),
                ("Ng", f"'{ngCount}'"),
                ("CriticalNg", f"'{CriticalNg}'"),
                ("UpdateDate", f"now()"),
                ("DateHour", f"'{today_Hour}'")
            }
            self.writeSql_count(inputDBData)

            if result == 'NG':
                totalCount += 1
                ngCount += 1
            elif result == 'CRI':
                totalCount += 1
                CriticalNg += 1
            else:
                totalCount += 1
                okCount += 1
            updateValue = [totalCount, okCount, ngCount, CriticalNg, today, model, today_Hour]
            self.updateSql(updateValue)

    # 토탈, ok, ng, 검사결과, 검사시간, 결과이미지 경로
    def readSql_Cone(
        self, ModelName=None, Result1=None, Result2=None, Result3=None, Result4=None, Result5=None, Result6=None, ImagePath1=None, ImagePath2=None, ImagePath3=None, ImagePath4=None, ImagePath5=None, ImagePath6=None, StartDate=None, EndDate=None,
    ):
        db = pymysql.connect(host=self.host, user=self.user, password=self.password, db=self.dbname, charset="utf8")
        curs = db.cursor()
        sql = "select * from data"
        addvalue = False
        DataChangeCheck = False

        if Result1 == DTT.OKdata or Result2 == DTT.OKdata or Result3 == DTT.OKdata or Result4 == DTT.OKdata or Result5 == DTT.OKdata or Result6 == DTT.OKdata:
            DataChangeCheck = True
        if Result1 == "NG" or Result2 == "NG" or Result3 == "NG" or Result4 == "NG" or Result5 == "NG" or Result6 == "NG":
            DataChangeCheck = None

        if ModelName is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "ModelName='{}'".format(ModelName)

        if Result1 is not None:
            add = " WHERE " if addvalue is False else " AND "

            if DataChangeCheck == None:
                # add = " WHERE " if addvalue is False else " OR "
                sql = sql + add + "(Result1 LIKE '%1%'"
            else:
                add = " WHERE " if addvalue is False else " AND "
                sql = sql + add + "(Result1='{}'".format(Result1)
            addvalue = True

        if Result2 is not None:
            if DataChangeCheck == True:
                add = " WHERE " if addvalue is False else " AND "
                sql = sql + add + "Result2='{}'".format(Result2)
            elif DataChangeCheck == None:
                add = " WHERE " if addvalue is False else " OR "
                sql = sql + add + "Result2 LIKE '%1%'"
            else:
                add = " WHERE " if addvalue is False else " OR "
                sql = sql + add + "Result2='{}'".format(Result2)
            addvalue = True

        if Result3 is not None:
            if DataChangeCheck == True:
                add = " WHERE " if addvalue is False else " AND "
                sql = sql + add + "Result3='{}'".format(Result3)
            elif DataChangeCheck == None:
                add = " WHERE " if addvalue is False else " OR "
                sql = sql + add + "Result3 LIKE '%1%'"
            else:
                add = " WHERE " if addvalue is False else " OR "
                sql = sql + add + "Result3='{}'".format(Result3)
            addvalue = True

        if Result4 is not None:
            if DataChangeCheck == True:
                add = " WHERE " if addvalue is False else " AND "
                sql = sql + add + "Result4='{}'".format(Result4)
            elif DataChangeCheck == None:
                add = " WHERE " if addvalue is False else " OR "
                sql = sql + add + "Result4 LIKE '%1%'"
            else:
                add = " WHERE " if addvalue is False else " OR "
                sql = sql + add + "Result4='{}'".format(Result4)
            addvalue = True

        if Result5 is not None:
            if DataChangeCheck == True:
                add = " WHERE " if addvalue is False else " AND "
                sql = sql + add + "Result5='{}'".format(Result5)
            elif DataChangeCheck == None:
                add = " WHERE " if addvalue is False else " OR "
                sql = sql + add + "Result5 LIKE '%1%'"
            else:
                add = " WHERE " if addvalue is False else " OR "
                sql = sql + add + "Result5='{}'".format(Result5)
            addvalue = True

        if Result6 is not None:
            if DataChangeCheck == True:
                add = " WHERE " if addvalue is False else " AND "
                sql = sql + add + "Result6='{}')".format(Result6)
            elif DataChangeCheck == None:
                add = " WHERE " if addvalue is False else " OR "
                sql = sql + add + "Result6 LIKE '%1%')"
            else:
                add = " WHERE " if addvalue is False else " OR "
                sql = sql + add + "Result6='{}')".format(Result6)
            addvalue = True

        if ImagePath1 is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "ImagePath1='{}'".format(ImagePath1)

        if ImagePath2 is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "ImagePath2='{}'".format(ImagePath2)

        if ImagePath3 is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "ImagePath3='{}'".format(ImagePath3)

        if ImagePath4 is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "ImagePath4='{}'".format(ImagePath4)

        if ImagePath5 is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "ImagePath5='{}'".format(ImagePath5)

        if ImagePath6 is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "ImagePath6='{}'".format(ImagePath6)

        if StartDate is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "InputTime >= '{}'".format(StartDate)

        if EndDate is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            if len(EndDate) > 10:
                sql = sql + add + "InputTime <= '{}'".format(EndDate)
            else:
                sql = sql + add + "InputTime <= '{} 23:59:59'".format(EndDate)

        sql = sql + " order by LotNum DESC LIMIT 0,10000"

        # print(sql)

        curs.execute(sql)
        rows = curs.fetchall()

        # self.writeExcel(rows, text, dirName=dirs)
        db.close()

        # print(rows)

        return rows

    def readSql_Cup(
        self, ModelName=None, Result1=None, Result2=None, Result3=None, Result4=None, Result5=None, ImagePath1=None, ImagePath2=None, ImagePath3=None, ImagePath4=None, ImagePath5=None, StartDate=None, EndDate=None,
    ):
        db = pymysql.connect(host=self.host, user=self.user, password=self.password, db=self.dbname, charset="utf8")
        curs = db.cursor()
        sql = "select * from data"
        addvalue = False
        DataChangeCheck = None

        if Result1 == DTT.OKdata or Result2 == DTT.OKdata or Result3 == DTT.OKdata or Result4 == DTT.OKdata or Result5 == DTT.OKdata:
            DataChangeCheck = True
        if Result1 == "NG" or Result2 == "NG" or Result3 == "NG" or Result4 == "NG" or Result5 == "NG":
            DataChangeCheck = False

        if ModelName is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "ModelName='{}'".format(ModelName)

        if Result1 is not None:
            add = " WHERE " if addvalue is False else " AND "

            if DataChangeCheck == True:
                # add = " WHERE " if addvalue is False else " OR "
                add = " WHERE " if addvalue is False else " AND "
                sql = sql + add + "(Result1='{}'".format(Result1)
            elif DataChangeCheck == None:
                add = " WHERE " if addvalue is False else " AND "
                sql = sql + add + "(Result1='{}'".format(Result1)
            else:
                sql = sql + add + "(Result1 LIKE '%1%'"
            addvalue = True

        if Result2 is not None:
            if DataChangeCheck == True:
                add = " WHERE " if addvalue is False else " AND "
                sql = sql + add + "Result2='{}'".format(Result2)
            elif DataChangeCheck == False:
                add = " WHERE " if addvalue is False else " OR "
                sql = sql + add + "Result2 LIKE '%1%'"
            else:
                add = " WHERE " if addvalue is False else " OR "
                sql = sql + add + "Result2='{}'".format(Result2)
            addvalue = True

        if Result3 is not None:
            if DataChangeCheck == True:
                add = " WHERE " if addvalue is False else " AND "
                sql = sql + add + "Result3='{}'".format(Result3)
            elif DataChangeCheck == False:
                add = " WHERE " if addvalue is False else " OR "
                sql = sql + add + "Result3 LIKE '%1%'"
            else:
                add = " WHERE " if addvalue is False else " OR "
                sql = sql + add + "Result3='{}'".format(Result3)
            addvalue = True

        if Result4 is not None:
            if DataChangeCheck == True:
                add = " WHERE " if addvalue is False else " AND "
                sql = sql + add + "Result4='{}'".format(Result4)
            elif DataChangeCheck == False:
                add = " WHERE " if addvalue is False else " OR "
                sql = sql + add + "Result4 LIKE '%1%'"
            else:
                add = " WHERE " if addvalue is False else " OR "
                sql = sql + add + "Result4='{}'".format(Result4)
            addvalue = True

        if Result5 is not None:
            if DataChangeCheck == True:
                add = " WHERE " if addvalue is False else " AND "
                sql = sql + add + "Result5='{}')".format(Result5)
            elif DataChangeCheck == False:
                add = " WHERE " if addvalue is False else " OR "
                sql = sql + add + "Result5 LIKE '%1%')"
            else:
                add = " WHERE " if addvalue is False else " OR "
                sql = sql + add + "Result5='{}')".format(Result5)
            addvalue = True

        if ImagePath1 is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "ImagePath1='{}'".format(ImagePath1)

        if ImagePath2 is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "ImagePath2='{}'".format(ImagePath2)

        if ImagePath3 is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "ImagePath3='{}'".format(ImagePath3)

        if ImagePath4 is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "ImagePath4='{}'".format(ImagePath4)

        if ImagePath5 is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "ImagePath5='{}'".format(ImagePath5)

        if StartDate is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "InputTime >= '{}'".format(StartDate)

        if EndDate is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            if len(EndDate) > 10:
                sql = sql + add + "InputTime <= '{}'".format(EndDate)
            else:
                sql = sql + add + "InputTime <= '{} 23:59:59'".format(EndDate)

        sql = sql + " order by LotNum DESC LIMIT 0,10000"

        print(sql)

        curs.execute(sql)
        rows = curs.fetchall()

        # self.writeExcel(rows, text, dirName=dirs)
        db.close()

        print(rows)

        return rows

    def writeSql(self, valueList):
        db = pymysql.connect(host=self.host, user=self.user, password=self.password, db=self.dbname, charset="utf8")
        param = values = ""
        curs = db.cursor()
        for name, val in valueList:
            param = param + name + ", "
            values = values + val + ", "

        sql = "INSERT INTO data ({}) VALUES ({})".format(param[:-2], values[:-2])

        curs.execute(sql)
        db.commit()
        db.close()

########################################### 엑셀 ##########################################################################################

class TimeExcel:
    def __init__(self):
        self.auto_reset = False
        self.reset = False # OH
        self.send_mail_success = 'yet' # yet : 대기, true : 성공, false : 실패

    def run(self):
        while 1 :
            try :
                time.sleep(0.8) 
                save_time = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
                save_time = save_time.split("-")

                hour = save_time[3] 
                minute = save_time[4]
                sec = save_time[5]

                nowtime = hour + minute + sec
                
                # print(nowtime)
                if nowtime == '080000' :
                    self.reset = True
                    print('08시 카운트 리셋')
                    Email.SendingListUp()
                    logger.info(f"08시 카운트 리셋")
                    self.auto_reset = True
                    main_frame.total_count = main_frame.ok_count = main_frame.ng_count = 0
                    main_frame.main_canvas.itemconfig(main_frame.totalText, text=main_frame.total_count)
                    main_frame.main_canvas.itemconfig(main_frame.okText, text=main_frame.ok_count)
                    main_frame.main_canvas.itemconfig(main_frame.ngText, text=main_frame.ng_count)

                    WriteData = [main_frame.total_count, main_frame.ok_count, main_frame.ng_count]

                    try:
                        try:
                            if not (os.path.isdir(f"CountSave")):
                                os.makedirs(os.path.join(f"CountSave"))
                        except OSError as e:
                            pass

                        with open(f"CountSave/Count.pickle", "wb") as file:
                            pickle.dump(WriteData, file)
                    except Exception as ex:
                        logger.info(f"Warning : 피클파일 저장 에러 - {ex} \n {traceback.format_exc()}")

                    d_today = datetime.date.today()
                    PXM.Today = d_today.strftime('%Y-%m-%d')
                    PXM.Yesterday = d_today - datetime.timedelta(1)

                    PXM.ExcelDataLoad('BearingArt_DaysReport.xlsx')
                    returnDict = DB.DB_DataLoad_Refine()
                    # PXM.ExcelSheetChoice('Sheet1')

                    dictKeyList = list(returnDict.keys())
                    # print(dictKeyList)
                    
                    #db에 생산이력 없을 경우
                    if dictKeyList == []:
                        i = 0
                        x = 0
                        PXM.ExcelSheetChoice('전수검사기 검사 실적')
                            
                        PXM.ExcelDataWrite('B1', f'▣ {PXM.lineName} {PXM.lineType} 전수검사기 검사 실적')
                        PXM.ExcelDataWrite('J3', f'{PXM.Today}')
                        PXM.ExcelDataWrite('B8', f'{PXM.lineName}\n{PXM.lineType}')
                        PXM.ExcelDataWrite('C8', f'{PXM.Yesterday} 08:00\n~\n{PXM.Today} 07:59')
                        PXM.ExcelDataWrite('D8' if i == 0 else 'D'+str(8+i), '생산 없음')
                        PXM.ExcelDataWrite('E8' if i == 0 else 'E'+str(8+i), 0)
                        PXM.ExcelDataWrite('F8' if i == 0 else 'F'+str(8+i), 0)
                        PXM.ExcelDataWrite('G8' if i == 0 else 'G'+str(8+i), 0)
                        PXM.ExcelDataWrite('H8' if i == 0 else 'H'+str(8+i), 0)
                        PXM.ExcelDataWrite('I8' if i == 0 else 'I'+str(8+i), 0)

                        PXM.ExcelDataWrite('B19', f'{PXM.lineName}\n{PXM.lineType}')
                        PXM.ExcelDataWrite('C19', f'{PXM.Yesterday} 08:00\n~\n{PXM.Today} 07:59')
                        PXM.ExcelDataWrite('D19' if i == 0 else 'D'+str(19+i), '생산 없음')
                        PXM.ExcelDataWrite('E19' if i == 0 else 'E'+str(19+i), 0)
                        PXM.ExcelDataWrite('F19' if i == 0 else 'F'+str(19+i), 0)
                        PXM.ExcelDataWrite('G19' if i == 0 else 'G'+str(19+i), 0)
                        PXM.ExcelDataWrite('H19' if i == 0 else 'H'+str(19+i), 0)
                        PXM.ExcelDataWrite('I19' if i == 0 else 'I'+str(19+i), 0)
                        PXM.ExcelSheetChoice('전수검사기 불량 검사 상세 실적')
                        
                        PXM.ExcelDataWrite('B1', f'▣ {PXM.lineName} {PXM.lineType} 전수검사기 불량 검사 상세 실적')
                        PXM.ExcelDataWrite('B8', f'{PXM.lineName}\n{PXM.lineType}')
                        PXM.ExcelDataWrite('G3', f'{PXM.Today}')
                        pass
                    #db에 생산이력 존재할 경우
                    else:
                        UpdateLineCount = 0
                        IndexUpdate = 0
                        for i in range(len(dictKeyList)):
                            IndexUpdate += UpdateLineCount
                            PXM.ExcelSheetChoice('전수검사기 검사 실적')
                            
                            PXM.ExcelDataWrite('B1', f'▣ {PXM.lineName} {PXM.lineType} 전수검사기 검사 실적')
                            PXM.ExcelDataWrite('J3', f'{PXM.Today}')
                            PXM.ExcelDataWrite('B8', f'{PXM.lineName}\n{PXM.lineType}')
                            PXM.ExcelDataWrite('C8', f'{PXM.Yesterday} 08:00\n~\n{PXM.Today} 07:59')
                            PXM.ExcelDataWrite('D8' if i == 0 else 'D'+str(8+i), dictKeyList[i])
                            PXM.ExcelDataWrite('E8' if i == 0 else 'E'+str(8+i), returnDict[dictKeyList[i]]['TotalCount'])
                            PXM.ExcelDataWrite('F8' if i == 0 else 'F'+str(8+i), returnDict[dictKeyList[i]]['OKCount'])
                            PXM.ExcelDataWrite('G8' if i == 0 else 'G'+str(8+i), returnDict[dictKeyList[i]]['NGCount']+returnDict[dictKeyList[i]]['CriticalNGCount'])
                            PXM.ExcelDataWrite('H8' if i == 0 else 'H'+str(8+i), returnDict[dictKeyList[i]]['NGCount'])
                            PXM.ExcelDataWrite('I8' if i == 0 else 'I'+str(8+i), returnDict[dictKeyList[i]]['CriticalNGCount'])

                            PXM.ExcelDataWrite('B19', f'{PXM.lineName}\n{PXM.lineType}')
                            PXM.ExcelDataWrite('C19', f'{PXM.Yesterday} 08:00\n~\n{PXM.Today} 07:59')
                            PXM.ExcelDataWrite('D19' if i == 0 else 'D'+str(19+i), dictKeyList[i])
                            PXM.ExcelDataWrite('E19' if i == 0 else 'E'+str(19+i), returnDict[dictKeyList[i]]['CriticalNGCount'])
                            PXM.ExcelDataWrite('F19' if i == 0 else 'F'+str(19+i), returnDict[dictKeyList[i]]['CriticalNGPartCount'][0])
                            PXM.ExcelDataWrite('G19' if i == 0 else 'G'+str(19+i), returnDict[dictKeyList[i]]['CriticalNGPartCount'][1])
                            PXM.ExcelDataWrite('H19' if i == 0 else 'H'+str(19+i), returnDict[dictKeyList[i]]['CriticalNGPartCount'][2])
                            PXM.ExcelDataWrite('I19' if i == 0 else 'I'+str(19+i), returnDict[dictKeyList[i]]['CriticalNGPartCount'][3])
                            PXM.ExcelSheetChoice('전수검사기 불량 검사 상세 실적') 
                            
                            PXM.ExcelDataWrite('B1', f'▣ {PXM.lineName} {PXM.lineType} 전수검사기 불량 검사 상세 실적')
                            PXM.ExcelDataWrite('B8', f'{PXM.lineName}\n{PXM.lineType}')
                            PXM.ExcelDataWrite('G3', f'{PXM.Today}')
                            
                            passCount = 0
                            for IDX, x in enumerate(range(len(returnDict[dictKeyList[i]]['Product']))):
                                try:
                                    if 'None' in  returnDict[dictKeyList[i]]['Product'][x][2]:
                                        passCount +=1
                                        continue
                                except:
                                    passCount +=1
                                    continue
                                check_X = x-passCount
                                # print(returnDict[dictKeyList[i]]['Product'][x][2], type(returnDict[dictKeyList[i]]['Product'][x][2]))
                                PXM.ExcelDataWrite('C'+str(check_X+8+IndexUpdate), returnDict[dictKeyList[i]]['Product'][x][0])
                                PXM.ExcelDataWrite('E'+str(check_X+8+IndexUpdate), returnDict[dictKeyList[i]]['Product'][x][1])
                                PXM.ExcelDataWrite('F'+str(check_X+8+IndexUpdate), returnDict[dictKeyList[i]]['Product'][x][2])
                                # print(returnDict[dictKeyList[i]]['Product'][check_X][2])
                                if "S/F 누락" in returnDict[dictKeyList[i]]['Product'][x][2] or "롤러 누락" in returnDict[dictKeyList[i]]['Product'][x][2] or "각인 누락" in returnDict[dictKeyList[i]]['Product'][x][2] or "식별각인 누락" in returnDict[dictKeyList[i]]['Product'][x][2] or '케이지 깨짐' in returnDict[dictKeyList[i]]['Product'][x][2]:
                                    font = Font(name='굴림', color = 'ff0000')
                                    PXM.Worksheet['F'+str(check_X+8+IndexUpdate)].font = font
                                    PXM.ExcelDataWrite('D'+str(check_X+8+IndexUpdate), '중대 불량')
                                    PXM.Worksheet['D'+str(check_X+8+IndexUpdate)].font = font
                                else:
                                    PXM.ExcelDataWrite('D'+str(check_X+8+IndexUpdate), '일반 불량')
                                    pass
                                UpdateLineCount += 1

                    PXM.ExcelFileWrite()
                    
                    # 메일 전송 프로세스
                    self.send_mail_success = Email.mailSendProcess() # True : 성공, False : 실패
                    if self.send_mail_success == True:
                        self.send_mail_success = 'yet'
                        print("전송 완료")
                        self.remove_xlsx()
                    elif self.send_mail_success == False :
                        print("전송 실패")

                if self.send_mail_success == False :
                    time.sleep(600) # 10분 마다 재시도
                    print('메일 전송 재시도')
                    logger.info(f"mail send retry")
                    self.send_mail_success = Email.mailSendProcess()
                    if self.send_mail_success == True:
                        self.send_mail_success = 'yet'
                        print("전송 완료")
                        self.remove_xlsx()
                    elif self.send_mail_success == False :
                        print("전송 실패")

            except Exception as ex:
                print(ex)
                logger.info(f"Warning : 자동 카운트 리셋 메일 송부 에러 - {ex} \n {traceback.format_exc()}")

    # 5일전 엑셀 삭제
    def remove_xlsx(self) :
        try:
            os.remove(f'C:/Users/sinplat/Desktop/{my_folder}/{PXM.five_days_ago_file_name}).xlsx')
            print('엑셀 삭제 완료')
            logger.info('엑셀 삭제 완료')
        except :
            print(f'엑셀 삭제 실패 {traceback.format_exc()}')
            logger.info(f'엑셀 삭제 실패 {traceback.format_exc()}')

class OpenpyxlModul():
    def __init__(self):
        import datetime
        self.lineName = 'SS#8'
        self.lineType = 'CONE'

        d_today = datetime.date.today()
        self.Today = d_today.strftime('%Y-%m-%d')
        self.Yesterday = d_today - datetime.timedelta(1)
        self.five_days_ago = d_today - datetime.timedelta(5)

        # self.Today = '2022-10-10'
        # self.Yesterday = '2022-10-09'
        self.Workbook = None    #로딩 및 저장에 쓰이는 최종 데이터
        self.Worksheet = None   #작업이 진행될 워크시트
        self.LoadData = None
        self.writeExcelFilename = ''

        self.xlsx_file_name = '' # 일일현황 파일 이름
        
    def ExcelDataLoad(self, Path):
        try:
            #"C:\\Users\\ParkJH\\Desktop\\엑셀프로젝트\\test.xlsx"
            self.Workbook = load_workbook(Path, data_only=True)
        except:
            print(traceback.format_exc())

    def CreateSheet(self, SheetName):
        try:
            self.Workbook.create_sheet(SheetName)
        except:
            print(traceback.format_exc())

    def ExcelSheetChoice(self, SheetName):
        try:
            self.Worksheet = self.Workbook[SheetName]
        except:
            print(traceback.format_exc())

    def ExcelDataRecv(self, rowIndex):
        try:
            returnData = self.Worksheet[rowIndex].value
            return returnData        
        except:
            print(traceback.format_exc())

    def ExcelDataWrite(self, rowIndex, WriteData):
        try:
            self.Worksheet[rowIndex] = WriteData
        except:
            print(traceback.format_exc())

    def ExcelAllValueLoad(self):
        for index_Y, row in enumerate(self.Worksheet.values):
            for index_X, value in enumerate(row):
                print(f'{index_X}, {index_Y} : {value}')

    def ExcelFileWrite(self):
        try:
            self.xlsx_file_name = f'{self.lineName} 라인 {self.lineType} 전수검사기 검사 실적({self.Today}).xlsx'
            self.five_days_ago_file_name = f'{self.lineName} 라인 {self.lineType} 전수검사기 검사 실적({self.five_days_ago}).xlsx'
            self.Workbook.save(filename = self.xlsx_file_name)
            self.writeExcelFilename = self.xlsx_file_name
        except:
            print(traceback.format_exc())

class DB_DataRequest():
    def __init__(self):
        import datetime
        self.host = 'localhost' #"localhost"
        self.user = 'root' #"root"
        self.password = 'yous7051' #"yous7051"
        self.dbname = 'bearingart' #"bearingart"
        self.CriticalIndex = [1, 2, 3, 4, 5] #[8,9,10,11,12]
        self.badTypeName = {
                            0: "데이터 부족", 
                            1: "S/F 누락", 
                            2: "롤러 누락", 
                            3: "각인 누락", 
                            4: "식별각인 누락",
                            5: '케이지 깨짐',
                            6: "이종 혼입",
                            7: "소단면 불량", 
                            8: "대단면 불량", 
                            9: "소단 내경 불량", 
                            10: "소단 외경 불량", 
                            11: "대단 외경 불량", 
                            12: "대단 내경 불량", 
                            13: "외경 불량",
                            }

    def readSql(self, StartDate=None, EndDate=None,):
        db = pymysql.connect(host=self.host, user=self.user, password=self.password, db=self.dbname, charset="utf8")
        curs = db.cursor()
        sql = "select * from data"
        addvalue = False

        if StartDate is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "InputTime >= '{} 08:00:00'".format(StartDate)

        if EndDate is not None:
            add = " WHERE " if addvalue is False else " AND "
            addvalue = True

            sql = sql + add + "InputTime <= '{} 08:00:00'".format(EndDate)

        sql = sql# + " order by LotNum DESC LIMIT 0,10000"

        print(sql)

        curs.execute(sql)
        rows = curs.fetchall()

        # self.writeExcel(rows, text, dirName=dirs)
        db.close()

        # print(rows)

        return rows

    def DB_DataLoad_Refine(self, TotalCount = None):
        DataDict = {}

        print(PXM.Today, PXM.Yesterday)
        
        returnData = DB.readSql(PXM.Yesterday, PXM.Today)

        # print(type(returnData))

        for i in range(len(returnData)):
            # print('----------------------------------------------------------------------')
            checkValue = [0,0]  #일반불량, 중대불량 체크
            keyCheck = ''
            UpdateValue = [None, None, None]
            for x in range(len(returnData[i])):
                # print(x, returnData[i][x])
                if x == 1:  #형번
                    if returnData[i][x] in DataDict:
                        keyCheck = returnData[i][x]
                        UpdateValue[0] = returnData[i][x]
                        pass
                    else:
                        # Productionmodel.append(returnData[i][x])
                        DataDict[returnData[i][x]] = {'TotalCount' : 0, 'OKCount' : 0, 'NGCount' : 0, 'CriticalNGCount' : 0, 'CriticalNGPartCount' : [0,0,0,0], 'Product' : []}
                        keyCheck = returnData[i][x]
                        UpdateValue[0] = returnData[i][x]
                if 2 <= x <= 6: #불량 유형
                    if '1' in returnData[i][x]:
                        Index = returnData[i][x].index('1')
                        # print(f'{Index} - 불량 체크 ({self.badTypeName[Index]})')
                        if Index in self.CriticalIndex:
                            checkValue[1] += 1
                            if self.badTypeName[Index] == 'S/F 누락':
                                DataDict[keyCheck]['CriticalNGPartCount'][0] += 1
                            elif self.badTypeName[Index] == '롤러 누락':
                                DataDict[keyCheck]['CriticalNGPartCount'][1] += 1
                            elif self.badTypeName[Index] == '각인 누락' or self.badTypeName[Index] == '식별각인 누락':
                                DataDict[keyCheck]['CriticalNGPartCount'][2] += 1
                            elif self.badTypeName[Index] == '케이지 깨짐':
                                DataDict[keyCheck]['CriticalNGPartCount'][3] += 1

                        else:
                            checkValue[0] += 1
                        if UpdateValue[2] == None:
                            UpdateValue[2] = self.badTypeName[Index]
                        else:
                            UpdateValue[2] += f', {self.badTypeName[Index]}'
                if x == 12: #업데이트 시간
                    UpdateValue[1] = returnData[i][x]

            if checkValue[1] != 0:
                DataDict[keyCheck]['CriticalNGCount'] += 1
            elif checkValue[0] != 0:
                DataDict[keyCheck]['NGCount'] += 1
            else:
                DataDict[keyCheck]['OKCount'] += 1
            DataDict[keyCheck]['TotalCount'] += 1
            DataDict[keyCheck]['Product'].append(UpdateValue)

        # print(DataDict)

        return DataDict

class EmailProcessClass():
    def __init__(self):
        import datetime
        self.SendingListUp()
        pass

    def SendingListUp(self):
        with open("Mail_Sending_List.json", "r", encoding='UTF8') as st_json:
            SendingList = json.load(st_json)
        self.Mymail = SendingList["MY"][0]
        self.Tomail = SendingList["MAIN"][0]
        joined_str = ""
        for v in SendingList["SUB"]:
            # 문자열 변수에 값이 있을 경우
            if(len(joined_str) > 0):
                # 콤마를 추가한다
                joined_str += ","
            joined_str += str(v)
        self.CCmail = joined_str

        #db load data
        DB.host = SendingList["DBHOST"]
        DB.user = SendingList["DBUSER"]
        DB.password = SendingList["DBPASSWORD"]
        DB.dbname = SendingList["DBNAME"]
        DB.CriticalIndex = SendingList["CRITICALINDEX"]
        DB.badTypeName = SendingList["BADTYPENAME"]
        DB.badTypeName = {int(k):str(v) for k,v in DB.badTypeName.items()}

        # openpyxl module load data
        PXM.lineName = SendingList["LINENAME"]
        PXM.lineType = SendingList["LINETYPE"]

        print(self.Mymail, self.Tomail, self.CCmail)
        print(DB.host,DB.user,DB.password,DB.dbname,DB.CriticalIndex,DB.badTypeName,DB.badTypeName)
        print(PXM.lineName,PXM.lineType)

    def mailSendProcess(self):
        import smtplib
        from email.mime.multipart import MIMEMultipart # 메일의 Data 영역의 메시지를 만드는 모듈 
        from email.mime.text import MIMEText # 메일의 본문 내용을 만드는 모듈
        from email.mime.application import MIMEApplication # 메일의 첨부 파일을 base64 형식으로 변환
        from email.mime.image import MIMEImage # 메일의 이미지 파일을 base64 형식으로 변환

        # smpt 서버와 연결
        try : 
            gmail_smtp = "smtp.gmail.com"#"smtp.mailplug.co.kr"  #gmail smtp 주소
            gmail_port = 465  #gmail smtp 포트번호
            smpt = smtplib.SMTP_SSL(gmail_smtp, gmail_port)
            print(smpt)

            # 로그인
            my_id = self.Mymail #"khaki798@sinplat.com"
            my_password = "iwppyeqlvjumzaoy"
            smpt.login(my_id, my_password)

            # 메일 기본 정보 설정
            msg = MIMEMultipart()
            msg["Subject"] = PXM.writeExcelFilename#f"{PXM.lineName} {PXM.lineType} line Days Report"
            msg["From"] = self.Mymail#"khaki798@sinplat.com"  #송신자
            msg["To"] = self.Tomail  #수신자
            msg['Cc'] = self.CCmail

            # 메일 내용 쓰기
            content = f""
            content_part = MIMEText(content, "plain")
            msg.attach(content_part)

            # 데이터 파일 첨부하기
            file_name = PXM.writeExcelFilename#"write.xlsx"
            with open(file_name, 'rb') as excel_file : 
                attachment = MIMEApplication( excel_file.read() )
                #첨부파일의 정보를 헤더로 추가
                attachment.add_header('Content-Disposition','attachment', filename=file_name) 
                msg.attach(attachment)
            
            sender = msg['From']
            receiver = msg['To'].split(",")
            if msg['Cc'] is not None:
                receiver += msg['Cc'].split(",")
            if msg['Bcc'] is not None:
                receiver += msg['Bcc'].split(",")

            # 메일 보내고 서버 끄기
            smpt.sendmail(sender, receiver, msg.as_string())  
            smpt.quit()
            return True

        except :
            print(f"Warning : 메일 송부 에러 - {traceback.format_exc()}")
            logger.info(f"Warning : 메일 송부 에러 - {traceback.format_exc()}")
            return False


# ★ 불량 확인 클래스
class NgcheckFrame(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        # 배경 이미지
        self.bg = ImageTk.PhotoImage(file=f"bg/ngcheck_bg.png")

        self.part_txt = [None] * 2
        self.part_txt[0] = ImageTk.PhotoImage(file=f"bg/part_txt0.png")
        self.part_txt[1] = ImageTk.PhotoImage(file=f"bg/part_txt1.png")

        self.choose_part_txt = [None] * 2
        self.choose_part_txt[0] = ImageTk.PhotoImage(file=f"bg/choose_part0.png")
        self.choose_part_txt[1] = ImageTk.PhotoImage(file=f"bg/choose_part1.png")

        self.ok_image = ''
        self.ng_image = ''
        
        self.cam_num = ''  # 101 ~ 105, 111 ~ 116
        self.part = ''

        # 카메라별 검사 파트 정보
        image_files = {
            "101": { # ●
                "filepath": "bg/part/101_.jpg",
                "parts": [
                    {"name": "1", "x_range": (585, 898), "y_range": (132, 477)},
                    {"name": "2", "x_range": (1019, 1341), "y_range": (573, 923)}
                ]
            },
            "102": { # ●
                "filepath": "bg/part/102_.jpg",
                "parts": [
                    {"name": "1", "x_range": (841, 1069), "y_range": (219, 471)}
                ]
            },
            "103": { # ●
                "filepath": "bg/part/103_.jpg",
                "parts": [
                    {"name": "1", "x_range": (855, 1076), "y_range": (404, 663)}
                ]
            },
            "104": { # ●
                "filepath": "bg/part/104_.jpg",
                "parts": [
                    {"name": "0", "x_range": (1182, 1430), "y_range": (70, 168)},
                    {"name": "1", "x_range": (594, 867), "y_range": (143, 448)},
                    {"name": "2", "x_range": (1055, 1331), "y_range": (603, 913)},
                    {"name": "3", "x_range": (1163, 1260), "y_range": (221, 349)}
                ]
            },
            "105": { # ●
                "filepath": "bg/part/105_.jpg",
                "parts": [
                    {"name": "1", "x_range": (832, 1099), "y_range": (602, 898)}
                ]
            },

            "111": { # ●
                "filepath": "bg/part/111_.jpg",
                "parts": [
                    {"name": "0", "x_range": (1186, 1446), "y_range": (63, 164)},
                    {"name": "1", "x_range": (988, 1302), "y_range": (158, 514)},
                    {"name": "2", "x_range": (618, 931), "y_range": (534, 883)},
                    {"name": "3", "x_range": (1030, 1343), "y_range": (566, 914)},
                    {"name": "4", "x_range": (688, 798), "y_range": (223, 368)}
                ]
            },
            "112": { # ●
                "filepath": "bg/part/112_.jpg",
                "parts": [
                    {"name": "1", "x_range": (846, 1066), "y_range": (154, 422)}
                ]
            },
            "113": { # ●
                "filepath": "bg/part/113_.jpg",
                "parts": [
                    {"name": "1", "x_range": (631, 1314), "y_range": (400, 864)}
                ]
            },
            "114": { # ●
                "filepath": "bg/part/114_.jpg",
                "parts": [
                    {"name": "1", "x_range": (1001, 1271), "y_range": (198, 505)},
                    {"name": "2", "x_range": (656, 923), "y_range": (552, 853)},
                    {"name": "3", "x_range": (1035, 1293), "y_range": (570, 866)},
                    {"name": "4", "x_range": (715, 869), "y_range": (247, 441)}
                ]
            },
            "115": { # ●
                "filepath": "bg/part/115_.jpg",
                "parts": [
                    {"name": "1", "x_range": (745, 1198), "y_range": (615, 896)},
                    {"name": "2", "x_range": (600, 1340), "y_range": (204, 608)}
                ]
            },
            "116": { # ●
                "filepath": "bg/part/116_.jpg",
                "parts": [
                    {"name": "1", "x_range": (859, 1081), "y_range": (194, 462)}
                ]
            }
        }

        self.images = {}

        for key, data in image_files.items():
            self.images[key] = {
                "image": ImageTk.PhotoImage(file=data["filepath"]),
                "parts": data["parts"]
            }

        # 선택된 카메라, 파트에 대한 이미지 정보
        self.image_data = {
            'OK': {'filepaths': [], 'current_index': 0},
            'NG': {'filepaths': [], 'current_index': 0}
        }

        self.client_txt_mapping = {
            '111': '소단면 & 각인',
            '112': '내경',
            '113': '롤러 & 케이지',
            '114': '대단면 & 롤러단면',
            '115': '대단면 & 케이지 & 면취',
            '116': '내경',
            '101': '소단면',
            '102': '궤도',
            '103': '외경',
            '104': '대단면 & 각인',
            '105': '면취 & 외경'
        }

        self.label_txt_mapping = {
            'CRACK': '케이지 깨짐',
            'ROLLER': '롤러 누락',
            'SHORT_NG': '롤러 길이 미달',
            'NG': '일반 불량',
            '1_MISS': 'SF누락',
            'MIX': '이종 혼입',
            'R_NG': '롤러 단면 이상',
            'D_NG': '대단 갈림',
            'ED' : '각인 불량'
        }

        self.part_click = False

        self.master.attributes("-fullscreen", True)
        self.master.bind(
            "<F11>", lambda event: self.master.attributes("-fullscreen", not self.master.attributes("-fullscreen")),
        )
        self.master.bind("<Escape>", lambda event: self.master.attributes("-fullscreen", False))
        self.create_widgets()

    def create_widgets(self):
        self.grid(row=0, column=0)
        self.ng_canvas = tk.Canvas(self, width=1920, height=1080)
        self.BG = self.ng_canvas.create_image(0, 0, image=self.bg, anchor="nw")

        self.ok_place = self.ng_canvas.create_image(86, 160, anchor="nw", state = 'normal')
        self.ng_place = self.ng_canvas.create_image(1035, 160, anchor="nw", state = 'normal') 

        # 팝업창 1
        self.popup_place1 = self.ng_canvas.create_image(21, 14, anchor="nw", state="hidden")
        # 팝업창 2
        self.popup_place2 = self.ng_canvas.create_image(50, 300, anchor="nw", state="hidden")

        self.part_image_place = self.ng_canvas.create_image(960, 540, anchor="center", state = 'hidden') 

        self.part_name = self.ng_canvas.create_text(1286, 64, text='', font=("gothic", 15, "bold"), fill="white", anchor="center")
        self.ng_name = self.ng_canvas.create_text(1733, 64, text='', font=("gothic", 15, "bold"), fill="white", anchor="center")

        self.ng_canvas.bind("<Button-1>", self.main_btn)
        self.ng_canvas.pack()

    def main_btn(self, event):
        x = event.x
        y = event.y

        if 1685 < x < 1685 + 194 and 958 < y < 958 + 85:
            self.ng_canvas.itemconfig(self.part_image_place, state = 'hidden')
            main_frame.tkraise()

        # 검사 파트 선택 클릭
        if 30 < x < 30 + 240 and 23 < y < 23 + 78:
            self.part_click = True
            if self.part_click == True :
                print('파트 선택 : ', self.cam_num)
                self.show_part_image()
            
        if self.part_click == False :
            if 29 < x < 29 + 59 and 512 < y < 512 + 56:
                print('양품 이미지 이전')
                self.show_previous_image('OK')
            if 880 < x < 880 + 59 and 512 < y < 512 + 56:
                print('양품 이미지 다음')
                self.show_next_image('OK')

            if 978 < x < 978 + 59 and 512 < y < 512 + 56:
                print('불량 이미지 이전')
                self.show_previous_image('NG')
            if 1828 < x < 1828 + 59 and 512 < y < 512 + 56:
                print('불량 이미지 다음')
                self.show_next_image('NG')

        if self.part_click == True :
            for part in self.images[self.cam_num]["parts"]:
                x_min, x_max = part["x_range"]
                y_min, y_max = part["y_range"]
                if x_min < x < x_max and y_min < y < y_max:
                    print('PART 선택 : ', part["name"])  # 해당 파트의 이름 출력
                    self.part = part["name"]
                    self.image_data = {
                        'OK': {'filepaths': [], 'current_index': 0},
                        'NG': {'filepaths': [], 'current_index': 0}
                    }
                    NF.show_result_image('OK', self.cam_num)
                    NF.show_result_image('NG', self.cam_num)

                    self.ng_canvas.itemconfig(self.part_image_place, state = 'hidden')
                    self.ng_canvas.itemconfig(self.part_name, state = 'normal')
                    self.ng_canvas.itemconfig(self.ng_name, state = 'normal')
                    self.part_click = False
                    break

    # 파트 이미지 출력
    def show_part_image(self):
        self.ng_canvas.itemconfig(self.part_name, state = 'hidden')
        self.ng_canvas.itemconfig(self.ng_name, state = 'hidden')
        self.ng_canvas.itemconfig(self.ok_place, image = '', state = 'hidden')
        self.ng_canvas.itemconfig(self.ng_place, image = '', state = 'hidden')
        self.ng_canvas.itemconfig(self.part_image_place, image = self.images[self.cam_num]['image'], state = 'normal')

    # 클라이언트에서 받은 이미지 저장
    def save_new_image_and_remove_oldest(self, result, img, client_num, part_num, ngname):  # result, decimg, ok_client_num, ok_part_num, okname
        save_path = f'{result}/{client_num}/{part_num}'
        if not os.path.exists(save_path):
            os.makedirs(save_path)

        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        img_name = f"{timestamp}_{ngname}.jpg"

        # List all files in the directory
        all_files = [os.path.join(save_path, f) for f in os.listdir(save_path) if os.path.isfile(os.path.join(save_path, f))]
        
        # If more than 100 images, delete the oldest
        if len(all_files) > 100:
            all_files.sort(key=os.path.getctime)  # Sort by creation time
            os.remove(all_files[0])  # Remove the oldest file
        
        # Save the new image
        cv2.imwrite(os.path.join(save_path, img_name), img)

    # 이미지 출력
    def show_result_image(self, result, client_num):
        # 이미지 경로 불러오기
        self.image_data[result]['filepaths'] = self.get_images_from_path(result, client_num)

        # 확인 영역 정보 출력
        txt = self.client_txt_mapping.get(client_num, '')
        self.ng_canvas.itemconfig(self.part_name, text=f'{txt} 파트 {self.part}')

        self.show_image_by_index(result)
        
    def show_image_by_index(self, result):
        try : 
            img = Image.open(self.image_data[result]['filepaths'][self.image_data[result]['current_index']])
            img_resized = img.resize((794, 794))

            if result == 'OK':
                self.ok_image = ImageTk.PhotoImage(img_resized)
                self.ng_canvas.itemconfig(self.ok_place, image=self.ok_image, state='normal')
            else:
                self.ng_image = ImageTk.PhotoImage(img_resized)
                self.ng_canvas.itemconfig(self.ng_place, image=self.ng_image, state='normal')
                label_idx = self.image_data[result]['filepaths'][self.image_data[result]['current_index']].split('\\')[1].split('_', 2)[2][:-4] # 라벨만 추출
                label_name = self.label_txt_mapping.get(label_idx, '일반 불량')
                self.ng_canvas.itemconfig(self.ng_name, text=f'{label_name}')
        except:
            pass

    def change_image_index(self, result, increment=True):
        if increment:
            self.image_data[result]['current_index'] += 1
            if self.image_data[result]['current_index'] >= len(self.image_data[result]['filepaths']):
                self.image_data[result]['current_index'] = len(self.image_data[result]['filepaths']) - 1
        else:
            self.image_data[result]['current_index'] -= 1
            if self.image_data[result]['current_index'] < 0:
                self.image_data[result]['current_index'] = 0

    def show_next_image(self, result):
        self.change_image_index(result, increment=True)
        self.show_image_by_index(result)

    def show_previous_image(self, result):
        self.change_image_index(result, increment=False)
        self.show_image_by_index(result)

    def get_images_from_path(self, result, client_num):
        try : 
            directory_path = f'{result}/{client_num}/{self.part}'
            all_files = [os.path.join(directory_path, f) for f in os.listdir(directory_path) if os.path.isfile(os.path.join(directory_path, f))]
            return [f for f in all_files if f.lower().endswith(('.jpg'))]
        except :
            print('리스트 없음')
            return []

    # 팝업창 gif 효과 
    def twinkle(self, imagelist, cnt, done) :
        if imagelist == self.part_txt :
            if done == True :
                self.ng_canvas.itemconfig(self.popup_place1, state="hidden")
            elif cnt % 2 == 0 :
                self.ng_canvas.itemconfig(self.popup_place1, image=imagelist[0], state="normal")
            elif cnt % 2 == 1 :
                self.ng_canvas.itemconfig(self.popup_place1, image=imagelist[1], state="normal")
        
        elif imagelist == self.choose_part_txt :
            if done == True :
                self.ng_canvas.itemconfig(self.popup_place2, state="hidden")
            elif cnt % 2 == 0 :
                self.ng_canvas.itemconfig(self.popup_place2, image=imagelist[0], state="normal")
            elif cnt % 2 == 1 :
                self.ng_canvas.itemconfig(self.popup_place2, image=imagelist[1], state="normal")

    # 팝업창 스레드
    def popup_thread(self):
        cnt = 0
        while 1:
            try :
                time.sleep(0.5)
                cnt += 1
                if self.part == '' :
                    self.twinkle(self.part_txt, cnt, False)

                    if self.part_click == True :
                        self.twinkle(self.part_txt, cnt, True)
                        self.twinkle(self.choose_part_txt, cnt, False)
                    else : 
                        self.twinkle(self.choose_part_txt, cnt, True)
                else :
                    if self.part_click == True :
                        self.twinkle(self.choose_part_txt, cnt, False)
                    else : 
                        self.twinkle(self.part_txt, cnt, True)
                        self.twinkle(self.choose_part_txt, cnt, True)
            
                # 초기화
                if cnt == 20 :
                    cnt = 0
            except :
                pass
        

############################################################################################################################################


def on_closing():
    result = messagebox.askyesno("Questions", "Do you want to exit the program?")
    if result == True:
        print("프로그램 종료 확인")
        WriteData = [main_frame.total_count, main_frame.ok_count, main_frame.ng_count]
        try:
            try:
                if not (os.path.isdir(f"CountSave")):
                    os.makedirs(os.path.join(f"CountSave"))
            except OSError as e:
                pass

            os.remove("CountSave/Count.pickle")

            with open(f"CountSave/Count.pickle", "wb") as file:
                pickle.dump(WriteData, file)
        except Exception as ex:
            logger.info(f"Warning : 피클파일 저장 에러 - {ex} \n {traceback.format_exc()}")
        CountDataSave()
        # STH.SendDataTrim("W", "D3000", "0")
        # STH.SendDataTrim("W", "D3001", "0")
        # STH.SendDataTrim("W", "D3002", "0")
        # STH.SendDataTrim("W", "D3003", "0")
        # STH.SendDataTrim("W", "D3006", "0")
        # STH.SendDataTrim("W", "D3007", "0")
        STH.ModbusSignalSend("RESET")
        exit()
    else:
        print("프로그램 종료하지 아니함")


def CountDataSave():
    WriteData = [main_frame.total_count, main_frame.ok_count, main_frame.ng_count]
    try:
        try:
            if not (os.path.isdir(f"CountSave")):
                os.makedirs(os.path.join(f"CountSave"))
        except OSError as e:
            pass

        with open(f"CountSave/Count.pickle", "wb") as file:
            pickle.dump(WriteData, file)
    except Exception as ex:
        logger.info(f"Warning : 피클파일 저장 에러 - {ex} \n {traceback.format_exc()}")


def loadCount():
    print("초기 검사 수량 데이터 로드")
    try:
        # self.vision_value = []
        with open(f"CountSave/Count.pickle", "rb") as file:
            data = pickle.load(file)

        loadData = data.copy()
        main_frame.total_count = loadData[0]
        main_frame.ok_count = loadData[1]
        main_frame.ng_count = loadData[2]

        main_frame.main_canvas.itemconfig(main_frame.totalText, text=main_frame.total_count)
        main_frame.main_canvas.itemconfig(main_frame.okText, text=main_frame.ok_count)
        main_frame.main_canvas.itemconfig(main_frame.ngText, text=main_frame.ng_count)

        print("피클 파일 로딩 완료")
    except:
        logger.info(f"피클파일이 존재하지 않습니다. {traceback.format_exc()}")
        print("피클파일이 존재하지 않습니다.")


if __name__ == "__main__":
    print("[INFO] '24.02.26 Update")
    DTT = DataTrimThread()
    DB = Database('bearingart', 'model_data', LINE) # DB명, table명, line명

    main_frame = MainFrame(master=root)
    history_frame = HistoryFrame(master=root)
    history_produce_frame = historyProduceFrame(master=root)
    ngcage_frame = ngCageShowFrame(master=root)
    management_frame = AdminManagementFrame(master=root)
    management_frame.management_State_Load()
    model_frame = ModelManageFrame(master=root, db_inst=DB)
    NF = NgcheckFrame(master=root)
    main_frame.tkraise()

    # 데이터 정리 쓰레드
    DTT.daemon = True
    DTT.start()

    # 서버클라이언트 소켓 쓰레드
    TSD = TransferSocketData()
    TSD.daemon = True
    TSD.start()

    STH = serialTRAN()
    STH.daemon = True
    STH.start()

    bearingartDB = MysqlDB()

    MQ = MQueue()
    proc = Process(target=SaveImages, args=(MQ,))
    proc.daemon = True
    proc.start()

    transfer = Transfer(LINE, CodeSetup)       # 데이터 송, 수신 모듈
    jsonparser = JASONparser(LINE, CodeSetup)  # json 파싱 모듈

    loadCount()

    DB = DB_DataRequest()
    PXM = OpenpyxlModul()
    Email = EmailProcessClass()

    TE = TimeExcel()
    threading.Thread(target=TE.run, daemon=True).start()

    
    Thread(target=main_frame.popup_thread, daemon=True).start()
    Thread(target=NF.popup_thread, daemon=True).start() 
    
    main_frame.ngSettingLoad("load")

    root.protocol("WM_DELETE_WINDOW", on_closing)
    root.mainloop()
